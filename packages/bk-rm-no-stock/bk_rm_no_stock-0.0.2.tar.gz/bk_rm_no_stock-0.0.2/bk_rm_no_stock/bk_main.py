#beki
from openpyxl import load_workbook


def remove_row(file_path):
	wb = load_workbook(file_path, data_only=True)
	ws = wb.active
	
	# Iterate in reverse to prevent shifting issues when deleting rows
	for row in range(160, 0, -1):  
	    cell_j = str(ws[f"J{row}"].value)
	    cell_m = str(ws[f"M{row}"].value)
	    cell_r = str(ws[f"R{row}"].value)
	
	
	
	
	    if cell_j == "0" and cell_m == "0" and cell_r == "0":
	        ws.delete_rows(row)
	
	
	wb.save(file_path)
	wb.close()

