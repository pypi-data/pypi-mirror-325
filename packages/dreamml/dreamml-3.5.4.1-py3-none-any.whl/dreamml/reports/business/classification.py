import os
from typing import Optional, List, Dict, Tuple, Any

import numpy as np
import shap
from matplotlib import pyplot as plt
from openpyxl.drawing.image import Image
from openpyxl.styles import NamedStyle, Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

from dreamml.reports._classification_metrics import (
    CalculateDataStatistics,
    CalculateDetailedMetrics,
)
from dreamml.reports.reports.style_utils import set_row_auto_height
from dreamml.validation._task_description import TaskDescriptionTest
from dreamml.validation.classification import ValidationReport

from dreamml.logging import get_logger
from dreamml.utils.prepare_artifacts_config import ClassificationArtifactsConfig
from dreamml.validation.classification._detailed_model_statistics import (
    ModelDetailedStatistics,
)
from dreamml.visualization.plots import (
    plot_quality_dynamics_per_segment_graph,
    plot_data_decile_statistics_graph,
)

_logger = get_logger(__name__)


def prepare_artifacts_config(config: dict) -> Tuple[Any, Any]:
    prepare_artifacts = ClassificationArtifactsConfig(config=config)
    artifacts_config, eval_set = prepare_artifacts.prepare_artifacts_config()

    artifacts_config["path_to_save"] = os.path.join(
        prepare_artifacts.experiment_dir_path,
        "docs",
        f"business_report_{prepare_artifacts.model_name}.xlsx",
    )

    return artifacts_config, eval_set


class ClassificationBusinessReport(ValidationReport):
    """
    Отчет о финальной модели для отправки
    на валидацию / бизнес заказчику.

    """

    def __init__(
        self,
        config,
        estimator,
        metric_name,
        vectorizer: callable = None,
        path_to_save: str = "./business_report.xlsx",
        used_features: Optional[List[str]] = None,
        categorical_features: Optional[List[str]] = None,
        create_file: bool = True,
        metric_col_name: str = "gini",
        metric_params: dict = None,
        images_dir_path: str = None,
        task: str = "binary",
        subtask: str = "tabular",  # [tabular, nlp]
        multiclass_artifacts: Optional[Dict] = None,
        custom_model: bool = False,
        user_config=None,
        text_column: str = None,
        text_preprocessed_column: str = None,
        group_column: str = None,
        time_column: str = None,
        create_pdf: bool = False,
    ):
        super().__init__(
            config=config,
            estimator=estimator,
            metric_name=metric_name,
            vectorizer=vectorizer,
            path_to_save=path_to_save,
            used_features=used_features,
            categorical_features=categorical_features,
            create_file=create_file,
            metric_col_name=metric_col_name,
            metric_params=metric_params,
            images_dir_path=images_dir_path,
            task=task,
            subtask=subtask,
            multiclass_artifacts=multiclass_artifacts,
            custom_model=custom_model,
            user_config=user_config,
            text_column=text_column,
            text_preprocessed_column=text_preprocessed_column,
            group_column=group_column,
            time_column=time_column,
            create_pdf=create_pdf,
        )
        self._init_styles()

    def create_report(self, **data):
        self._create_business_task_description_page(**data)
        self._create_data_samples_statistics_page(**data)
        self._create_pipeline_description_page(**data)
        self._create_model_params_page(**data)
        self._create_detailed_stats_page(**data)
        self._create_quality_dynamics_page(**data)
        self._create_data_distribution_page(**data)
        self._create_features_importance_page(**data)

        _logger.info("Расчет результатов тестов.")
        self._create_data_quality_page(**data)
        self._create_model_quality_page(**data)
        self._create_model_calibration_page(**data)

        if "OOT" in data:
            self._create_model_stability_page(**data)
        else:
            self._create_model_stability_page_without_oot(**data)

        drop_sheets = [
            "Качество данных (PSI-анализ)",
            "Качество (Точность) модели",
            "Калибровка модели",
            "Стабильность модели",
        ]

        for sheet in drop_sheets:
            del self.wb[sheet]
            del self.sheets[sheet]

        self._count_total_result()  # Заполняем таблицу итоговых результатов
        self._create_total_result_page()

        for sheet_name, ws in self.sheets.items():
            set_row_auto_height(ws)

        self.writer.save()
        self.create_pdf()

    def _init_styles(self):
        cell_format = NamedStyle(name="cell_format")
        cell_format.font = Font(color="808080")
        cell_format.alignment = Alignment(
            wrap_text=True, horizontal="center", vertical="center"
        )
        brd = Side(border_style="thin", color="000000")
        cell_format.border = Border(left=brd, right=brd, top=brd, bottom=brd)

        row_title_format = NamedStyle(name="row_title_format")
        row_title_format.font = Font(bold=True)
        row_title_format.alignment = Alignment(
            wrap_text=True, horizontal="center", vertical="center"
        )
        brd = Side(border_style="thin", color="000000")
        row_title_format.border = Border(left=brd, right=brd, top=brd, bottom=brd)

        col_title_format = NamedStyle(name="col_title_format")
        col_title_format.font = Font(bold=True)
        col_title_format.fill = PatternFill(
            start_color="00CC99", end_color="00CC99", fill_type="solid"
        )
        col_title_format.alignment = Alignment(horizontal="center", vertical="center")
        brd = Side(border_style="thin", color="000000")
        col_title_format.border = Border(left=brd, right=brd, top=brd, bottom=brd)

        self.wb.add_named_style(cell_format)
        self.wb.add_named_style(row_title_format)
        self.wb.add_named_style(col_title_format)

    def _create_business_task_description_page(self, **data):
        """
        Страница с описанием бизнес-задачи
        """
        _logger.info("1. Сбор статистики о решаемой бизнес-задаче.")
        sheet_name = "1. Задача. 2. Выборка"

        ws = self.wb.create_sheet(title=sheet_name)
        self.sheets[sheet_name] = ws
        self.sheet_descriptions[sheet_name] = ""

        business_task_text = (
            "<Заполните название модели>\n"
            "например: Модель оттока клиентов по пакету услуг Премьер"
        )
        task_description_text = (
            "<Заполните описание модели: цель моделирования, "
            "для какого бизнес процесса строится модель, какие данные используются >\n"
            "например: Модель для выделения клиентов банка склонных к закрытию Пакета Услуг Премьер. \n"
            "Модель строится на данных ЦОД по информации об открытии/закрытии пакетов услуг клиентами банка. \n"
            "В качестве факторов используются данные клиентского профиля трайба Массовая Персонализация."
        )
        task_type_text = self.artifacts_config["task"]
        data_selection_date_text = (
            "<Укажите дату сбора данных>\n" "например: 31.01.2024"
        )
        target_description_text = f"<Заполните определние целевого события/переменной>"
        data_selection_text = (
            "<Заполните критерии отбора популяции: фильтры, исключения>\n"
            "например: 1. Клиенты, по которым есть информация в витрине \n"
            "клиентского профиля на отчетную дату."
        )
        used_algo_text = (
            f"<DreamML {self.artifacts_config['estimator'].__class__.__name__}>"
        )

        ws.merge_cells("B2:E2")
        ws.merge_cells("B3:E3")
        ws.merge_cells("B4:E4")
        ws.merge_cells("B5:E5")
        ws.merge_cells("B6:E6")
        ws.merge_cells("B7:E7")
        ws.merge_cells("B8:E8")
        ws.merge_cells("B9:E9")

        ws["B2"] = business_task_text
        ws["B3"] = task_description_text
        ws["B4"] = task_type_text
        ws["B5"] = target_description_text
        ws["B6"] = data_selection_date_text
        ws["B7"] = data_selection_text
        ws["B8"] = used_algo_text

        metrics = {self.metric_name, "gini", "precision_recall_auc"}
        ws["B9"] = ", ".join(metrics)

        ws["A1"] = "1. Описание задачи"
        ws["A1"].font = Font(size=22, bold=True)
        ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[1].height = 40

        ws["A2"] = "Название модели"
        ws["A3"] = "Описание задачи"
        ws["A4"] = "Тип задачи"
        ws["A5"] = "Целевая переменная"
        ws["A6"] = "Дата сбора данных"
        ws["A7"] = "Отбор наблюдений"
        ws["A8"] = "ML-алгоритм"
        ws["A9"] = "Метрики качества"

        for row in ws["B2":"E9"]:
            for cell in row:
                cell.style = "cell_format"
                cell.alignment = Alignment(
                    horizontal="left", vertical="center", wrap_text=True
                )
        for row in ws["A2":"A9"]:
            for cell in row:
                cell.style = "row_title_format"

        ws.column_dimensions["A"].width = 22
        for i in range(2, 6):
            ws.column_dimensions[get_column_letter(i)].width = 18
        # for i in range(1, 10):
        #     ws.row_dimensions[i].height = None

    def _create_data_samples_statistics_page(self, **data):
        _logger.info("2. Сбор статистики о выборке.")
        sheet_name = "1. Задача. 2. Выборка"

        features = data["train"][0].columns.to_series()
        transformer = CalculateDataStatistics(
            None,
            features,
            self.config,
            task=self.artifacts_config["task"],
            business=True,
        )

        result = transformer._calculate_samples_stats(**data)

        result.to_excel(
            self.writer,
            startrow=11,
            sheet_name=sheet_name,
            index=False,
            float_format="%.2f",
        )

        ws = self.sheets[sheet_name]

        ws["A11"] = "2. Выборка"
        ws["A11"].font = Font(size=22, bold=True)
        ws["A11"].alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[11].height = 40

        for row in ws["A13" :f"G{12+len(result)}"]:
            for cell in row:
                cell.style = "cell_format"
                cell.alignment = Alignment(
                    horizontal="left", vertical="center", wrap_text=True
                )
        for row in ws["A12":"G12"]:
            for cell in row:
                cell.style = "row_title_format"

        ws.column_dimensions["A"].width = 22
        for i in range(2, 8):
            ws.column_dimensions[get_column_letter(i)].width = 18

    def _create_pipeline_description_page(self, **data):
        """
        Страница с описанием всех стадий пайплайна
        """
        _logger.info("3. Сбор статистики об используемом пайплайне.")
        sheet_name = "3. Пайплайн"
        self.sheet_descriptions[sheet_name] = (
            "Страница с описанием всех стадий пайплайна"
        )
        pst = TaskDescriptionTest(self.artifacts_config, self.validation_test_config)
        stats = pst._create_description(**data)
        stats.to_excel(self.writer, sheet_name=sheet_name, index=False, startrow=1)

        ws = self.sheets[sheet_name]

        ws["A1"] = "3. Pipeline моделирования"
        ws["A1"].font = Font(size=22, bold=True)
        ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[1].height = 40

        ws.column_dimensions["A"].width = 35
        ws.column_dimensions["B"].width = 130

        for row in ws["A3":"B9"]:
            for cell in row:
                cell.style = "cell_format"
                cell.alignment = Alignment(
                    horizontal="left", vertical="center", wrap_text=True
                )
        for row in ws["A2":"B2"]:
            for cell in row:
                cell.style = "row_title_format"

    def _create_model_params_page(self, **data):
        _logger.info("4. Cбор гиперпарметров модели.")
        sheet_name = "4. Гиперпараметры. 5. Качество"
        self.sheet_descriptions[sheet_name] = ""

        params = self.estimator.get_estimator_params
        params["Настройка"] = ""
        params.index = np.arange(1, len(params) + 1)

        params.to_excel(
            self.writer,
            sheet_name=sheet_name,
            index=True,
            startrow=1,
        )

        ws = self.sheets[sheet_name]

        ws["A1"] = "4. Гиперпараметры модели"
        ws["A1"].font = Font(size=22, bold=True)
        ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[1].height = 40

        ws.column_dimensions["A"].width = 4
        ws.column_dimensions["B"].width = 25
        ws.column_dimensions["C"].width = 12
        ws.column_dimensions["D"].width = 10

        for row in ws["B3" :f"D{2+len(params)}"]:
            for cell in row:
                cell.style = "cell_format"
                cell.alignment = Alignment(
                    horizontal="left", vertical="center", wrap_text=True
                )
        for row in ws["B2":"D2"]:
            for cell in row:
                cell.style = "row_title_format"

        for i in range(3, 3 + len(params)):
            ws.row_dimensions[i].height = 14

    def _create_detailed_stats_page(self, **data):
        _logger.info("5. Детальный расчёт качества модели.")
        sheet_name = "4. Гиперпараметры. 5. Качество"

        stats_test = ModelDetailedStatistics(
            self.artifacts_config,
            self.validation_test_config,
            self.metric_name,
            self.metric_col_name,
            self.metric_params,
            business=True,
        )
        stats = stats_test.create_report(**data)

        ws = self.sheets[sheet_name]

        ws["F1"] = "5. Качество модели (Gini)"
        ws["F1"].font = Font(size=22, bold=True)
        ws["F1"].alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[1].height = 40

        stats.to_excel(
            self.writer,
            sheet_name=sheet_name,
            startrow=1,
            startcol=5,
            index=False,
        )

        for row in ws["F3" :f"{get_column_letter(5+stats.shape[1])}{2+stats.shape[0]}"]:
            for cell in row:
                cell.style = "cell_format"
                cell.alignment = Alignment(
                    horizontal="left", vertical="center", wrap_text=True
                )

        for row in ws[
            "G3" :f"{get_column_letter(5 + stats.shape[1])}{2 + stats.shape[0]}"
        ]:
            for cell in row:
                cell.style = "cell_format"
                cell.alignment = Alignment(
                    horizontal="left", vertical="center", wrap_text=True
                )
                cell.number_format = "0.0%"
        for row in ws["F2" :f"{get_column_letter(5+stats.shape[1])}2"]:
            for cell in row:
                cell.style = "row_title_format"

        picture_name = f"business_model_quality_per_segment.png"
        path_to_picture = os.path.join(self.images_dir_path, picture_name)
        if os.path.exists(path_to_picture):
            img = Image(path_to_picture)
            img.anchor = f"F{2+stats.shape[0]+2}"
            ws.add_image(img)

    def _create_quality_dynamics_page(self, **data):
        sheet_name = "6. Динамика качества"

        time_column = self.artifacts_config["time_column"]
        if time_column is None:
            _logger.warning(
                f'Страница "{sheet_name}" не будет включена в отчет в связи с отсутствием параметра time_column'
            )
            return

        _logger.info("6. Расчёт динамики качества модели.")

        ws = self.wb.create_sheet(title=sheet_name)
        self.sheets[sheet_name] = ws
        self.sheet_descriptions[sheet_name] = ""

        ws["A1"] = "6. Динамика метрики качества модели"
        ws["A1"].font = Font(size=22, bold=True)
        ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[1].height = 40

        group_column = self.artifacts_config["group_column"]

        if group_column is None:
            unique_groups = [""]
        else:
            unique_groups = set()
            for sample_name in data:
                unique_groups.update(data[sample_name][0][group_column].unique())

            unique_groups = sorted(list(unique_groups))

        target_per_group = {}
        time_per_group = {}
        for group in unique_groups:
            target_per_group[group] = {}
            time_per_group[group] = {}

            for sample_name in data:
                x, y_true = data[sample_name]

                if group_column is None:
                    x_group = x
                    y_true_group = y_true
                else:
                    x_group = x[x[group_column] == group]
                    y_true_group = y_true[x[group_column] == group]

                if len(x_group) == 0:
                    target_per_group[group][sample_name] = ([], [])
                    time_per_group[group][sample_name] = []
                    continue

                time = x_group[time_column]

                y_pred_group = self.estimator.transform(x_group)

                target_per_group[group][sample_name] = (y_true_group, y_pred_group)
                time_per_group[group][sample_name] = time

        picture_name = f"business_quality_dynamics_segment.png"

        save_path = os.path.join(self.images_dir_path, picture_name)
        plot_quality_dynamics_per_segment_graph(
            time_per_group, target_per_group, save_path=save_path
        )

        path_to_picture = os.path.join(self.images_dir_path, picture_name)
        if os.path.exists(path_to_picture):
            img = Image(path_to_picture)
            img.anchor = f"A2"
            ws.add_image(img)

    def _create_data_distribution_page(self, **data):
        _logger.info("7. Расчёт распределения данных.")
        sheet_name = "7. Распределение"

        ws = self.wb.create_sheet(title=sheet_name)
        self.sheets[sheet_name] = ws
        self.sheet_descriptions[sheet_name] = ""

        ws["A1"] = "7. Распределение по децилям"
        ws["A1"].font = Font(size=22, bold=True)
        ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[1].height = 40

        group_column = self.artifacts_config["group_column"]

        if group_column is None:
            unique_groups = [""]
        else:
            unique_groups = set()
            for sample_name in data:
                unique_groups.update(data[sample_name][0][group_column].unique())

            unique_groups = sorted(list(unique_groups))

        n_bins = 10

        decile_index = [
            f"[{int(100 * i/n_bins)}-{int(100 * (i+1)/n_bins)}]" for i in range(n_bins)
        ]

        x, y_true = data["test"]
        start_row = 1
        for idx, group in enumerate(unique_groups):
            ws[f"A{start_row+1}"] = f"Segment {group}"
            ws[f"A{start_row+1}"].font = Font(size=14, bold=True)
            ws[f"A{start_row+1}"].alignment = Alignment(
                horizontal="left", vertical="bottom"
            )
            ws.row_dimensions[start_row + 1].height = 30

            start_row += 1

            if group_column is None:
                x_group = x
                y_true_group = y_true
            else:
                x_group = x[x[group_column] == group]
                y_true_group = y_true[x[group_column] == group]

            if len(x_group) == 0:
                continue

            y_pred_group = self.estimator.transform(x_group)

            transformer = CalculateDetailedMetrics(
                n_bins, "gini", {}, task=self.artifacts_config["task"]
            )
            try:
                data = transformer.transform(y_true_group, y_pred_group)
            except ValueError as e:
                _logger.warning(
                    f"Can't calculate data distribution for group {group}: {e}"
                )
                continue

            data = data.iloc[:n_bins]

            data = data[["cum # obs", "cum_eventrate"]]
            data = data.reset_index()

            data["index"] = decile_index
            data.columns = [
                "Дециль (модельные прогнозы)",
                "Кол-во наблюдений",
                "Event-rate (факт.)",
            ]

            data.to_excel(
                self.writer, sheet_name=sheet_name, startrow=start_row, index=False
            )
            for row in ws[
                f"A{start_row+2}" :f"{get_column_letter(data.shape[1])}{start_row + 1 + data.shape[0]}"
            ]:
                for cell in row:
                    cell.style = "cell_format"
            for row in ws[
                f"A{start_row+1}" :f"{get_column_letter(data.shape[1])}{start_row+1}"
            ]:
                for cell in row:
                    cell.style = "row_title_format"
            for row in ws[f"A{start_row+2}" :f"A{start_row + 1 + data.shape[0]}"]:
                for cell in row:
                    cell.style = "row_title_format"

            picture_name = f"data_decile_statistics_segment_{idx}.png"
            path_to_picture = os.path.join(self.images_dir_path, picture_name)
            plot_data_decile_statistics_graph(data, save_path=path_to_picture)

            if os.path.exists(path_to_picture):
                img = Image(path_to_picture)
                img.anchor = f"{get_column_letter(data.shape[1]+1)}{start_row+1}"
                ws.add_image(img)

            start_row += len(data) + 1

    def _create_features_importance_page(self, **data):
        _logger.info("8. Оценка важности признаков модели по метрике SHAP-values.")
        sheet_name = "8. Признаки"

        if "valid" in data:
            x, _ = data["valid"]
        elif "train" in data:
            x, _ = data["train"]
        elif "test" in data:
            x, _ = data["test"]
        elif "OOT" in data:
            x, _ = data["OOT"]

        shap_values, imp = self.estimator.get_shap_importance(x)
        imp.to_excel(self.writer, sheet_name=sheet_name, startrow=1, index=True)
        ws = self.sheets[sheet_name]

        ws["A1"] = "8. Важность модельных признаков"
        ws["A1"].font = Font(size=22, bold=True)
        ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[1].height = 40

        for row in ws["B3" :f"C{2+len(imp)}"]:
            for cell in row:
                cell.style = "cell_format"
                cell.alignment = Alignment(
                    horizontal="left", vertical="center", wrap_text=True
                )
        for row in ws["A2":"C2"]:
            for cell in row:
                cell.style = "row_title_format"
        for row in ws["A3" :f"A{2+len(imp)}"]:
            for cell in row:
                cell.style = "row_title_format"

        for i in range(3, 3 + len(imp)):
            ws.row_dimensions[i].height = 20

        ws.column_dimensions["A"].width = 4
        ws.column_dimensions["B"].width = 25
        ws.column_dimensions["C"].width = 12

        if isinstance(shap_values, np.ndarray):
            plt.clf()
            shap.initjs()
            shap.summary_plot(
                shap_values,
                features=x[self.estimator.used_features],
                feature_names=self.estimator.used_features,
                show=False,
            )
            plt.savefig(
                os.path.join(self.images_dir_path, "shap_summary_plot.png"),
                bbox_inches="tight",
                pad_inches=0.1,
            )
            plt.close()

            img = Image(os.path.join(self.images_dir_path, "shap_summary_plot.png"))
            img.anchor = f"D2"
            ws.add_image(img)

        msg = "Лист с оценкой важности признаков модели по метрике SHAP-values"
        self.sheet_descriptions[sheet_name] = msg

    def _create_total_result_page(self):
        """
        Страница с результатами по все тестам.
        """
        _logger.info("9. Подведение итогов тестов.")
        sheet_name = "9. Валидация DML"

        self.total_result_df.to_excel(
            self.writer,
            sheet_name=sheet_name,
            startrow=0,
            startcol=0,
            index=False,
        )

        self.add_table_borders(
            self.total_result_df[
                [
                    "Блок тестов",
                    "N",
                    "Краткое описание теста",
                    "Детали теста",
                    "Использование OOS",
                    "Использование OOT",
                    "Итог",
                ]
            ],
            sheet_name=sheet_name,
            startcol=0,
        )
        self.add_header_color(
            self.total_result_df,
            sheet_name=sheet_name,
            color="00CC99",
            startcol=0,
        )
        self.add_traffic_light_color(sheet_name=sheet_name)

        ws = self.sheets[sheet_name]

        # Настройка ширины столбцов
        ws.column_dimensions[get_column_letter(1)].width = 33
        ws.column_dimensions[get_column_letter(2)].width = 5
        ws.column_dimensions[get_column_letter(3)].width = 110
        ws.column_dimensions[get_column_letter(4)].width = 10
        ws.column_dimensions[get_column_letter(5)].width = 15
        ws.column_dimensions[get_column_letter(6)].width = 15
        ws.column_dimensions[get_column_letter(7)].width = 10
        ws.column_dimensions[get_column_letter(8)].width = 10
        ws.column_dimensions[get_column_letter(9)].width = 15

        # Выравнивание текста в ячейках
        for row in ws[2 : 2 + self.total_result_df.shape[0]]:
            for col in [1, 3, 4, 5, 6]:
                cell = row[col]
                cell.alignment = Alignment(horizontal="center")
        for row in ws[2 : 2 + self.total_result_df.shape[0]]:
            for col in [0]:
                cell = row[col]
                cell.alignment = Alignment(horizontal="center", vertical="center")

        # Границы ячеек
        for row in [ws[5:5], ws[11:11], ws[14:14], ws[17:17], ws[25:25], ws[27:27]]:
            for col in range(7):
                cell = row[col]
                cell.border = Border(
                    left=cell.border.left,
                    right=cell.border.right,
                    top=cell.border.top,
                    bottom=Side(border_style="thin", color="000000"),
                )

        # Объединение ячеек
        ws.merge_cells("A2:A5")
        ws.merge_cells("A6:A11")
        ws.merge_cells("A12:A14")
        ws.merge_cells("A15:A17")
        ws.merge_cells("A18:A25")