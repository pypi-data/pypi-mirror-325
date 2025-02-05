import pandas as pd
from openpyxl import load_workbook
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.pdfgen import canvas
from reportlab.platypus import Table, TableStyle
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4

from dreamml.logging import get_logger


font_scale = 0.6
try:
    pdfmetrics.registerFont(TTFont("Arial", "Arial.ttf"))
    pdfmetrics.registerFont(TTFont("Arial-Bold", "Arial_Bold.ttf"))
    pdf_maker_available = True
except Exception as e:
    pdf_maker_available = False

_logger = get_logger(__name__)


def read_sheet_with_styles(wb, sheet_name=0):
    # Load the workbook and select the worksheet
    ws = wb[sheet_name if isinstance(sheet_name, str) else wb.sheetnames[sheet_name]]

    # Create a Pandas DataFrame from the sheet
    #     df = pd.DataFrame(ws.values)

    from itertools import islice

    data = ws.values

    first_row = next(data)
    cols = first_row[1:]
    index_name = first_row[0]

    data = list(data)
    idx = [r[0] for r in data]
    data = (islice(r, 1, None) for r in data)
    df = pd.DataFrame(data, index=idx, columns=cols)
    df.index.name = index_name

    # Extract styles for each cell
    styles = {}
    for row in ws.iter_rows():
        for cell in row:
            styles[(cell.row - 1, cell.column - 1)] = {
                "font": cell.font,
                "fill": cell.fill,
                "border": cell.border,
            }

    merged_bounds = [r.bounds for r in ws.merged_cells.ranges]

    return df, styles, merged_bounds


def style_table_on_canvas(c, df, styles, merged_bounds, pagesize):
    width, height = pagesize

    # Create a table with the DataFrame content
    df = df.reset_index()
    data = [df.columns.to_list()] + df.values.tolist()
    table = Table(data, colWidths=[(width - 100) / len(df.columns)] * len(df.columns))

    # Create a default style
    default_style = TableStyle(
        [
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("FONTNAME", (0, 0), (-1, -1), "Arial"),
            ("FONTSIZE", (0, 0), (-1, -1), 11 * font_scale),
            #         ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ]
    )

    # Apply the extracted styles to the table
    for (row, col), style in styles.items():
        if (
            style["fill"].fgColor.rgb is not None
            and style["fill"].patternType is not None
        ):
            if isinstance(style["fill"].bgColor.rgb, str):
                color = f"#{style['fill'].fgColor.rgb[-6:]}"
                default_style.add("BACKGROUND", (col, row), (col, row), color)

        #         fontname = style['font'].name
        #         fontname = "Times-Bold" if style['font'].bold else "Times-Roman"
        if style["font"].bold:
            default_style.add("FONTNAME", (col, row), (col, row), "Arial-Bold")

        if style["font"].size:
            default_style.add(
                "FONTSIZE", (col, row), (col, row), style["font"].size * font_scale
            )

        if style["border"].bottom.style is not None:
            default_style.add("LINEBELOW", (col, row), (col, row), 1, colors.black)
        if style["border"].top.style is not None:
            default_style.add("LINEABOVE", (col, row), (col, row), 1, colors.black)
        if style["border"].left.style is not None:
            default_style.add("LINEBEFORE", (col, row), (col, row), 1, colors.black)
        if style["border"].right.style is not None:
            default_style.add("LINEAFTER", (col, row), (col, row), 1, colors.black)

    for bounds in merged_bounds:
        default_style.add(
            "SPAN", (bounds[0] - 1, bounds[1] - 1), (bounds[2] - 1, bounds[3] - 1)
        )

    table.setStyle(default_style)

    # Draw the table on the canvas
    table.wrapOn(c, width, height)
    table.drawOn(c, 20, 100)


def convert_excel_to_pdf(excel_file, pdf_file, pagesize=A4):
    _logger.debug(f"Converting {excel_file} to {pdf_file}...")

    wb = load_workbook(excel_file, data_only=True)

    # Create a PDF canvas
    c = canvas.Canvas(pdf_file, pagesize=pagesize)
    width, height = pagesize

    for sheet_name in wb.sheetnames:
        df, styles, merged_bounds = read_sheet_with_styles(wb, sheet_name)

        style_table_on_canvas(c, df, styles, merged_bounds, pagesize)

        c.showPage()

    # Save the PDF
    c.save()