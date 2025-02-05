import sys
import os
import json
import time
import uuid
from pathlib import Path
import yaml
from copy import deepcopy
from typing import Optional, List, Dict, Tuple, Any

import shap
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
from dreamml.validation._base import ValidationStyles

import dreamml as dml
from dreamml.modeling.metrics import metrics_mapping

from dreamml.validation.classification._gini_test import GiniTest
from dreamml.validation.classification._ks_test import KSTest
from dreamml.validation.classification._f_test import FTest
from dreamml.validation.classification._ciec_test import CIECTest
from dreamml.validation.classification._data_statistics import DataStatisticsTest
from dreamml.validation.classification._permutation_importance import PermutationImportanceChecker
from dreamml.validation._task_description import TaskDescriptionTest
from dreamml.validation._task_description import BusinessCaseDescription

from dreamml.validation.wrappers._estimator import Estimator

from dreamml.configs import traffic_lights
from openpyxl.styles import NamedStyle, Font, PatternFill, Border, Side, Alignment
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
from dreamml.utils.get_last_experiment_directory import get_experiment_dir_path
from dreamml.logging import get_logger
from dreamml.logging.monitoring import ReportStartedLogData, ReportFinishedLogData
from dreamml.utils.prepare_artifacts_config import ClassificationArtifactsConfig

_logger = get_logger(__name__)


def prepare_validation_test_config(config, raise_exception: bool = True):
    experiment_dir_path = get_experiment_dir_path(
        config.get("results_path", ""),
        experiment_dir_name=config.get("dir_name"),
        use_last_experiment_directory=config.get(
            "use_last_experiment_directory", False
        ),
        raise_exception=raise_exception,
    )
    try:
        with open(f"{experiment_dir_path}/config/config.yaml", "r") as f:
            experiment_config = yaml.unsafe_load(f)
    except FileNotFoundError as e:
        experiment_config = {}
        _logger.exception(e)
        _logger.info("validation method by default: hold-out")
    # если кастомная модель (нет config.yaml, то светофоры по дефолту эти)
    # TODO: при получении флагов от валидаторов убрать светофоры из конфига и захардкодить сюда
    return {
        "n_folds": config.get("n_folds", 4),
        "psi_threshold": config.get("psi_threshold", 0.2),
        "bootstrap_samples": config.get("bootstrap_samples", 50),
        "fairness_threshold": config.get("fairness_threshold", 0.5),
        "fairness_features": config.get("fairness_features", []),
        "validation_method_for_model_training": experiment_config.get(
            "validation", "hold-out"
        ),
        "traffic_lights": experiment_config.get("traffic_lights", traffic_lights),
    }

def prepare_artifacts_config(config: dict) -> Tuple[Any, Any]:
    """
    Подготовка конфигурационного файла для передачи в
    объект валидационного отчета ValidationReport.

    Parameters
    ----------
    config: dict
        Конфигурационный файл config с указанием
        названия финальной модели и директории с
        артефактами вычислительного эксперимента.

    Returns
    -------
    artifacts_config: dict
        Конфигурационный файл для отчета.

    """
    prepare_artifacts = ClassificationArtifactsConfig(config=config)
    artifacts_config, eval_set = prepare_artifacts.prepare_artifacts_config()
    return artifacts_config, eval_set


class ValidationReport(ValidationStyles):
    """
    Отчет о финальной модели для отправки
    на валидацию / бизнес заказчику.

    """

    def __init__(
        self,
        config,
        estimator,
        metric_name,
        vectorizer: callable = None,
        path_to_save: str = "./val_report.xlsx",
        used_features: Optional[List[str]] = None,
        categorical_features: Optional[List[str]] = None,
        create_file: bool = True,
        metric_col_name: str = "gini",
        metric_params: dict = None,
        images_dir_path: str = None,
        task: str = "binary",
        subtask: str = "tabular",  # [tabular, nlp]
        multiclass_artifacts: Optional[Dict] = None,
        custom_model: bool = False,
        user_config=None,
        text_column: str = None,
        text_preprocessed_column: str = None,
        group_column: str = None,
        time_column: str = None,
        create_pdf: bool = False,
        number_of_simulations_1_1: int = 200,
        number_of_simulations_3_2: int = 100
    ):

        self.artifacts_config = {
            "estimator": estimator,
            "used_features": used_features,
            "vectorizer": vectorizer,
            "categorical_features": categorical_features,
            "task": task,
            "subtask": subtask,
            "metric_name": metric_name,
            "images_dir_path": images_dir_path,
            "multiclass_artifacts": multiclass_artifacts,
            "metric_params": metric_params,
            "text_column": text_column,
            "text_preprocessed_column": text_preprocessed_column,
            "log_target_transformer": None,  # заглушка, означает, что log_target не будет применяться
            "group_column": group_column,
            "time_column": time_column,
            "number_of_simulations_1_1": number_of_simulations_1_1,
            "number_of_simulations_3_2": number_of_simulations_3_2,
            "custom_model": custom_model,
        }
        self.custom_model = custom_model
        self.user_config = user_config
        super().__init__(path_to_save, create_file, create_pdf)
        self.validation_test_config = prepare_validation_test_config(
            config, raise_exception=(not self.custom_model)
        )

        artifacts_config = deepcopy(self.artifacts_config)
        estimator_artifacts_config = {
            k: v
            for k, v in artifacts_config.items()
            if k
            not in [
                "text_column",
                "text_preprocessed_column",
                "subtask",
                "images_dir_path",
                "multiclass_artifacts",
                "group_column",
                "metric_params",
                "time_column",
            ]
        }

        if hasattr(estimator, "model"):
            artifacts_config = deepcopy(estimator_artifacts_config)
            artifacts_config["estimator"] = estimator.model

        elif hasattr(estimator, "estimator"):
            artifacts_config["estimator"] = estimator.estimator

        self.estimator = Estimator(
            estimator=artifacts_config["estimator"],
            vectorizer=artifacts_config["vectorizer"],
            log_target_transformer=artifacts_config["log_target_transformer"],
            used_features=artifacts_config["used_features"],
            categorical_features=artifacts_config["categorical_features"],
            task=artifacts_config["task"],
            metric_name=artifacts_config["metric_name"],
        )

        self.config = config
        self.test_results = {}
        self.test_5_results = []
        self.test_5_1_result = None
        self.images_dir_path = images_dir_path
        self.sheet_descriptions = {}
        # self.hyperlinks_for_total_results_page = {}
        self.metric_name = metric_name
        self.metric_col_name = metric_col_name
        self.metric_params = metric_params
        self.multiclass_artifacts = multiclass_artifacts

    @property
    def stability_block_light(self):
        """
        Получение финального светофора по блоку 5,
        "Стабильность модели".

        """
        alternative = np.where(
            "red" in self.test_5_results,
            "red",
            np.where("yellow" in self.test_5_results, "yellow", "green"),
        )
        result = pd.DataFrame(
            {
                "Ожидаемый результат теста 5.2": [self.test_5_1_result],
                "Худший результат по остальным тестам": [alternative.tolist()],
            }
        )
        result["Ожидаемый результат по блоку «Стабильность модели»"] = "-"
        if self.test_5_1_result == "green":
            result["Ожидаемый результат по блоку «Стабильность модели»"] = np.where(
                alternative.tolist() == "red", "yellow", "green"
            )
        elif self.test_5_1_result == "yellow":
            result["Ожидаемый результат по блоку «Стабильность модели»"] = np.where(
                alternative.tolist() == "red", "red", "yellow"
            )
        else:
            result["Ожидаемый результат по блоку «Стабильность модели»"] = "red"

        self.test_results["Стабильность модели"] = result[
            "Ожидаемый результат по блоку «Стабильность модели»"
        ].values.tolist()[0]
        return result

    def _create_business_task_description_page(self, **data):
        """
        Страница с описанием бизнес-задачи

        """
        _logger.info("0. Сбор статистики о решаемой бизнес-задаче.")
        bcd = BusinessCaseDescription(
            self.artifacts_config, self.validation_test_config
        )
        stats = bcd._create_description(**data)
        stats.to_excel(self.writer, sheet_name="Описание бизнес-задачи", index=False)
        self.sheet_descriptions["Описание бизнес-задачи"] = ""

        ws = self.writer.sheets["Описание бизнес-задачи"]
        b_task_cell_format = NamedStyle(name="b_task_cell_format")
        b_task_cell_format.font = Font(color="808080")
        b_task_cell_format.alignment = Alignment(
            wrap_text=True, horizontal="center", vertical="center"
        )
        brd = Side(border_style="thin", color="000000")
        b_task_cell_format.border = Border(left=brd, right=brd, top=brd, bottom=brd)

        self.wb.add_named_style(b_task_cell_format)

        b_task_row_title_format = NamedStyle(name="b_task_row_title_format")
        b_task_row_title_format.font = Font(bold=True)
        b_task_row_title_format.alignment = Alignment(
            horizontal="center", vertical="center"
        )
        brd = Side(border_style="thin", color="000000")
        b_task_row_title_format.border = Border(
            left=brd, right=brd, top=brd, bottom=brd
        )

        self.wb.add_named_style(b_task_row_title_format)

        b_task_col_title_format = NamedStyle(name="b_task_col_title_format")
        b_task_col_title_format.font = Font(bold=True)
        b_task_col_title_format.fill = PatternFill(
            start_color="00CC99", end_color="00CC99", fill_type="solid"
        )
        b_task_col_title_format.alignment = Alignment(
            horizontal="center", vertical="center"
        )
        brd = Side(border_style="thin", color="000000")
        b_task_col_title_format.border = Border(
            left=brd, right=brd, top=brd, bottom=brd
        )

        self.wb.add_named_style(b_task_col_title_format)

        business_task_text = (
            "<Заполните название модели>\n"
            "например: Модель оттока клиентов по пакету услуг Премьер"
        )
        task_description_text = (
            "<Заполните описание модели: цель моделирования, "
            "для какого бизнес процесса строится модель, какие данные используются >\n"
            "например: Модель для выделения клиентов банка склонных к закрытию Пакета Услуг Премьер. \n"
            "Модель строится на данных ЦОД по информации об открытии/закрытии пакетов услуг клиентами банка. \n"
            "В качестве факторов используются данные клиентского профиля трайба Массовая Персонализация."
        )
        report_dt_description_text = (
            "<Укажите отчетные даты> \n" "например: 31.01.2024, 28.02.2024, 31.03.2024"
        )
        target_description_text = "<Заполните определние целевого события/переменной>"
        data_selection_text = (
            "<Заполните критерии отбора популяции: фильтры, исключения>\n"
            "например: 1. Клиенты, по которым есть информация в витрине \n"
            "клиентского профиля на отчетную дату.\n"
        )
        used_algo_text = "<DreamML XGBClassifier>"
        ws.merge_cells("B2:E2")
        ws.merge_cells("B3:E3")
        ws.merge_cells("B4:D4")
        ws.merge_cells("B5:D5")
        ws.merge_cells("B6:E6")
        ws.merge_cells("B7:E7")

        ws["B2"] = business_task_text
        ws["B3"] = task_description_text
        ws["B4"] = report_dt_description_text
        ws["B5"] = data_selection_text
        ws["B6"] = target_description_text
        ws["B7"] = used_algo_text
        ws["E4"] = report_dt_description_text
        ws["E5"] = data_selection_text
        ws["A1"] = "Параметр"
        ws["A2"] = "Бизнес-задача"
        ws["A3"] = "Описание задачи"
        ws["A4"] = "Даты сбора данных"
        ws["A5"] = "Отбор наблюдений"
        ws["A6"] = "Описание целевой переменной"
        ws["A7"] = "Используемый ML-алгоритм"
        ws["B1"] = "Выборка train"
        ws["C1"] = "Выборка valid"
        ws["D1"] = "Выборка test"
        ws["E1"] = "Выборка Out-Of-Time"

        for row in ws["B2":"E7"]:
            for cell in row:
                cell.style = "b_task_cell_format"
        for row in ws["A1":"E1"]:
            for cell in row:
                cell.style = "b_task_col_title_format"
        for row in ws["A2":"A7"]:
            for cell in row:
                cell.style = "b_task_row_title_format"

        for i in range(1, 6):
            ws.column_dimensions[get_column_letter(i)].width = 30
        for i in range(1, 8):
            ws.row_dimensions[i].height = 75

        # Гиперссылка на оглавление
        ws["A9"] = "<<< Вернуться к оглавлению"
        ws["A9"].hyperlink = f"#'Оглавление'!A1"
        ws["A9"].style = "Hyperlink"

    def _create_pipeline_description_page(self, **data):
        """
        Страница с описанием всех стадий пайплайна
        """
        _logger.info("1. Сбор статистики об используемом пайплайне.")
        pst = TaskDescriptionTest(self.artifacts_config, self.validation_test_config)
        stats = pst._create_description(**data)
        stats.to_excel(self.writer, sheet_name="Описание пайплайна", index=False)
        self.add_table_borders(stats, sheet_name="Описание пайплайна")
        self.add_cell_width(sheet_name="Описание пайплайна")
        self.add_header_color(stats, sheet_name="Описание пайплайна", color="00CC99")
        self.add_numeric_format(stats, sheet_name="Описание пайплайна", startcol=1)

        ws = self.sheets["Описание пайплайна"]
        # Гиперссылка на оглавление
        ws[f"A{stats.shape[0] + 3}"] = "<<< Вернуться к оглавлению"
        ws[f"A{stats.shape[0] + 3}"].hyperlink = f"#'Оглавление'!A1"
        ws[f"A{stats.shape[0] + 3}"].style = "Hyperlink"

        self.sheet_descriptions["Описание пайплайна"] = (
            "Страница с описанием всех стадий пайплайна"
        )

    def _create_dml_info_page(self):
        kernel_folder_path = Path(sys.executable).parent.parent
        kernel_name = os.path.basename(kernel_folder_path)
        try:
            kernel_path = os.path.join(kernel_folder_path, "jh_kernel", "kernel.json")
            with open(kernel_path, "r") as file:
                kernel = json.load(file)
            kernel_version = kernel["display_name"]
        except FileNotFoundError:
            kernel_version = kernel_name

        df = pd.DataFrame(
            {
                "Module": ["DreamML", "Kernel"],
                "Version": [dml.__version__, kernel_version],
            }
        )
        df.to_excel(self.writer, sheet_name="V", startrow=0, startcol=0, index=False)
        self.add_header_color(df, sheet_name="V", color="00CC99", startcol=0)
        self.add_cell_width(
            sheet_name="V",
        )

        ws = self.sheets["V"]
        msg = "Made in DreamML"
        ws[f"A{df.shape[0] + 4}"] = msg
        ws[f"A{df.shape[0] + 4}"].font = Font(italic=True)

    def _create_data_description_page(self, **data):
        _logger.info("2. Сбор статистики о данных.")
        transformer = DataStatisticsTest(
            self.artifacts_config,
            self.validation_test_config,
        )
        result = transformer.transform(**data)

        startrows = [
            0,
            2 + len(result[0]),
            4 + len(result[0]) + len(result[1]),
        ]

        if self.artifacts_config["task"] not in ["multiclass", "multilabel"]:
            eventrate_columns = [[4], [], []]
            integer_columns = [[2, 3], [2, 3], [2]]
        else:
            idx_events_columns, idx_eventrate_columns = [], []
            for col_idx, column in enumerate(result[0].columns):
                if "events" in column:
                    idx_events_columns.append(col_idx + 1)
                if "eventrate" in column:
                    idx_eventrate_columns.append(col_idx + 1)

            eventrate_columns = [idx_eventrate_columns, [], []]
            integer_columns = [idx_events_columns, [2, 3], [2]]

        for result, start_index, e_col, int_col in zip(
            result, startrows, eventrate_columns, integer_columns
        ):
            result.to_excel(
                self.writer,
                startrow=start_index,
                sheet_name="Описание данных",
                index=False,
            )
            self.add_table_borders(
                result, sheet_name="Описание данных", startrow=start_index
            )
            self.add_header_color(
                result,
                sheet_name="Описание данных",
                startrow=start_index,
                color="00CC99",
            )
            self.add_numeric_format(
                result,
                sheet_name="Описание данных",
                startrow=start_index,
                to_eventrate_format=e_col,
                to_integer_format=int_col,
            )

        self.add_cell_width(sheet_name="Описание данных")

        ws = self.sheets["Описание данных"]
        # Гиперссылка на оглавление
        ws["E7"] = "<<< Вернуться к оглавлению"
        ws["E7"].hyperlink = f"#'Оглавление'!A1"
        ws["E7"].style = "Hyperlink"

        self.sheet_descriptions["Описание данных"] = "Страница со статистикой о данных"

    def _create_features_importance_page(self, **data):
        """
        Лист с важностью признаков.

        """
        msg = "7. Оценка важности признаков модели по метрике SHAP-values."
        _logger.info(msg)
        if "valid" in data:
            x, _ = data["valid"]
        elif "train" in data:
            x, _ = data["train"]
        elif "test" in data:
            x, _ = data["test"]
        elif "OOT" in data:
            x, _ = data["OOT"]

        shap_values, imp = self.estimator.get_shap_importance(x)
        imp.to_excel(
            self.writer, sheet_name="Важность признаков", startrow=0, index=False
        )
        self.add_numeric_format(imp, sheet_name="Важность признаков", min_value=100)
        self.add_table_borders(imp, sheet_name="Важность признаков")
        self.add_header_color(imp, sheet_name="Важность признаков", color="00CC99")
        self.add_cell_width(sheet_name="Важность признаков")
        ws = self.sheets["Важность признаков"]

        if isinstance(shap_values, np.ndarray):
            plt.clf()
            shap.initjs()
            shap.summary_plot(
                shap_values,
                features=x[self.estimator.used_features],
                feature_names=self.estimator.used_features,
                show=False,
            )
            plt.savefig(
                os.path.join(self.images_dir_path, "shap_summary_plot.png"),
                bbox_inches="tight",
                pad_inches=0.1,
            )
            plt.close()

            img = Image(os.path.join(self.images_dir_path, "shap_summary_plot.png"))
            img.anchor = f"A{imp.shape[0] + 2}"
            ws.add_image(img)

        # Гиперссылка на оглавление
        ws["E2"] = "<<< Вернуться к оглавлению"
        ws["E2"].hyperlink = f"#'Оглавление'!A1"
        ws["E2"].style = "Hyperlink"

        msg = "Лист с оценкой важности признаков модели по метрике SHAP-values"
        self.sheet_descriptions["Важность признаков"] = msg

    def _create_permutation_importance_page(self, **data):
        """
        Страница с проверкой важности признаков на основе перестановок.
        """
        msg = "8. Проверка важности признаков на основе перестановок."
        _logger.info(msg)

        perm_test = PermutationImportanceChecker(
            writer=self.writer,
            model=self.artifacts_config["estimator"],
            features_list=self.artifacts_config["used_features"],
            cat_features=self.artifacts_config["categorical_features"],
            images_dir_path=self.images_dir_path,
            metric_name=self.metric_name,
            metric_col_name=self.metric_col_name,
            metric_params=self.metric_params,
            task=self.artifacts_config["task"],
        )
        perm_df = perm_test.validate(**data)

        perm_df.to_excel(
            self.writer, sheet_name="Permutation importance", startrow=0, index=False
        )
        self.add_table_borders(perm_df, sheet_name="Permutation importance")
        self.add_header_color(
            perm_df, sheet_name="Permutation importance", color="00CC99"
        )
        self.add_cell_width(sheet_name="Permutation importance")
        self.add_bottom_table_borders(perm_df, sheet_name="Permutation importance")
        self.add_numeric_format(
            perm_df, sheet_name="Permutation importance", min_value=100
        )

        ws = self.sheets["Permutation importance"]
        ws["E2"] = (
            "Permutation importance - метрика важности признака в построенной модели"
        )
        ws["E3"] = (
            f"Считается как относительное изменение метрики качества модели при перемешивании значений признака"
        )
        ws["E5"] = (
            "Factors relevancy - доля факторов с важностью 20% и более от фактора с максимальной важностью"
        )
        ws["E7"] = "* - данный тест информативный"

        # Гиперссылка на оглавление
        ws["E9"] = "<<< Вернуться к оглавлению"
        ws["E9"].hyperlink = f"#'Оглавление'!A1"
        ws["E9"].style = "Hyperlink"

        msg = "Страница с проверкой важности признаков на основе перестановок"
        self.sheet_descriptions["Permutation importance"] = msg
        plt.close()

        img = Image(f"{self.images_dir_path}/Permutation importance.png")
        img.anchor = f"A{perm_df.shape[0] + 4}"
        ws.add_image(img)

    def _create_model_params_page(self, **data):
        msg = "9. Cбор гиперпарметров модели."
        _logger.info(msg)
        try:
            params = self.estimator.get_estimator_params
        except KeyError:
            params = pd.DataFrame(columns=["Гиперпараметр", "Значение"])

        params.to_excel(
            self.writer,
            sheet_name="Гиперпараметры модели",
            index=False,
            startrow=0,
        )
        self.add_table_borders(params, sheet_name="Гиперпараметры модели")
        self.add_header_color(
            params, sheet_name="Гиперпараметры модели", color="00CC99"
        )
        self.add_cell_width(
            sheet_name="Гиперпараметры модели",
        )

        ws = self.sheets["Гиперпараметры модели"]
        if params.empty:
            ws["A2"] = "У LAMA нет интерфейса для гиперпараметров"
        # Гиперссылка на оглавление
        ws["E2"] = "<<< Вернуться к оглавлению"
        ws["E2"].hyperlink = f"#'Оглавление'!A1"
        ws["E2"].style = "Hyperlink"

        self.sheet_descriptions["Гиперпараметры модели"] = ""

    def _create_gini_coef(self, **data):
        """
        Коэффициент Джини модели на всех семплах.
        """
        _sheet_name = "Коэффициент Джини модели"
        msg = ("3. Коэффициент Джини модели")
        _logger.info(msg)
        metric = metrics_mapping["gini"](
            model_name=self.config["model_name"],
            task="binary"
        )

        gini_test = GiniTest(self.estimator, metric)
        result, traffic_light_gini_degradation = gini_test.transform(**data)

        self.test_results[_sheet_name] = traffic_light_gini_degradation

        result.to_excel(
            self.writer,
            sheet_name=_sheet_name,
            startrow=3,
            index=False,
        )

        self.add_table_borders(result, sheet_name=_sheet_name)
        # self.add_header_color(result, sheet_name="Коэффициент Джини модели.", color="00CC99")
        self.add_numeric_format(result, sheet_name=_sheet_name, min_value=100)
        self.add_traffic_light_color(sheet_name=_sheet_name)
        self.add_cell_width(sheet_name=_sheet_name)

        ws = self.sheets[_sheet_name]
        msg = ("Тест 1 -  Коэффициент Джини модели")

        ws[f"A1"] = msg
        ws["A9"] = "<<< Вернуться к оглавлению"
        ws["A9"].hyperlink = f"#'Оглавление'!A1"
        ws["A9"].style = "Hyperlink"

        self.sheet_descriptions[_sheet_name] = msg
        # self.hyperlinks_for_total_results_page["1"] = _sheet_name

    def _create_ks_coef(self, **data):
        """
        Статистика Колмогорова-Смирнова.
        """
        _sheet_name = "Статистика Колмогорова-Смирнова"
        msg = ("4. Статистика Колмогорова-Смирнова")
        _logger.info(msg)

        ks_test = KSTest(self.estimator)
        result, result_traffic_light = ks_test.transform(**data)

        self.test_results[_sheet_name] = result_traffic_light

        result.to_excel(
            self.writer,
            sheet_name=_sheet_name,
            startrow=3,
            index=False,
        )

        self.add_table_borders(result, sheet_name=_sheet_name)
        self.add_numeric_format(result, sheet_name=_sheet_name, min_value=100)
        self.add_traffic_light_color(sheet_name=_sheet_name)
        self.add_cell_width(sheet_name=_sheet_name)

        ws = self.sheets[_sheet_name]
        msg = ("Тест 2 -  Статистика Колмогорова-Смирнова")

        ws[f"A1"] = msg
        ws["A9"] = "<<< Вернуться к оглавлению"
        ws["A9"].hyperlink = f"#'Оглавление'!A1"
        ws["A9"].style = "Hyperlink"

        self.sheet_descriptions[_sheet_name] = msg
        # self.hyperlinks_for_total_results_page["2"] = _sheet_name

    def _create_f1_coef(self, **data):
        """
        F1 - мера модели на всех семплах.
        """
        _sheet_name = "F1 - мера"
        msg = ("5. F1 - мера модели")
        _logger.info(msg)
        metric = metrics_mapping["f1_score"](
            model_name=self.config["model_name"],
            task="binary"
        )

        gini_test = FTest(self.estimator, metric)
        result, result_traffic_light = gini_test.transform(**data)

        self.test_results[_sheet_name] = result_traffic_light


        result.to_excel(
            self.writer,
            sheet_name=_sheet_name,
            startrow=3,
            index=False,
        )

        self.add_table_borders(result, sheet_name=_sheet_name)
        self.add_numeric_format(result, sheet_name=_sheet_name, min_value=100)
        self.add_traffic_light_color(sheet_name=_sheet_name)
        self.add_cell_width(sheet_name=_sheet_name)

        ws = self.sheets[_sheet_name]
        msg = ("Тест 3 -  F1 - мера модели")

        ws[f"A1"] = msg
        ws["A9"] = "<<< Вернуться к оглавлению"
        ws["A9"].hyperlink = f"#'Оглавление'!A1"
        ws["A9"].style = "Hyperlink"

        self.sheet_descriptions[_sheet_name] = msg
        # self.hyperlinks_for_total_results_page["1"] = _sheet_name

    def _create_ciec_coef(self, **data):
        """
        Conditional Information Entropy Coefficient на всех семплах.
        """
        _sheet_name = "CIEC"
        msg = ("6. Conditional Information Entropy Coefficient")
        _logger.info(msg)

        gini_test = CIECTest(self.estimator)
        result, result_traffic_light = gini_test.transform(**data)

        self.test_results[_sheet_name] = result_traffic_light

        result.to_excel(
            self.writer,
            sheet_name=_sheet_name,
            startrow=3,
            index=False,
        )

        self.add_table_borders(result, sheet_name=_sheet_name)
        self.add_numeric_format(result, sheet_name=_sheet_name, min_value=100)
        self.add_traffic_light_color(sheet_name=_sheet_name)
        self.add_cell_width(sheet_name=_sheet_name)

        ws = self.sheets[_sheet_name]
        msg = ("Тест 4 -  Conditional Information Entropy Coefficient")

        ws[f"A1"] = msg
        ws["A9"] = "<<< Вернуться к оглавлению"
        ws["A9"].hyperlink = f"#'Оглавление'!A1"
        ws["A9"].style = "Hyperlink"

        self.sheet_descriptions[_sheet_name] = msg
        # self.hyperlinks_for_total_results_page["1"] = _sheet_name

    def _create_total_results(self):
        _sheet_name = "Итоговый результат"

        ws = self.sheets[_sheet_name]
        msg = ("Таблица с итоговыми результатами")
        ws["A1"] = msg

        result = pd.DataFrame(self.test_results.items())

        result.to_excel(
            self.writer,
            sheet_name=_sheet_name,
            startrow=3,
            index=False,
        )

        self.add_table_borders(result, startrow=3, sheet_name=_sheet_name)
        # self.add_numeric_format(result, sheet_name=_sheet_name, min_value=100)
        # self.add_table_borders(result, sheet_name="Описание пайплайна")
        self.add_traffic_light_color(sheet_name=_sheet_name)
        self.add_cell_width(sheet_name=_sheet_name)

        ws["A2"].value = None
        ws["B1"].value = None
        ws["A4"].value = None
        ws["B4"].value = None

        # self.add_header_color(result, sheet_name="Коэффициент Джини модели.", color="00CC99")

    def create_report(self, **data):
        """
        Создать полный валидационный отчет.
        """
        report_id = uuid.uuid4().hex

        task = self.artifacts_config["task"]
        custom_model = self.custom_model
        user_config = self.user_config

        if custom_model:
            experiment_name = None
        else:
            experiment_name = Path(self.dir).parent.parent.name

        start_time = time.time()
        _logger.monitor(
            f"Creating validation report for {task} task.",
            extra={
                "log_data": ReportStartedLogData(
                    task=task,
                    development=False,
                    custom_model=custom_model,
                    experiment_name=experiment_name,
                    report_id=report_id,
                    user_config=user_config,
                )
            },
        )

        self._create_dml_info_page()

        self._create_business_task_description_page(**data)
        self._create_pipeline_description_page(**data)

        pd.DataFrame([" "]).to_excel(self.writer, sheet_name="Итоговый результат")
        self._create_data_description_page(**data)

        self._create_gini_coef(**data)
        self._create_ks_coef(**data)
        self._create_f1_coef(**data)
        self._create_ciec_coef(**data)

        try:
            self._create_features_importance_page(**data)
        except:
            pass

        if self.artifacts_config["subtask"] != "nlp":
            self._create_permutation_importance_page(**data)

        try:
            self._create_model_params_page(**data)
        except:
            pass
        self._create_total_results()

        self.writer.save()

        elapsed_time = time.time() - start_time
        _logger.monitor(
            f"Validation report for {task} task is created in {elapsed_time:.1f} seconds.",
            extra={
                "log_data": ReportFinishedLogData(
                    report_id=report_id,
                    elapsed_time=elapsed_time,
                )
            },
        )

    def create_traffic_light(self, **data):
        """
        Создать итоговый светофор и провести только необходимые для него тесты.

        """

        results = []

        metric = metrics_mapping["gini"](
            model_name=self.config["model_name"],
            task="binary"
        )
        gini_test = GiniTest(self.estimator, metric)
        result, traffic_light_gini_degradation = gini_test.transform(**data)
        results.append(traffic_light_gini_degradation)

        ks_test = KSTest(self.estimator)
        result, result_traffic_light = ks_test.transform(**data)
        results.append(result_traffic_light)

        metric = metrics_mapping["f1_score"](
            model_name=self.config["model_name"],
            task="binary"
        )
        gini_test = FTest(self.estimator, metric)
        result, result_traffic_light = gini_test.transform(**data)
        results.append(result_traffic_light)

        gini_test = CIECTest(self.estimator)
        result, result_traffic_light = gini_test.transform(**data)
        results.append(result_traffic_light)

        if "red" in results:
            self.test_results["Итог"] = "red"
        elif "yellow" in results:
            self.test_results["Итог"] = "yellow"
        else:
            self.test_results["Итог"] = "green"
        results.append(str(self.test_results["Итог"]))

        return results