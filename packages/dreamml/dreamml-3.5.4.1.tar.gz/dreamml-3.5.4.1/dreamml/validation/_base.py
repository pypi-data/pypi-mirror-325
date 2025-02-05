import string
from typing import List
import numpy as np
import pandas as pd
from tqdm.auto import tqdm
from openpyxl.styles import NamedStyle, Font, PatternFill, Border, Side, Alignment
from openpyxl.styles.numbers import FORMAT_PERCENTAGE_00
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder
from openpyxl.utils import get_column_letter

from dreamml.logging import get_logger
from dreamml.reports.pdf_maker import convert_excel_to_pdf, pdf_maker_available

_logger = get_logger(__name__)


class ValidationStyles:

    def __init__(
        self,
        path_to_save: str = "./validation_report.xlsx",
        create_file: bool = True,
        create_pdf: bool = False,
    ):
        self.dir = path_to_save
        if create_file:
            # Сделано для использования с отчетом о разработки, что бы не создавались файлы отчета о валидации.
            self.writer = pd.ExcelWriter(
                path=path_to_save,
                engine="openpyxl",
                datetime_format="dd-mm-yy hh:mm:ss",
            )
            self.letters = string.ascii_uppercase
            self.sheets = self.writer.sheets
            self.wb = self.writer.book
        self._create_pdf = create_pdf

    def create_pdf(self):
        if not self._create_pdf:
            return

        pdf_file = self.dir[:-5] if self.dir.endswith(".xlsx") else self.dir

        pdf_file += ".pdf"

        if not pdf_maker_available:
            _logger.warning("Can't make pdf as fonts aren't installed.")
            return
        convert_excel_to_pdf(self.dir, pdf_file)

    def add_cell_width(self, sheet_name: str):
        """
        Установка ширины ячейки на листе Excel.
        Для вычисления ширины ячейки вычисляется длина
        заголовков всех таблиц и максимальные значения
        в таблице. Итоговая ширина равна длине максимального
        элемента в таблице + 2 (для запаса).

        Parameters
        ----------
        sheet_name: string
            Название листа в Excel-книге для записи данных.

        """
        ws = self.sheets[sheet_name]

        dim_holder = DimensionHolder(worksheet=ws)

        for col in ws.columns:
            max_col_width = 0
            for cell in col:
                col_width = len(str(cell.value))
                if max_col_width < col_width:
                    max_col_width = col_width

            max_col_width += 2

            dim_holder[get_column_letter(col[0].column)] = ColumnDimension(
                ws, min=col[0].column, max=col[0].column, width=max_col_width
            )

        ws.column_dimensions = dim_holder

    def add_header_color(
        self,
        data,
        sheet_name: str,
        startrow: int = 0,
        startcol: int = 0,
        color: str = "77d496",
    ):
        """
        Установка цвета заголовка на листе Excel.

        Parameters
        ----------
        data: pandas.DataFrame
            Набор данных для записи в Excel.

        sheet_name: string
            Название листа в Excel-книге для записи данных.

        startrow: integer, optional, default = 0
            Номер строки, с которой начинать запись данных.

        startcol: integer, optional, default = 0
            Номер столбца, с которого начинать запись данных.

        color: string, optional, default = "77d496"
            RGB-цвет заголовка.

        """
        ws = self.sheets[sheet_name]
        startrow += 1  # openPyXl итерируется с 1, а не с 0
        startcol += 1

        drm_header_format = NamedStyle(name="drm_header_format")
        drm_header_format.font = Font(bold=True)
        drm_header_format.fill = PatternFill(
            start_color=color, end_color=color, fill_type="solid"
        )
        drm_header_format.alignment = Alignment(
            wrap_text=True, horizontal="center", vertical="center"
        )
        brd = Side(border_style="thin", color="000000")
        drm_header_format.border = Border(left=brd, right=brd, top=brd, bottom=brd)

        style_list = [x.name for x in self.wb._named_styles]
        if drm_header_format.name not in style_list:
            self.wb.add_named_style(drm_header_format)

        for col in ws.iter_cols(
            min_row=startrow,
            max_row=startrow,
            min_col=startcol,
            max_col=startcol + data.shape[1] - 1,
        ):
            for cell in col:
                cell.style = "drm_header_format"

    def add_numeric_format(
        self,
        data,
        sheet_name: str,
        startrow: int = 0,
        startcol: int = 0,
        min_value: int = 10,
        to_eventrate_format: List[int] = [],
        to_integer_format: List[int] = [],
    ):
        """
        Установка формата для числовой таблицы.

        Parameters
        ----------
        data: pandas.DataFrame
            Набор данных для записи в Excel.

        sheet_name: string
            Название листа в Excel-книге для записи данных.

        startrow: integer, optional, default = 0
            Номер строки, с которой начинать запись данных.

        startcol: integer, optional, default = 0
            Номер столбца, с которого начинать запись данных.

        set_top: integer, optional, default = 0

        min_value: interger, optional, default = 10
            Минимальное значение, формат записи которого "х",
            если значение в ячейке Excel-книге меньше min_value,
            то формат записи - "x.yy".

        to_eventrate_format: List[int], default = []
            Номера столбцов для форматирования в процентах (номера начинаются с 1).

        to_integer_format: List[int], default = []
            Номера столбцов для целочисленного форматирования (номера начинаются с 1).

        """
        ws = self.sheets[sheet_name]
        startrow += 1  # openPyXl итерируется с 1, а не с 0
        startcol += 1

        for col in ws.iter_cols(
            min_row=startrow,
            max_row=startrow + data.shape[0],
            min_col=startcol,
            max_col=startcol + data.shape[1] - 1,
        ):
            for cell in col:
                if isinstance(cell.value, float or int):
                    if cell.column in to_integer_format:
                        cell.number_format = "0"
                    elif cell.column in to_eventrate_format:
                        cell.number_format = FORMAT_PERCENTAGE_00
                    else:
                        cell.number_format = "0.00" if cell.value < min_value else "0"

    def add_table_borders(
        self, data, sheet_name: str, startrow: int = 0, startcol: int = 0
    ):
        """
        Установка границ таблицы на листе Excel.

        Parameters
        ----------
        data: pandas.DataFrame
            Набор данных для записи в Excel.

        sheet_name: string
            Название листа в Excel-книге для записи данных.

        startrow: integer, optional, default = 0
            Номер строки, с которой начинать запись данных.

        startcol: integer, optional, default = 0
            Номер столбца, с которого начинать запись данных.

        num_format: integer, optional, default = 10
            Числовой формат записи данных.

        """
        ws = self.sheets[sheet_name]
        startrow += 1  # openPyXl итерируется с 1, а не с 0
        startcol += 1

        for col in ws.iter_cols(
            min_row=startrow,
            max_row=startrow + data.shape[0],
            min_col=startcol,
            max_col=startcol,
        ):
            for cell in col:
                cell.border = Border(
                    left=Side(border_style="thin", color="000000"),
                    right=cell.border.right,
                    top=cell.border.top,
                    bottom=cell.border.bottom,
                )

        for col in ws.iter_cols(
            min_row=startrow,
            max_row=startrow + data.shape[0],
            min_col=startcol + data.shape[1] - 1,
            max_col=startcol + data.shape[1] - 1,
        ):
            for cell in col:
                cell.border = Border(
                    left=cell.border.left,
                    right=Side(border_style="thin", color="000000"),
                    top=cell.border.top,
                    bottom=cell.border.bottom,
                )

        for row in ws.iter_rows(
            min_row=startrow + data.shape[0],
            max_row=startrow + data.shape[0],
            min_col=startcol,
            max_col=startcol + data.shape[1] - 1,
        ):
            for cell in row:
                cell.border = Border(
                    left=cell.border.left,
                    right=cell.border.right,
                    top=cell.border.top,
                    bottom=Side(border_style="thin", color="000000"),
                )

    def add_bottom_table_borders(
        self, data, sheet_name: str, startrow: int = 0, startcol: int = 0
    ):
        """
        Установка верхней и нижней границ таблицы на листе Excel.

        Parameters
        ----------
        data: pandas.DataFrame
            Столбец со значениями таблицы.

        sheet_name: string
            Название листа в Excel-книге для записи данных.

        startrow: integer, optional, default = 0
            Номер строки, с которой начинать запись данных.

        startcol: integer, optional, default = 0
            Номер столбца, с которого начинать запись данных.

        """
        ws = self.sheets[sheet_name]
        startrow += 1  # openPyXl итерируется с 1, а не с 0

        drm_result_format = NamedStyle(name="drm_result_format")
        drm_result_format.font = Font(bold=True)
        brd = Side(border_style="thin", color="000000")
        drm_result_format.border = Border(left=brd, right=brd, top=brd, bottom=brd)

        style_list = [x.name for x in self.wb._named_styles]
        if drm_result_format.name not in style_list:
            self.wb.add_named_style(drm_result_format)

        for col in ws.iter_rows(
            min_row=startrow + data.shape[0],
            max_row=startrow + data.shape[0],
            min_col=startcol,
            max_col=startcol + data.shape[1],
        ):
            for cell in col:
                cell.style = "drm_result_format"

    def add_traffic_light_color(self, sheet_name: str):
        """
        Установка света светофора в результате прохождения теста.

        """
        ws = self.sheets[sheet_name]

        drm_traffic_light_format = NamedStyle(name="drm_traffic_light_format")
        drm_traffic_light_format.alignment = Alignment(
            wrap_text=True, horizontal="center"
        )
        drm_traffic_light_format.font = Font(bold=True)
        brd = Side(border_style="thin", color="000000")
        drm_traffic_light_format.border = Border(
            left=brd, right=brd, top=brd, bottom=brd
        )

        style_list = [x.name for x in self.wb._named_styles]
        if drm_traffic_light_format.name not in style_list:
            self.wb.add_named_style(drm_traffic_light_format)

        for col in ws.columns:
            for cell in col:
                if cell.value == "green":
                    cell.style = "drm_traffic_light_format"
                    cell.fill = PatternFill(
                        start_color="66CC99", end_color="66CC99", fill_type="solid"
                    )
                if cell.value == "yellow":
                    cell.style = "drm_traffic_light_format"
                    cell.fill = PatternFill(
                        start_color="FFFF33", end_color="FFFF33", fill_type="solid"
                    )
                if cell.value == "red":
                    cell.style = "drm_traffic_light_format"
                    cell.fill = PatternFill(
                        start_color="CC0000", end_color="CC0000", fill_type="solid"
                    )


class BaseTest:

    def _predict(self, X) -> np.array:
        """
        Метод для применения модели к выборке X.

        Parameters
        ----------
        X: pandas.core.frame.DataFrame
            Матрица признаков для применения модели.

        Returns
        -------
        y_pred: np.array
            Вектор прогнозов.

        """
        if hasattr(self, "vectorizer") and self.vectorizer is not None:
            X = self.vectorizer.transform(X)

        if hasattr(self.estimator, "predict_proba"):
            return self.estimator.predict_proba(X[self.used_features])[:, 1]
        if hasattr(self.estimator, "transform"):
            return self.estimator.transform(X)

        return self.estimator.predict(X)

    def label_binarizing(self, task, label_binarizer, labels, target_name, **eval_set):
        if task == "multiclass":
            for sample_name, (X_sample, y_sample) in tqdm(eval_set.items()):
                if len(y_sample.shape) == 1 and label_binarizer is not None:
                    data = label_binarizer.transform(y_sample)
                    y_sample = pd.DataFrame(
                        data=data, columns=labels, index=y_sample.index
                    )
                eval_set[sample_name] = (X_sample, y_sample)

        return eval_set

    def _calculate_macro_score(
        self, metric_class, y_true: np.ndarray, y_pred: np.ndarray
    ):
        """
        Рассчет macro метрики с/без NaN значениями.

        Parameters
        ----------
        y_true: np.ndarray - матрица таргетов (n_samples, n_classes)
        y_pred: np.ndarray - матрица предсказанных вероятностей (n_samples, n_classes)

        Returns
        -------
        Macro метрика по каждому классу
        """
        y_true = y_true.values if isinstance(y_true, pd.DataFrame) else y_true
        y_pred = y_pred.values if isinstance(y_pred, pd.DataFrame) else y_pred

        metric_list = []
        num_classes = y_true.shape[1]
        mask = ~np.isnan(y_true)

        for class_idx in range(num_classes):
            mask_by_class_idx = mask[:, class_idx]
            y_true_idx_class = y_true[:, class_idx][mask_by_class_idx]
            y_pred_idx_class = y_pred[:, class_idx][mask_by_class_idx]

            if len(np.unique(y_true_idx_class)) == 1:
                continue
            if len(np.unique(y_pred_idx_class)) == 1:
                metric_list.append(0)
                continue

            metric_list.append(metric_class(y_true_idx_class, y_pred_idx_class))
        return np.mean(metric_list)