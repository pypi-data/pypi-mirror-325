
import requests
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.styles import NamedStyle
from io import BytesIO

class xlsp:
    def __init__(self, file_url, sheet_name, access_token):
        """
        Initialize the ExcelCM with the SharePoint file URL, sheet name, and access token.
        
        :param file_url: str, URL to the Excel file on SharePoint.
        :param sheet_name: str, Name of the sheet to work with.
        :param access_token: str, Access token for Microsoft Graph API authentication.
        """
        self.file_url = file_url
        self.sheet_name = sheet_name
        self.access_token = access_token

        # Download the file from SharePoint
        self.file_content = self._download_file_from_sharepoint()

        # Load the workbook from the downloaded content
        self.workbook = load_workbook(filename=BytesIO(self.file_content), data_only=True)
        self.sheet = self.workbook[sheet_name]
        self.header_map = self._map_headers()

    def _download_file_from_sharepoint(self):
        """
        Download the file from SharePoint using the Microsoft Graph API.
        
        :return: bytes, Content of the file.
        """
        headers = {
            "Authorization": f"Bearer {self.access_token}"
        }
        response = requests.get(self.file_url, headers=headers)
        response.raise_for_status()  # Raise an error for bad responses
        return response.content

    def _map_headers(self):
        """
        Map header column names to column letters.
        
        :return: dict, Mapping of column names to column letters.
        """
        header_map = {}
        for cell in next(self.sheet.iter_rows(values_only=False, max_row=1)):
            if cell.value:
                header_map[cell.column_letter] = cell.value  # Swap key-value for reverse lookup
        return header_map

    def column_name(self, column_letter):
        """
        Get the column name from a column letter.
        
        :param column_letter: str, The column letter (e.g., 'A', 'B', etc.).
        :return: str or None, The column name if found, otherwise None.
        """
        return self.header_map.get(column_letter, None)

    def get_sheet_name(self):
        """
        Get the name of the current sheet.
        
        :return: str, Sheet name.
        """
        return self.sheet.title

    def get_number_of_columns(self):
        """
        Get the number of columns in the sheet based on the header row.
        
        :return: int, Number of columns.
        """
        return len(self.header_map)

    def get_number_of_rows(self):
        """
        Get the total number of rows in the sheet.
        
        :return: int, Number of rows.
        """
        return self.sheet.max_row

    def get_cell_by_column_name(self, row_num, column_name):
        """
        Access a cell by row number and column name.
        
        :param row_num: int, Row number of the cell.
        :param column_name: str, Column name in the header row.
        :return: openpyxl.cell.Cell, The cell object.
        """
        # Search for the column letter corresponding to the column name
        column_letter = None
        for letter, name in self.header_map.items():
            if name == column_name:
                column_letter = letter
                break

        if not column_letter:
            raise ValueError(f"Column name '{column_name}' not found in header.")

        return self.sheet[f"{column_letter}{row_num}"]

    def get_headers(self):
        """
        Get the header row as a list of column names.
        
        :return: list, Column headers.
        """
        return [cell.value for cell in next(self.sheet.iter_rows(max_row=1))]

    def save(self, local_path="local_copy.xlsx"):
        """
        Save changes to a local file.
        
        :param local_path: str, Path to save the local copy of the file.
        """
        self.workbook.save(local_path)

    def save_to_sharepoint(self, upload_url):
        """
        Save changes back to SharePoint by uploading the modified file.
        
        :param upload_url: str, URL to upload the file to SharePoint using Microsoft Graph API.
        """
        # Save the workbook to a BytesIO object
        file_stream = BytesIO()
        self.workbook.save(file_stream)
        file_stream.seek(0)  # Reset the stream position to the beginning

        # Close workbook to ensure all changes are saved properly
        self.workbook.close()

        # Upload the file to SharePoint
        headers = {
            "Authorization": f"Bearer {self.access_token}",
            "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        }
        response = requests.put(upload_url, headers=headers, data=file_stream)
        
        # Raise error if upload fails
        response.raise_for_status()
        
        print("Workbook successfully saved to SharePoint.")

class xlOSMLog(xlsp):

    # OSM Reformat Log column names
    log_column_name = ['osm_log_id', 'osm_status', 'data_engineer_submit_date',
        'osm_submit_date', 'original_file_name', 'list_vendor', 'list_name',
        'report_campaign_title', 'actual_file_name', 'osm_aps', 'record_count',
        'treated', 'control', 'lead_date', 'aps_end_date', 'record_source',
        'node_1', 'node_2', 'node_3', 'media_type', 'touches',
        'cell_code_definition', 'cellcode_p1', 'p2', 'p3', 'p4', 'p5',
        'reference_counts_and_cell_coding', 'file_name_parse_1',
        'file_name_parse_2'
    ]

    def __init__(self, file_url, sheet_name, access_token):
        """
        Initialize the ExcelOSMLog with SharePoint file details and map headers.
        """
        super().__init__(file_url, sheet_name, access_token)  # Call parent class constructor
        self.column_mapping = self._generate_column_mapping()  # Store mapping as an exposed variable

    def _generate_column_mapping(self):
        """
        Create a mapping between log_column_name and header_map.

        :return: dict, Mapping of {Column Letter: (Log Column Name, Header Name)}
        """
        mapping = {}
        for index, column_letter in enumerate(self.header_map.keys()):
            log_name = self.log_column_name[index] if index < len(self.log_column_name) else "N/A"
            header_name = self.header_map.get(column_letter, "Not in header")
            mapping[column_letter] = (log_name, header_name)

        return mapping  # Exposed as self.column_mapping

    def get_row_by_osm_log_id(self, osm_log_id):
        """
        Get the row associated with a specific OSM Log ID.

        :param osm_log_id: int, The OSM Log ID to search for.
        :return: dict, Row data with column names as keys and cell values as values.
        """
        # Find the column letter for 'osm_log_id'
        osm_log_id_col_letter = None
        for col_letter, col_name in self.header_map.items():
            if col_name == 'OSM Log ID':
                osm_log_id_col_letter = col_letter
                break

        if not osm_log_id_col_letter:
            raise ValueError("OSM Log ID column not found in header.")

        # Get the index of the column letter in the header
        osm_log_id_index = list(self.header_map.keys()).index(osm_log_id_col_letter)

        # Iterate through rows and find the matching OSM Log ID
        for row in self.sheet.iter_rows(min_row=2, values_only=True):
            if row[osm_log_id_index] == osm_log_id:  # Match based on column index
                return {self.header_map[col_letter]: row[idx] for idx, col_letter in enumerate(self.header_map)}

        return None  # Return None if not found
