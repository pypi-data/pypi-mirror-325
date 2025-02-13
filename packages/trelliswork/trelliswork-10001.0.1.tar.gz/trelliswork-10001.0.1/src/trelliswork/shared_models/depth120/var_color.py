import re

from openpyxl.styles import PatternFill

from ..depth110 import ColorSystem, WebSafeColor


class VarColor():
    """様々な色指定
    """


    @classmethod
    @property
    def AUTO(clazz):
        return 1


    @classmethod
    @property
    def DARKNESS(clazz):
        return 2


    @classmethod
    @property
    def PAPER_COLOR(clazz):
        return 3


    @classmethod
    @property
    def TONE_AND_COLOR_NAME(clazz):
        return 4


    @classmethod
    @property
    def WEB_SAFE_COLOR(clazz):
        return 5


    @classmethod
    @property
    def XL_COLOR_CODE(clazz):
        return 6


    @staticmethod
    def what_am_i(var_color_value):
        """トーン名・色名の欄に何が入っているか判定します
        """

        # 何も入っていない、または False が入っている
        if not var_color_value:
            return False

        # ナンが入っている
        if var_color_value is None:
            return None

        if isinstance(var_color_value, dict):
            var_color_dict = var_color_value
            if 'darkness' in var_color_dict:
                return VarColor.DARKNESS
            
            else:
                raise ValueError(f'未定義の色指定。 {var_color_value=}')


        # ウェブ・セーフ・カラーが入っている
        #
        #   とりあえず、 `#` で始まるなら、ウェブセーフカラーとして扱う
        #
        #if var_color_value.startswith('#'):
        if re.match(r'^#[0-9a-fA-f]{6}$', var_color_value):
            return VarColor.WEB_SAFE_COLOR

        if re.match(r'^[0-9a-fA-f]{6}$', var_color_value):
            return VarColor.XL_COLOR_CODE

        # 色相名と色名だ
        #if '.' in var_color_value:
        if re.match(r'^[0-9a-zA-Z_]+\.[0-9a-zA-Z_]+$', var_color_value):
            return VarColor.TONE_AND_COLOR_NAME

        # 影の自動設定
        if var_color_value == 'auto':
            return VarColor.AUTO
        
        # 紙の色
        if var_color_value == 'paperColor':
            return VarColor.PAPER_COLOR
        
        raise ValueError(f"""ERROR: what_am_i: undefined {var_color_value=}""")


    def __init__(self, var_color_value):
        self._var_color_value = var_color_value
        self._var_type = VarColor.what_am_i(var_color_value)


    @property
    def var_color_value(self):
        return self._var_color_value


    @property
    def var_type(self):
        return self._var_type


    def to_web_safe_color_obj(self, contents_doc):
        """様々な色名をウェブ・セーフ・カラーの１６進文字列の色コードに変換します
        """

        # 色が指定されていないとき、この関数を呼び出してはいけません
        if not self.var_type:
            raise Exception(f'var_color_name_to_web_safe_color_code: 色が指定されていません')

        # 背景色を［なし］にします。透明（transparent）で上書きするのと同じです
        if self.var_type == VarColor.PAPER_COLOR:
            raise Exception(f'var_color_name_to_web_safe_color_code: 透明色には対応していません')

        # ［auto］は自動で影の色を設定する機能ですが、その機能をオフにしているときは、とりあえず黒色にします
        if self.var_type == VarColor.AUTO:
            web_safe_color_code = ColorSystem.alias_to_web_safe_color_dict(contents_doc=contents_doc)['xlTheme']['xlBlack']
            return WebSafeColor(web_safe_color_code)

        # ウェブセーフカラー
        if self.var_type == VarColor.WEB_SAFE_COLOR:
            return WebSafeColor(self.var_color_value)

        web_safe_color_code = ColorSystem.solve_tone_and_color_name(
                contents_doc=contents_doc,
                tone_and_color_name=self.var_color_value)
        return WebSafeColor(web_safe_color_code)


    def to_fill_obj(self, contents_doc):
        """様々な色名を FillPattern オブジェクトに変換します
        """

        try:

            # 色が指定されていないとき、この関数を呼び出してはいけません
            if not self.var_type:
                raise Exception(f'to_fill_obj: 色が指定されていません')

            # 背景色を［なし］にします。透明（transparent）で上書きするのと同じです
            if self.var_type == VarColor.PAPER_COLOR:
                return ColorSystem.none_pattern_fill

            if self.var_type == VarColor.XL_COLOR_CODE:
                return PatternFill(
                        patternType='solid',
                        fgColor=self.var_color_value)

            # ［auto］は自動で影の色を設定する機能ですが、その機能をオフにしているときは、とりあえず黒色にします
            if self.var_type == VarColor.AUTO:
                web_safe_color_code = ColorSystem.alias_to_web_safe_color_dict(contents_doc)['xlTheme']['xlBlack']
                web_safe_color_obj = WebSafeColor(web_safe_color_code)
                xl_color_name = web_safe_color_obj.to_xl()

                #if not re.match(r'^[0-9a-fA-f]{6}$', xl_color_name): #FIXME
                #    raise ValueError(f'色指定がおかしい {xl_color_name=}')
                # else:
                #     print(f'★ {xl_color_name=}')

                return PatternFill(
                        patternType='solid',
                        fgColor=xl_color_name)

            # ウェブ・セーフ・カラーを、エクセルの引数書式へ変換
            if self.var_type == VarColor.WEB_SAFE_COLOR:
                web_safe_color_obj = WebSafeColor(self.var_color_value)
                return PatternFill(
                        patternType='solid',
                        fgColor=web_safe_color_obj.to_xl())

            if self.var_type == VarColor.TONE_AND_COLOR_NAME:
                tone, color = self.var_color_value.split('.', 2)
                tone = tone.strip()
                color = color.strip()

                if tone in ColorSystem.alias_to_web_safe_color_dict(contents_doc):
                    if color in ColorSystem.alias_to_web_safe_color_dict(contents_doc)[tone]:
                        web_safe_color_code = ColorSystem.alias_to_web_safe_color_dict(contents_doc)[tone][color]
                        web_safe_color_obj = WebSafeColor(web_safe_color_code)
                        return PatternFill(
                                patternType='solid',
                                fgColor=web_safe_color_obj.to_xl())


            print(f'to_fill_obj: 色がない {self.var_color_value=}')
            return ColorSystem.none_pattern_fill

        except:
            print(f'{self.var_color_value=} {self.var_type=}')
            raise
