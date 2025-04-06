import os
import openpyxl as xl
from openpyxl.styles import Font, PatternFill
from openpyxl.styles.alignment import Alignment
from openpyxl.styles.borders import Border, Side
from openpyxl.drawing.image import Image as XlImage
import json

from ...shared_models import Canvas, ColorSystem, Share, VarColor, WebSafeColor, XlFont


#############
# MARK: Ruler
#############
def render_ruler(config_doc, contents_doc, ws):
    """定規の描画
    """

    # 処理しないフラグ
    if 'renderer' in config_doc and (renderer_dict := config_doc['renderer']):
        if 'features' in renderer_dict and (features_dict := renderer_dict['features']):
            if 'ruler' in features_dict and (feature_dict := features_dict['ruler']):
                if 'enabled' in feature_dict:
                    enabled = feature_dict['enabled'] # False 値を取りたい
                    if not enabled:
                        return

    print("🔧　定規の描画")

    HORIZONTAL_RULER_HEIGHT = 1     # 水平定規の縦幅
    VERTICAL_RULER_WIDTH = 2        # 垂直定規の横幅

    # Trellis では、タテ：ヨコ＝３：３ で、１ユニットセルとします。
    # また、上辺、右辺、下辺、左辺に、１セル幅の定規を置きます
    canvas_obj = Canvas.from_dict(contents_doc['canvas'])
    canvas_bounds_obj = canvas_obj.bounds_obj

    # 定規を描画しないケース
    if (
            # ruler 項目がない、 
            'ruler' not in contents_doc or
            # ruler 項目にヌルが設定されている
            (ruler_dict := contents_doc['ruler']) is None or
            # contents_doc.visibule プロパティがない
            'visible' not in ruler_dict or
            # contents_doc.visibule プロパティがヌルか偽だ
            ruler_dict['visible'] in [None, False]):
        return

    # 定規の文字色
    font_list = None
    black_font = Font(color='000000')

    # font_list 作成
    if 'foreground' in ruler_dict and (foreground_dict := ruler_dict['foreground']) is not None:
        if 'varColors' in foreground_dict and (var_color_list := foreground_dict['varColors']) is not None:
            if len(var_color_list) == 0:
                # フォントの色の既定値は黒が１つ
                font_list = [black_font]

            else:
                font_list = [None] * len(var_color_list)
                
                for index, fg_color_text in enumerate(var_color_list):
                    var_color_obj = VarColor(fg_color_text)

                    if fg_color_text == 'paperColor':
                        #font_list[index] = Font(color=None)   # フォントに使うと黒になる
                        raise ValueError(f'foreground.varColors に paperColor を指定してはいけません {index=}')

                    elif (web_safe_color_obj_of_font := var_color_obj.to_web_safe_color_obj(
                            contents_doc=contents_doc)) and web_safe_color_obj_of_font is not None:

                        try:
                            xl_font_obj = XlFont(web_safe_color_code=web_safe_color_obj_of_font.code)
                            font_list[index] = Font(color=xl_font_obj.color_code_for_xl)
                        except:
                            print(f'ERROR: render_ruler: {index=}')
                            raise

        else:
            # フォントの色の既定値は黒が１つ
            font_list = [black_font]

    # 定規の背景色
    pattern_fill_list = None

    # pattern_fill_list 作成
    if 'background' in ruler_dict and (background_dict := ruler_dict['background']) is not None:
        if 'varColors' in background_dict and (var_color_list := background_dict['varColors']) is not None:
            if len(var_color_list) == 0:
                # 背景色の既定値は［塗りつぶし無し］
                pattern_fill_list = [PatternFill(patternType=None)]
            
            else:
                pattern_fill_list = [None] * len(var_color_list)
                
                for index, bg_color_text in enumerate(var_color_list):
                    var_color_obj = VarColor(bg_color_text)

                    if bg_color_text == 'paperColor':
                        pattern_fill_list[index] = PatternFill(patternType=None)

                    elif (web_safe_color_obj := var_color_obj.to_web_safe_color_obj(
                            contents_doc=contents_doc)) and web_safe_color_obj is not None:
                        try:
                            pattern_fill_list[index] = PatternFill(
                                    patternType='solid',
                                    fgColor=web_safe_color_obj.to_xl())
                        except:
                            print(f'ERROR: render_ruler: {index=}')
                            raise

        else:
            # 背景色の既定値は［塗りつぶし無し］
            pattern_fill_list = [PatternFill(patternType=None)]


    center_center_alignment = Alignment(horizontal='center', vertical='center')


    def render_coloring_of_top_edge():
        """定規の着色　＞　上辺
        """
        row_th = canvas_bounds_obj.top_th

        for column_th in range(
                canvas_bounds_obj.left_th + Share.OUT_COUNTS_THAT_CHANGE_INNING,
                canvas_bounds_obj.right_th - Share.OUT_COUNTS_THAT_CHANGE_INNING,
                Share.OUT_COUNTS_THAT_CHANGE_INNING):
            column_letter = xl.utils.get_column_letter(column_th)
            cell = ws[f'{column_letter}{row_th}']

            ruler_number = (column_th - canvas_bounds_obj.left_th) // Share.OUT_COUNTS_THAT_CHANGE_INNING
            cell.fill = pattern_fill_list[ruler_number % len(pattern_fill_list)]


    def render_coloring_of_left_edge():
        """定規の着色　＞　左辺
        """

        # 幅が４アウト未満の場合、左辺のルーラーは描かないものとします（上、右、下、左の辺の定規のセル結合が被ってしまうため、上辺だけ残します）
        if canvas_bounds_obj.width < 4:
            return

        column_th = canvas_bounds_obj.left_th
        column_letter = xl.utils.get_column_letter(column_th)
        shrink = canvas_bounds_obj.height % Share.OUT_COUNTS_THAT_CHANGE_INNING

        for row_th in range(
                canvas_bounds_obj.top_th,
                canvas_bounds_obj.bottom_th - shrink,
                Share.OUT_COUNTS_THAT_CHANGE_INNING):
            cell = ws[f'{column_letter}{row_th}']

            ruler_number = (row_th - canvas_bounds_obj.top_th) // Share.OUT_COUNTS_THAT_CHANGE_INNING
            cell.fill = pattern_fill_list[ruler_number % len(pattern_fill_list)]


    def render_ruler_coloring_of_left_edge_bottom_spacing():
        """左辺の最後の要素が端数のとき、左辺の最後の要素の左上へ着色

                最後の端数の要素に色を塗ってもらいたいから、もう１要素着色しておく
        """
        vertical_remain = canvas_bounds_obj.height % Share.OUT_COUNTS_THAT_CHANGE_INNING
        #print(f'左辺 h_qty={canvas_bounds_obj.height_obj.total_of_out_counts_qty} {shrink=} {vertical_remain=}')

        if vertical_remain != 0:
            column_th = canvas_bounds_obj.left_th
            column_letter = xl.utils.get_column_letter(column_th)
            row_th = canvas_bounds_obj.bottom_th - vertical_remain
            ruler_number = (row_th - canvas_bounds_obj.top_th) // Share.OUT_COUNTS_THAT_CHANGE_INNING
            #print(f"""左辺の最後の要素の左上へ着色 {row_th=} {ruler_number=}""")
            cell = ws[f'{column_letter}{row_th}']

            # 数字も振りたい
            if vertical_remain == 2:
                cell.value = ruler_number
                cell.alignment = center_center_alignment
                cell.font = font_list[ruler_number % len(font_list)]

            cell.fill = pattern_fill_list[ruler_number % len(pattern_fill_list)]


    def render_coloring_of_bottom_edge():
        """定規の着色　＞　下辺
        """

        # 高さが２投球回未満の場合、下辺のルーラーは描かないものとします（上、右、下、左の辺の定規のセル結合が被ってしまうため、上辺だけ残します）
        if canvas_bounds_obj.height < 2:
            return

        row_th = canvas_bounds_obj.bottom_th - 1

        for column_th in range(
                canvas_bounds_obj.left_th + Share.OUT_COUNTS_THAT_CHANGE_INNING,
                canvas_bounds_obj.right_th - Share.OUT_COUNTS_THAT_CHANGE_INNING,
                Share.OUT_COUNTS_THAT_CHANGE_INNING):
            column_letter = xl.utils.get_column_letter(column_th)
            cell = ws[f'{column_letter}{row_th}']
            ruler_number = (column_th - canvas_bounds_obj.left_th) // Share.OUT_COUNTS_THAT_CHANGE_INNING
            cell.fill = pattern_fill_list[ruler_number % len(pattern_fill_list)]


    def render_coloring_of_right_edge():
        """定規の着色　＞　右辺
        """

        # 幅が４アウト未満の場合、右辺のルーラーは描かないものとします（上、右、下、左の辺の定規のセル結合が被ってしまうため、上辺だけ残します）
        if canvas_bounds_obj.width < 4:
            return

        column_th = canvas_bounds_obj.right_th - VERTICAL_RULER_WIDTH
        column_letter = xl.utils.get_column_letter(column_th)
        shrink = canvas_bounds_obj.height % Share.OUT_COUNTS_THAT_CHANGE_INNING

        for row_th in range(
                canvas_bounds_obj.top_th,
                canvas_bounds_obj.bottom_th - shrink,
                Share.OUT_COUNTS_THAT_CHANGE_INNING):
            cell = ws[f'{column_letter}{row_th}']

            ruler_number = (row_th - canvas_bounds_obj.top_th) // Share.OUT_COUNTS_THAT_CHANGE_INNING
            cell.fill = pattern_fill_list[ruler_number % len(pattern_fill_list)]


    def render_ruler_coloring_of_right_edge_bottom_spacing():
        """右辺の最後の要素が端数のとき、右辺の最後の要素の左上へ着色

                最後の端数の要素に色を塗ってもらいたいから、もう１要素着色しておく
        """
        vertical_remain = canvas_bounds_obj.height % Share.OUT_COUNTS_THAT_CHANGE_INNING
        #print(f'右辺 h_qty={canvas_bounds_obj.height} {shrink=} {vertical_remain=}')

        if vertical_remain != 0:
            column_th = canvas_bounds_obj.right_th - VERTICAL_RULER_WIDTH
            column_letter = xl.utils.get_column_letter(column_th)
            row_th = canvas_bounds_obj.bottom_th - vertical_remain
            ruler_number = (row_th - canvas_bounds_obj.top_th) // Share.OUT_COUNTS_THAT_CHANGE_INNING
            #print(f"""右辺の最後の要素の左上へ着色 {row_th=} {ruler_number=}""")
            cell = ws[f'{column_letter}{row_th}']

            # 数字も振りたい
            if vertical_remain == 2:
                cell.value = ruler_number
                cell.alignment = center_center_alignment
                cell.font = font_list[ruler_number % len(font_list)]

            cell.fill = pattern_fill_list[ruler_number % len(pattern_fill_list)]


    def render_ruler_coloring_of_top_left_spacing():
        """定規の着色　＞　左上の１セルの隙間
        """
        column_th = canvas_bounds_obj.left_th + VERTICAL_RULER_WIDTH
        row_th = canvas_bounds_obj.top_th
        ruler_number = (column_th - canvas_bounds_obj.left_th) // Share.OUT_COUNTS_THAT_CHANGE_INNING
        column_letter = xl.utils.get_column_letter(column_th)
        cell = ws[f'{column_letter}{row_th}']
        cell.fill = pattern_fill_list[ruler_number % len(pattern_fill_list)]


    def render_ruler_coloring_right_end_spacing_on_top():
        """定規の着色　＞　上の水平定規の右端の隙間の先頭
        """
        horizontal_remain = canvas_bounds_obj.width % Share.OUT_COUNTS_THAT_CHANGE_INNING
        if horizontal_remain in [1, 2]:
            return

        row_th = canvas_bounds_obj.top_th

        # 何アウト余るか
        spacing = (canvas_bounds_obj.width - VERTICAL_RULER_WIDTH) % Share.OUT_COUNTS_THAT_CHANGE_INNING

        # 隙間の先頭
        column_th = canvas_bounds_obj.right_th - VERTICAL_RULER_WIDTH - spacing
        column_letter = xl.utils.get_column_letter(column_th)

        # 隙間に表示される定規の番号
        ruler_number = column_th // Share.OUT_COUNTS_THAT_CHANGE_INNING

        cell = ws[f'{column_letter}{row_th}']
        cell.fill = pattern_fill_list[ruler_number % len(pattern_fill_list)]


    def render_ruler_coloring_of_bottom_left_spacing():
        """定規の着色　＞　左下の１セルの隙間
        """
        column_th = canvas_bounds_obj.left_th + VERTICAL_RULER_WIDTH
        row_th = canvas_bounds_obj.bottom_th - 1

        ruler_number = (column_th - canvas_bounds_obj.left_th) // Share.OUT_COUNTS_THAT_CHANGE_INNING
        column_letter = xl.utils.get_column_letter(column_th)
        cell = ws[f'{column_letter}{row_th}']
        cell.fill = pattern_fill_list[ruler_number % len(pattern_fill_list)]


    def render_ruler_coloring_right_end_spacing_on_bottom():
        """定規の着色　＞　下の水平定規の右端の隙間の先頭
        """
        horizontal_remain = canvas_bounds_obj.width % Share.OUT_COUNTS_THAT_CHANGE_INNING
        if horizontal_remain in [1, 2]:
            return

        row_th = canvas_bounds_obj.bottom_th - 1

        # 何アウト余るか
        spacing = (canvas_bounds_obj.width - VERTICAL_RULER_WIDTH) % Share.OUT_COUNTS_THAT_CHANGE_INNING

        # 隙間の先頭
        column_th = canvas_bounds_obj.right_th - VERTICAL_RULER_WIDTH - spacing
        column_letter = xl.utils.get_column_letter(column_th)

        # 隙間に表示される定規の番号
        ruler_number = column_th // Share.OUT_COUNTS_THAT_CHANGE_INNING

        cell = ws[f'{column_letter}{row_th}']
        cell.fill = pattern_fill_list[ruler_number % len(pattern_fill_list)]


    def render_ruler_merge_cells_of_top_edge():
        """定規のセル結合　＞　上辺

        横幅が３で割り切れるとき、１投球回は 4th から始まる。２投球回を最終表示にするためには、横幅を 3 シュリンクする
        ■■□[  1 ][  2 ]□■■
        ■■                ■■

        横幅が３で割ると１余るとき、１投球回は 4th から始まる。２投球回を最終表示にするためには、横幅を 4 シュリンクする
        ■■□[  1 ][  2 ]□□■■
        ■■                  ■■

        横幅が３で割ると２余るとき、１投球回は 4th から始まる。２投球回を最終表示にするためには、横幅を 2 シュリンクする
        ■■□[  1 ][  2 ][  3 ]■■
        ■■                    ■■
        """
        skip_left = Share.OUT_COUNTS_THAT_CHANGE_INNING
        horizontal_remain = canvas_bounds_obj.width % Share.OUT_COUNTS_THAT_CHANGE_INNING
        if horizontal_remain == 0:
            shrink_right = 3
        elif horizontal_remain == 1:
            shrink_right = 4
        else:
            shrink_right = 2

        row_th = canvas_bounds_obj.top_th

        for column_th in range(
                canvas_bounds_obj.left_th + skip_left,
                canvas_bounds_obj.right_th - shrink_right,
                Share.OUT_COUNTS_THAT_CHANGE_INNING):
            column_letter = xl.utils.get_column_letter(column_th)
            column_letter2 = xl.utils.get_column_letter(column_th + 2)
            ws.merge_cells(f'{column_letter}{row_th}:{column_letter2}{row_th}')


    def render_ruler_merge_cells_of_left_edge():
        """定規のセル結合　＞　左辺

        縦幅が３で割り切れるとき、１投球回は 1th から始まる。最後の投球回は、端数なしで表示できる
        [  0 ][  1 ][  2 ][  3 ]
        ■                    ■

        縦幅が３で割ると１余るとき、１投球回は 1th から始まる。最後の投球回は、端数１になる
        [  0 ][  1 ][  2 ][  3 ]□
        ■                      ■

        縦幅が３で割ると２余るとき、１投球回は 1th から始まる。最後の投球回は、端数２になる
        [  0 ][  1 ][  2 ][  3 ]□□
        ■                        ■
        """

        # 幅が４アウト未満の場合、左辺のルーラーは描かないものとします（上、右、下、左の辺の定規のセル結合が被ってしまうため、上辺だけ残します）
        if canvas_bounds_obj.width < 4:
            return

        column_th = canvas_bounds_obj.left_th
        column_letter = xl.utils.get_column_letter(column_th)
        column_letter2 = xl.utils.get_column_letter(column_th + 1)

        for row_th in range(
                canvas_bounds_obj.top_th,
                canvas_bounds_obj.bottom_th - Share.OUT_COUNTS_THAT_CHANGE_INNING,
                Share.OUT_COUNTS_THAT_CHANGE_INNING):
            ws.merge_cells(f'{column_letter}{row_th}:{column_letter2}{row_th + 2}')

        # 高さが１イニング未満の場合、最後の要素はありません
        if canvas_bounds_obj.height < Share.OUT_COUNTS_THAT_CHANGE_INNING:
            return
        
        # 最後の要素
        spacing = canvas_bounds_obj.height % Share.OUT_COUNTS_THAT_CHANGE_INNING
        if spacing == 0:
            row_th = canvas_bounds_obj.height * Share.OUT_COUNTS_THAT_CHANGE_INNING + canvas_bounds_obj.top_th - Share.OUT_COUNTS_THAT_CHANGE_INNING
            #print(f'マージセルA h_qty={canvas_bounds_obj.height} {row_th=} {spacing=}')
            try:
                ws.merge_cells(f'{column_letter}{row_th}:{column_letter2}{row_th + 2}')
            except:
                print(f"""★マージセルＡ失敗：
{column_letter=}{row_th=}:{column_letter2=}{row_th + 2=}
{canvas_bounds_obj.height=}
{Share.OUT_COUNTS_THAT_CHANGE_INNING=}
{canvas_bounds_obj.top_th=}
{Share.OUT_COUNTS_THAT_CHANGE_INNING=}
""")
                raise
        elif spacing == 1:
            row_th = canvas_bounds_obj.height * Share.OUT_COUNTS_THAT_CHANGE_INNING + canvas_bounds_obj.top_th
            #print(f'マージセルB {row_th=} {spacing=} {column_letter=} {column_letter2=} {canvas_bounds_obj.height=}')
            ws.merge_cells(f'{column_letter}{row_th}:{column_letter2}{row_th}')
        elif spacing == 2:
            row_th = canvas_bounds_obj.height * Share.OUT_COUNTS_THAT_CHANGE_INNING + canvas_bounds_obj.top_th
            #print(f'マージセルH h_qty={canvas_bounds_obj.height} {row_th=} {spacing=}')
            ws.merge_cells(f'{column_letter}{row_th}:{column_letter2}{row_th + 1}')


    def render_ruler_merge_cells_of_bottom_edge():
        """定規のセル結合　＞　下辺"""

        # 高さが２投球回未満の場合、下辺のルーラーは描かないものとします（上、右、下、左の辺の定規のセル結合が被ってしまうため、上辺だけ残します）
        if canvas_bounds_obj.height < 2:
            return

        skip_left = Share.OUT_COUNTS_THAT_CHANGE_INNING
        horizontal_remain = canvas_bounds_obj.width % Share.OUT_COUNTS_THAT_CHANGE_INNING
        if horizontal_remain == 0:
            shrink_right = 3
        elif horizontal_remain == 1:
            shrink_right = 4
        else:
            shrink_right = 2

        row_th = canvas_bounds_obj.bottom_th - 1

        for column_th in range(
                canvas_bounds_obj.left_th + skip_left,
                canvas_bounds_obj.right_th - shrink_right,
                Share.OUT_COUNTS_THAT_CHANGE_INNING):
            column_letter = xl.utils.get_column_letter(column_th)
            column_letter2 = xl.utils.get_column_letter(column_th + 2)
            ws.merge_cells(f'{column_letter}{row_th}:{column_letter2}{row_th}')


    def render_ruler_merge_cells_of_right_edge():
        """定規のセル結合　＞　右辺"""

        # 幅が４アウト未満の場合、右辺のルーラーは描かないものとします（上、右、下、左の辺の定規のセル結合が被ってしまうため、上辺だけ残します）
        if canvas_bounds_obj.width < 4:
            return

        column_th = canvas_bounds_obj.right_th - VERTICAL_RULER_WIDTH
        column_letter = xl.utils.get_column_letter(column_th)
        column_letter2 = xl.utils.get_column_letter(column_th + 1)

        for row_th in range(
                canvas_bounds_obj.top_th,
                canvas_bounds_obj.bottom_th - Share.OUT_COUNTS_THAT_CHANGE_INNING,
                Share.OUT_COUNTS_THAT_CHANGE_INNING):
            ws.merge_cells(f'{column_letter}{row_th}:{column_letter2}{row_th + 2}')

        # 高さが１イニング未満の場合、最後の要素はありません
        if canvas_bounds_obj.height < Share.OUT_COUNTS_THAT_CHANGE_INNING:
            return
        
        # 最後の要素
        spacing = canvas_bounds_obj.height % Share.OUT_COUNTS_THAT_CHANGE_INNING
        if spacing == 0:
            row_th = canvas_bounds_obj.height * Share.OUT_COUNTS_THAT_CHANGE_INNING + canvas_bounds_obj.top_th - Share.OUT_COUNTS_THAT_CHANGE_INNING
            #print(f'マージセルC h_qty={canvas_bounds_obj.height} {row_th=} {spacing=}')
            ws.merge_cells(f'{column_letter}{row_th}:{column_letter2}{row_th + 2}')
        elif spacing == 1:
            row_th = canvas_bounds_obj.height * Share.OUT_COUNTS_THAT_CHANGE_INNING + canvas_bounds_obj.top_th
            ws.merge_cells(f'{column_letter}{row_th}:{column_letter2}{row_th}')
        elif spacing == 2:
            row_th = canvas_bounds_obj.height * Share.OUT_COUNTS_THAT_CHANGE_INNING + canvas_bounds_obj.top_th
            #print(f'マージセルD h_qty={canvas_bounds_obj.height.total_of_out_counts_qty} {row_th=} {spacing=}')
            ws.merge_cells(f'{column_letter}{row_th}:{column_letter2}{row_th + 1}')


    def render_ruler_merge_cells_right_end_fraction_on_top():
        """上側の水平［定規］の右端の端数のセル結合"""

        # 隙間の幅
        spacing = (canvas_bounds_obj.width - VERTICAL_RULER_WIDTH) % Share.OUT_COUNTS_THAT_CHANGE_INNING
        if spacing == 2:
            column_th = canvas_bounds_obj.right_th - VERTICAL_RULER_WIDTH - spacing
            row_th = canvas_bounds_obj.top_th
            column_letter = xl.utils.get_column_letter(column_th)
            column_letter2 = xl.utils.get_column_letter(column_th + spacing - 1)
            #print(f"""マージセルE {column_th=} {row_th=} {column_letter=} {column_letter2=}""")
            ws.merge_cells(f'{column_letter}{row_th}:{column_letter2}{row_th}')


    def render_ruler_merge_cells_right_end_fraction_on_bottom():
        """下側の水平［定規］の右端の端数のセル結合"""

        # 隙間の幅
        spacing = (canvas_bounds_obj.width - VERTICAL_RULER_WIDTH) % Share.OUT_COUNTS_THAT_CHANGE_INNING
        if spacing == 2:
            column_th = canvas_bounds_obj.right_th - VERTICAL_RULER_WIDTH - spacing
            row_th = canvas_bounds_obj.bottom_th - 1
            column_letter = xl.utils.get_column_letter(column_th)
            column_letter2 = xl.utils.get_column_letter(column_th + spacing - 1)
            #print(f"""マージセルF {column_th=} {row_th=} {column_letter=} {column_letter2=}""")
            ws.merge_cells(f'{column_letter}{row_th}:{column_letter2}{row_th}')


    # 定規上のテキスト表示
    __print_all_texts(
            ws=ws,
            vertical_ruler_width=VERTICAL_RULER_WIDTH,
            horizontal_ruler_height=HORIZONTAL_RULER_HEIGHT,
            font_list=font_list,
            center_center_alignment=center_center_alignment,
            canvas_bounds_obj=canvas_bounds_obj)


    # 定規の着色　＞　上辺
    render_coloring_of_top_edge()
    # 定規の着色　＞　左辺
    render_coloring_of_left_edge()
    # 定規の着色　＞　下辺
    render_coloring_of_bottom_edge()
    # 定規の着色　＞　右辺
    render_coloring_of_right_edge()


    # 左辺の最後の要素が端数のとき、左辺の最後の要素の左上へ着色
    render_ruler_coloring_of_left_edge_bottom_spacing()

    # 右辺の最後の要素が端数のとき、右辺の最後の要素の左上へ着色
    render_ruler_coloring_of_right_edge_bottom_spacing()

    # NOTE 上下の辺の両端の端数の処理

    # 定規の着色　＞　左上の１セルの隙間
    render_ruler_coloring_of_top_left_spacing()

    # 定規の着色　＞　上の水平定規の右端の隙間の先頭
    render_ruler_coloring_right_end_spacing_on_top()

    # 定規の着色　＞　左下の１セルの隙間
    render_ruler_coloring_of_bottom_left_spacing()

    # 定規の着色　＞　下の水平定規の右端の隙間の先頭
    render_ruler_coloring_right_end_spacing_on_bottom()

    # NOTE セル結合すると read only セルになるから、セル結合は、セルを編集が終わったあとで行う

    # 定規のセル結合　＞　上辺
    render_ruler_merge_cells_of_top_edge()

    # 定規のセル結合　＞　左辺
    render_ruler_merge_cells_of_left_edge()

    # 定規のセル結合　＞　下辺
    render_ruler_merge_cells_of_bottom_edge()

    # 定規のセル結合　＞　右辺
    render_ruler_merge_cells_of_right_edge()

    # 上側の水平［定規］の右端の端数のセル結合
    render_ruler_merge_cells_right_end_fraction_on_top()

    # 下側の水平［定規］の右端の端数のセル結合
    render_ruler_merge_cells_right_end_fraction_on_bottom()


def __print_all_texts(ws, vertical_ruler_width, horizontal_ruler_height, font_list, center_center_alignment, canvas_bounds_obj):
    """定規上のテキスト表示
    """


    def print_all_texts_on_top_edge():
        """定規の採番　＞　上辺

                横幅が３で割り切れるとき、１投球回は 4th から始まる。２投球回を最終表示にするためには、横幅を 3 シュリンクする
                ■■□[  1 ][  2 ]□■■
                ■■                ■■

                横幅が３で割ると１余るとき、１投球回は 4th から始まる。２投球回を最終表示にするためには、横幅を 4 シュリンクする
                ■■□[  1 ][  2 ]□□■■
                ■■                  ■■

                横幅が３で割ると２余るとき、１投球回は 4th から始まる。２投球回を最終表示にするためには、横幅を 2 シュリンクする
                ■■□[  1 ][  2 ][  3 ]■■
                ■■                    ■■
        """
        row_th = canvas_bounds_obj.top_th

        for column_th in range(
                canvas_bounds_obj.left_th + Share.OUT_COUNTS_THAT_CHANGE_INNING,
                canvas_bounds_obj.right_th - Share.OUT_COUNTS_THAT_CHANGE_INNING,
                Share.OUT_COUNTS_THAT_CHANGE_INNING):
            column_letter = xl.utils.get_column_letter(column_th)
            cell = ws[f'{column_letter}{row_th}']

            # 1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12,
            # -------- -------- -------- -----------
            # dark      light    dark     light
            #
            # - 1 する
            #
            # 0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11,
            # -------- -------- -------- ----------
            # dark     light    dark     light
            #
            # 3 で割って端数を切り捨て
            #
            # 0, 0, 0, 1, 1, 1, 2, 2, 2, 3, 3, 3,
            # -------- -------- -------- --------
            # dark     light    dark     light
            #
            # 2 で割った余り
            #
            # 0, 0, 0, 1, 1, 1, 0, 0, 0, 1, 1, 1,
            # -------- -------- -------- --------
            # dark     light    dark     light
            #
            ruler_number = (column_th - canvas_bounds_obj.left_th) // Share.OUT_COUNTS_THAT_CHANGE_INNING
            cell.value = ruler_number
            cell.alignment = center_center_alignment
            cell.font = font_list[ruler_number % len(font_list)]


    def print_all_texts_on_left_edge():
        """定規の採番　＞　左辺

        縦幅が３で割り切れるとき、１投球回は 1th から始まる。最後の投球回は、端数なしで表示できる
        [  0 ][  1 ][  2 ][  3 ]
        ■                    ■

        縦幅が３で割ると１余るとき、１投球回は 1th から始まる。最後の投球回は、端数１になる
        [  0 ][  1 ][  2 ][  3 ]□
        ■                      ■

        縦幅が３で割ると２余るとき、１投球回は 1th から始まる。最後の投球回は、端数２になる
        [  0 ][  1 ][  2 ][  3 ]□□
        ■                        ■
        """

        # 幅が４アウト未満の場合、左辺のルーラーは描かないものとします（上、右、下、左の辺の定規のセル結合が被ってしまうため、上辺だけ残します）
        if canvas_bounds_obj.width < 4:
            return

        column_th = canvas_bounds_obj.left_th
        column_letter = xl.utils.get_column_letter(column_th)
        shrink = canvas_bounds_obj.height % Share.OUT_COUNTS_THAT_CHANGE_INNING

        for row_th in range(
                canvas_bounds_obj.top_th,
                canvas_bounds_obj.bottom_th - shrink,
                Share.OUT_COUNTS_THAT_CHANGE_INNING):
            cell = ws[f'{column_letter}{row_th}']

            ruler_number = (row_th - canvas_bounds_obj.top_th) // Share.OUT_COUNTS_THAT_CHANGE_INNING
            cell.value = ruler_number
            cell.alignment = center_center_alignment
            cell.font = font_list[ruler_number % len(font_list)]


    def print_all_texts_on_bottom_edge():
        """定規の採番　＞　下辺
        """

        # 高さが２投球回未満の場合、下辺のルーラーは描かないものとします（上、右、下、左の辺の定規のセル結合が被ってしまうため、上辺だけ残します）
        if canvas_bounds_obj.height < 2:
            return

        row_th = canvas_bounds_obj.bottom_th - horizontal_ruler_height

        for column_th in range(
                canvas_bounds_obj.left_th + Share.OUT_COUNTS_THAT_CHANGE_INNING,
                canvas_bounds_obj.right_th - Share.OUT_COUNTS_THAT_CHANGE_INNING,
                Share.OUT_COUNTS_THAT_CHANGE_INNING):
            column_letter = xl.utils.get_column_letter(column_th)
            cell = ws[f'{column_letter}{row_th}']
            ruler_number = (column_th - canvas_bounds_obj.left_th) // Share.OUT_COUNTS_THAT_CHANGE_INNING
            cell.value = ruler_number
            cell.alignment = center_center_alignment
            cell.font = font_list[ruler_number % len(font_list)]


    def print_all_texts_on_right_edge():
        """定規の採番　＞　右辺
        """

        # 幅が４アウト未満の場合、右辺のルーラーは描かないものとします（上、右、下、左の辺の定規のセル結合が被ってしまうため、上辺だけ残します）
        if canvas_bounds_obj.width < 4:
            return

        column_th = canvas_bounds_obj.right_th - vertical_ruler_width
        column_letter = xl.utils.get_column_letter(column_th)
        shrink = canvas_bounds_obj.height % Share.OUT_COUNTS_THAT_CHANGE_INNING

        for row_th in range(
                canvas_bounds_obj.top_th,
                canvas_bounds_obj.bottom_th - shrink,
                Share.OUT_COUNTS_THAT_CHANGE_INNING):
            cell = ws[f'{column_letter}{row_th}']

            ruler_number = (row_th - canvas_bounds_obj.top_th) // Share.OUT_COUNTS_THAT_CHANGE_INNING
            cell.value = ruler_number
            cell.alignment = center_center_alignment
            cell.font = font_list[ruler_number % len(font_list)]


    # 定規の採番　＞　上辺
    print_all_texts_on_top_edge()
    # 定規の採番　＞　左辺
    print_all_texts_on_left_edge()
    # 定規の採番　＞　下辺
    print_all_texts_on_bottom_edge()
    # 定規の採番　＞　右辺
    print_all_texts_on_right_edge()
