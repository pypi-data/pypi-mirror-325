import openpyxl as xl
import os

from openpyxl.drawing.image import Image as XlImage
from openpyxl.styles import Font, PatternFill
from openpyxl.styles.alignment import Alignment
from openpyxl.styles.borders import Border, Side

from ..shared_models import ColorSystem, Share, VarColor, WebSafeColor


def fill_rectangle(ws, contents_doc, column_th, row_th, columns, rows, color):
    """矩形を塗りつぶします
    """
    var_color_obj = VarColor(color)
    color_type = var_color_obj.var_type

    if not color_type:
        return


    if color_type != VarColor.DARKNESS:
        fill_obj = var_color_obj.to_fill_obj(
                contents_doc=contents_doc)


    # 横へ
    for cur_column_th in range(column_th, column_th + columns):
        column_letter = xl.utils.get_column_letter(cur_column_th)

        # 縦へ
        for cur_row_th in range(row_th, row_th + rows):
            cell = ws[f'{column_letter}{cur_row_th}']

            if color_type == VarColor.DARKNESS:
                # TODO セルの背景色を取得
                # TODO ウェブ・セーフ・カラーに変換
                # TODO さらに影の色に変換
                # TODO 指定によりそれを複数回
                fill_obj = var_color_obj.to_fill_obj(
                        contents_doc=contents_doc)
            

            cell.fill = fill_obj


def draw_xl_border_on_rectangle(ws, contents_doc, xl_border_dict, column_th, row_th, columns, rows):
    """境界線の描画
    """
    top_side = None
    right_side = None
    bottom_side = None
    left_side = None

    # 罫線の style の種類
    # 📖 [openpyxl.styles.borders module](https://openpyxl.readthedocs.io/en/3.1/api/openpyxl.styles.borders.html)
    # ‘mediumDashed’, ‘mediumDashDotDot’, ‘dashDot’, ‘dashed’, ‘slantDashDot’, ‘dashDotDot’, ‘thick’, ‘thin’, ‘dotted’, ‘double’, ‘medium’, ‘hair’, ‘mediumDashDot’

    if 'top' in xl_border_dict and (top_dict := xl_border_dict['top']):
        web_safe_color_code = None
        style = None

        if 'color' in top_dict and (color := top_dict['color']):
            var_color_obj = VarColor(color)
            web_safe_color_obj = var_color_obj.to_web_safe_color_obj(
                    contents_doc=contents_doc)

        if 'xlStyle' in top_dict and (style := top_dict['xlStyle']):
            pass

        try:
            top_side = Side(style=style, color=web_safe_color_obj.to_xl())
        except:
            print(f'draw_xl_border_on_rectangle: いずれかが、未対応の指定： {style=} {web_safe_color_obj.code=}')


    if 'right' in xl_border_dict and (right_dict := xl_border_dict['right']):
        style = None

        if 'color' in right_dict and (color := right_dict['color']):
            var_color_obj = VarColor(color)
            web_safe_color_obj = var_color_obj.to_web_safe_color_obj(
                    contents_doc=contents_doc)

        if 'xlStyle' in right_dict and (style := right_dict['xlStyle']):
            pass

        try:
            right_side = Side(style=style, color=web_safe_color_obj.to_xl())
        except:
            print(f'draw_xl_border_on_rectangle: スタイルか、ウェブセーフカラーのいずれかが、未対応の指定： {style=}')


    if 'bottom' in xl_border_dict and (bottom_dict := xl_border_dict['bottom']):
        style = None

        if 'color' in bottom_dict and (color := bottom_dict['color']):
            var_color_obj = VarColor(color)
            web_safe_color_obj = var_color_obj.to_web_safe_color_obj(
                    contents_doc=contents_doc)

        if 'xlStyle' in bottom_dict and (style := bottom_dict['xlStyle']):
            pass

        try:
            bottom_side = Side(style=style, color=web_safe_color_obj.to_xl())
        except:
            print(f'draw_xl_border_on_rectangle: スタイルか、ウェブセーフカラーのいずれかが、未対応の指定： {style=}')


    if 'left' in xl_border_dict and (left_dict := xl_border_dict['left']):
        style = None

        if 'color' in left_dict and (color := left_dict['color']):
            var_color_obj = VarColor(color)
            web_safe_color_obj = var_color_obj.to_web_safe_color_obj(
                    contents_doc=contents_doc)

        if 'xlStyle' in left_dict and (style := left_dict['xlStyle']):
            pass

        try:
            left_side = Side(style=style, color=web_safe_color_obj.to_xl())
        except:
            print(f'draw_xl_border_on_rectangle: スタイルか、ウェブセーフカラーのいずれかが、未対応の指定： {style=}')


    # TODO 厚みが１のケースや、角は、２辺に線を引く

    
    top_border = Border(top=top_side)           # 上辺
    right_border = Border(right=right_side)     # 右辺
    bottom_border = Border(bottom=bottom_side)  # 下辺
    left_border = Border(left=left_side)        # 左辺

    # 水平方向
    if rows == 0 or rows == 1:
        if rows == 0:
            # 上辺だけ引く
            horizontal_border = Border(top=top_side)
        else:
            # 上辺と下辺の両方を引く
            horizontal_border = Border(top=top_side, bottom=bottom_side)

        # （角を除く）横へ
        for cur_column_th in range(column_th + 1, column_th + columns - 1):
            column_letter = xl.utils.get_column_letter(cur_column_th)
            cell = ws[f'{column_letter}{row_th}']
            cell.border = horizontal_border

    # 上辺を引くのと、下辺を引くのとがある
    else:
        top_border = Border(top=top_side)
        bottom_border = Border(bottom=bottom_side)

        # （角を除く）横へ
        for cur_column_th in range(column_th + 1, column_th + columns - 1):
            column_letter = xl.utils.get_column_letter(cur_column_th)

            cell = ws[f'{column_letter}{row_th}']
            cell.border = top_border

            cell = ws[f'{column_letter}{row_th + rows - 1}']
            cell.border = bottom_border


    # 垂直方向
    if columns == 0 or columns == 1:
        if columns == 0:
            # 左辺だけ引く
            vertical_border = Border(left=left_side)
        else:
            # 左辺と右辺の両方を引く
            vertical_border = Border(left=left_side, right=right_side)

        # （角を除く）縦へ
        for cur_row_th in range(row_th + 1, row_th + rows - 1):
            column_letter = xl.utils.get_column_letter(columns)
            cell = ws[f'{column_letter}{cur_row_th}']
            cell.border = vertical_border

    # 左辺を引くのと、右辺を引くのとがある
    else:
        left_border = Border(left=left_side)
        right_border = Border(right=right_side)

        # （角を除く）縦へ
        for cur_row_th in range(row_th + 1, row_th + rows - 1):
            column_letter = xl.utils.get_column_letter(column_th)
            cell = ws[f'{column_letter}{cur_row_th}']
            cell.border = left_border

            column_letter = xl.utils.get_column_letter(column_th + columns - 1)
            cell = ws[f'{column_letter}{cur_row_th}']
            cell.border = right_border


    # 左上隅
    if 1 < columns and 1 < rows:
        column_letter = xl.utils.get_column_letter(column_th)
        cell = ws[f'{column_letter}{row_th}']
        cell.border = Border(top=top_side, left=left_side)

    # 右上隅
    if 1 < columns and 1 < rows:
        column_letter = xl.utils.get_column_letter(column_th + columns - 1)
        cell = ws[f'{column_letter}{row_th}']
        cell.border = Border(top=top_side, right=right_side)

    # 左下隅
    if 1 < columns and 1 < rows:
        column_letter = xl.utils.get_column_letter(column_th)
        cell = ws[f'{column_letter}{row_th + rows - 1}']
        cell.border = Border(left=left_side, bottom=bottom_side)

    # 右下隅
    if 1 < columns and 1 < rows:
        column_letter = xl.utils.get_column_letter(column_th + columns - 1)
        cell = ws[f'{column_letter}{row_th + rows - 1}']
        cell.border = Border(right=right_side, bottom=bottom_side)

    # 四方
    if columns == 1 and rows == 1:
        column_letter = xl.utils.get_column_letter(column_th)
        cell = ws[f'{column_letter}{row_th}']
        cell.border = Border(top=top_side, right=right_side, bottom=bottom_side, left=left_side)


def print_text(ws, location_obj, text, xl_alignment_obj, xl_font_obj):
    """テキスト描画
    """

    # テキストの位置
    column_th = location_obj.x_obj.total_of_out_counts_th
    row_th = location_obj.y_obj.total_of_out_counts_th

    # テキスト設定
    column_letter = xl.utils.get_column_letter(column_th)
    cell = ws[f'{column_letter}{row_th}']
    cell.value = text

    # フォント設定
    if xl_font_obj:
        cell.font = Font(color=xl_font_obj.color_code_for_xl)

    # テキストの位置揃え
    if xl_alignment_obj:
        cell.alignment = Alignment(
                horizontal=xl_alignment_obj.xlHorizontal,
                vertical=xl_alignment_obj.xlVertical)


def draw_rectangle(ws, column_th, row_th, columns, rows):
    """矩形の枠線の描画
    """

    # 赤はデバッグ用
    red_side = Side(style='thick', color='FF0000')
    black_side = Side(style='thick', color='000000')

    red_top_border = Border(top=red_side)
    red_top_right_border = Border(top=red_side, right=red_side)
    red_right_border = Border(right=red_side)
    red_bottom_right_border = Border(bottom=red_side, right=red_side)
    red_bottom_border = Border(bottom=red_side)
    red_bottom_left_border = Border(bottom=red_side, left=red_side)
    red_left_border = Border(left=red_side)
    red_top_left_border = Border(top=red_side, left=red_side)

    black_top_border = Border(top=black_side)
    black_top_right_border = Border(top=black_side, right=black_side)
    black_right_border = Border(right=black_side)
    black_bottom_right_border = Border(bottom=black_side, right=black_side)
    black_bottom_border = Border(bottom=black_side)
    black_bottom_left_border = Border(bottom=black_side, left=black_side)
    black_left_border = Border(left=black_side)
    black_top_left_border = Border(top=black_side, left=black_side)

    # 罫線で四角を作る　＞　左上
    cur_column_th = column_th
    column_letter = xl.utils.get_column_letter(cur_column_th)
    cur_row_th = row_th
    cell = ws[f'{column_letter}{cur_row_th}']
    cell.border = black_top_left_border

    # 罫線で四角を作る　＞　上辺
    for cur_column_th in range(column_th + 1, column_th + columns - 1):
        column_letter = xl.utils.get_column_letter(cur_column_th)
        cell = ws[f'{column_letter}{cur_row_th}']
        cell.border = black_top_border

    # 罫線で四角を作る　＞　右上
    cur_column_th = column_th + columns - 1
    column_letter = xl.utils.get_column_letter(cur_column_th)
    cell = ws[f'{column_letter}{cur_row_th}']
    cell.border = black_top_right_border

    # 罫線で四角を作る　＞　左辺
    cur_column_th = column_th
    for cur_row_th in range(row_th + 1, row_th + rows - 1):
        column_letter = xl.utils.get_column_letter(cur_column_th)
        cell = ws[f'{column_letter}{cur_row_th}']
        cell.border = black_left_border

    # 罫線で四角を作る　＞　左下
    cur_row_th = row_th + rows - 1
    cell = ws[f'{column_letter}{cur_row_th}']
    cell.border = black_bottom_left_border

    # 罫線で四角を作る　＞　下辺
    for cur_column_th in range(column_th + 1, column_th + columns - 1):
        column_letter = xl.utils.get_column_letter(cur_column_th)
        cell = ws[f'{column_letter}{cur_row_th}']
        cell.border = black_bottom_border

    # 罫線で四角を作る　＞　右下
    cur_column_th = column_th + columns - 1
    column_letter = xl.utils.get_column_letter(cur_column_th)
    cell = ws[f'{column_letter}{cur_row_th}']
    cell.border = black_bottom_right_border

    # 罫線で四角を作る　＞　右辺
    for cur_row_th in range(row_th + 1, row_th + rows - 1):
        cell = ws[f'{column_letter}{cur_row_th}']
        cell.border = black_right_border


def render_paper_strip(ws, contents_doc, paper_strip, column_th, row_th, columns, rows):
    """短冊１行の描画
    """

    # 柱のヘッダーの背景色
    if 'background' in paper_strip and (background_dict := paper_strip['background']):
        if 'varColor' in background_dict and (bg_color := background_dict['varColor']):
            # 矩形を塗りつぶす
            fill_rectangle(
                    ws=ws,
                    contents_doc=contents_doc,
                    column_th=column_th,
                    row_th=row_th,
                    columns=columns,
                    rows=1 * Share.OUT_COUNTS_THAT_CHANGE_INNING,   # １行分
                    color=bg_color)


    # インデント
    if 'indent' in paper_strip:
        indent = paper_strip['indent']
    else:
        indent = 0

    # アイコン（があれば画像をワークシートのセルに挿入）
    if 'icon' in paper_strip:
        image_basename = paper_strip['icon']  # 例： 'white_game_object.png'

        cur_column_th = column_th + (indent * Share.OUT_COUNTS_THAT_CHANGE_INNING)
        column_letter = xl.utils.get_column_letter(cur_column_th)
        #
        # NOTE 元の画像サイズで貼り付けられるわけではないの、何でだろう？ 60x60pixels の画像にしておくと、90x90pixels のセルに合う？
        #
        # TODO 📖 [PythonでExcelファイルに画像を挿入する/列の幅を調整する](https://qiita.com/kaba_san/items/b231a41891ebc240efc7)
        # 難しい
        #
        try:
            ws.add_image(XlImage(os.path.join('./assets/icons', image_basename)), f"{column_letter}{row_th}")
        except FileNotFoundError as e:
            print(f'FileNotFoundError {e=} {image_basename=}')


    # テキスト（があれば）
    if 'text0' in paper_strip:
        text = paper_strip['text0']

        # 左に１マス分のアイコンを置く前提
        icon_columns = Share.OUT_COUNTS_THAT_CHANGE_INNING
        cur_column_th = column_th + icon_columns + (indent * Share.OUT_COUNTS_THAT_CHANGE_INNING)
        column_letter = xl.utils.get_column_letter(cur_column_th)
        cell = ws[f'{column_letter}{row_th}']
        cell.value = text

    if 'text1' in paper_strip:
        text = paper_strip['text1']

        # 左に１マス分のアイコンを置く前提
        icon_columns = Share.OUT_COUNTS_THAT_CHANGE_INNING
        cur_column_th = column_th + icon_columns + (indent * Share.OUT_COUNTS_THAT_CHANGE_INNING)
        column_letter = xl.utils.get_column_letter(cur_column_th)
        cell = ws[f'{column_letter}{row_th + 1}']
        cell.value = text

    if 'text3' in paper_strip:
        text = paper_strip['text2']

        # 左に１マス分のアイコンを置く前提
        icon_columns = Share.OUT_COUNTS_THAT_CHANGE_INNING
        cur_column_th = column_th + icon_columns + (indent * Share.OUT_COUNTS_THAT_CHANGE_INNING)
        column_letter = xl.utils.get_column_letter(cur_column_th)
        cell = ws[f'{column_letter}{row_th + 2}']
        cell.value = text


def fill_start_terminal(ws, column_th, row_th):
    """始端を描きます
    """
    # ドット絵を描きます
    fill_pixel_art(
            ws=ws,
            column_th=column_th,
            row_th=row_th,
            columns=9,
            rows=9,
            pixels=[
                [1, 1, 1, 1, 1, 1, 1, 1, 1],
                [1, 1, 1, 0, 0, 0, 1, 1, 1],
                [1, 0, 0, 1, 1, 1, 0, 0, 1],
                [1, 1, 0, 1, 1, 1, 1, 0, 1],
                [1, 1, 1, 0, 0, 0, 1, 1, 1],
                [1, 0, 1, 1, 1, 1, 0, 1, 1],
                [1, 0, 0, 1, 1, 1, 0, 0, 1],
                [1, 1, 1, 0, 0, 0, 1, 1, 1],
                [1, 1, 1, 1, 1, 1, 1, 1, 1],
            ])


def fill_end_terminal(ws, column_th, row_th):
    """終端を描きます
    """
    # ドット絵を描きます
    fill_pixel_art(
            ws=ws,
            column_th=column_th,
            row_th=row_th,
            columns=9,
            rows=9,
            pixels=[
                [1, 1, 1, 1, 1, 1, 1, 1, 1],
                [1, 0, 0, 0, 0, 0, 0, 0, 1],
                [1, 0, 1, 1, 1, 1, 1, 1, 1],
                [1, 0, 1, 1, 1, 1, 1, 1, 1],
                [1, 0, 0, 0, 0, 0, 0, 0, 1],
                [1, 0, 1, 1, 1, 1, 1, 1, 1],
                [1, 0, 1, 1, 1, 1, 1, 1, 1],
                [1, 0, 0, 0, 0, 0, 0, 0, 1],
                [1, 1, 1, 1, 1, 1, 1, 1, 1],
            ])


def fill_pixel_art(ws, column_th, row_th, columns, rows, pixels):
    """ドット絵を描きます
    """
    # 背景色
    mat_black = PatternFill(patternType='solid', fgColor='080808')
    mat_white = PatternFill(patternType='solid', fgColor='E8E8E8')

    # 横へ
    for cur_column_th in range(column_th, column_th + columns):
        for cur_row_th in range(row_th, row_th + rows):
            column_letter = xl.utils.get_column_letter(cur_column_th)
            cell = ws[f'{column_letter}{cur_row_th}']

            pixel = pixels[cur_row_th - row_th][cur_column_th - column_th]
            if pixel == 1:
                cell.fill = mat_black
            else:
                cell.fill = mat_white
