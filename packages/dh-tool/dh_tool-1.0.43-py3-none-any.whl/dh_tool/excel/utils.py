from openpyxl.utils import get_column_letter
import pandas as pd


def map_column_names_to_letters(worksheet, width_map):
    """
    DataFrame의 컬럼 이름 또는 엑셀 열 문자(A, B, C)를 자동 매핑하여 열 너비 설정
    """
    # 엑셀 시트의 헤더 가져오기
    headers = [cell.value for cell in next(worksheet.iter_rows(min_row=1, max_row=1))]

    # 최종 매핑 결과 저장
    col_letter_map = {}

    for key, width in width_map.items():
        if key in headers:
            # ✅ 컬럼 이름을 엑셀 열 문자로 변환
            col_idx = headers.index(key) + 1
            col_letter = get_column_letter(col_idx)
            col_letter_map[col_letter] = width
        else:
            print(f"컬럼 '{key}'을 찾을 수 없습니다.")

    return col_letter_map


def get_column_indices_by_condition(worksheet, condition):
    """
    조건을 만족하는 열의 인덱스를 반환
    - condition: 각 열의 데이터 리스트를 받아 True/False를 반환하는 함수
    """
    headers = [cell.value for cell in next(worksheet.iter_rows(min_row=1, max_row=1))]
    indices = []

    for idx, col in enumerate(worksheet.iter_cols(min_row=2), start=1):
        col_data = [cell.value for cell in col]
        if condition(col_data):
            indices.append(idx)

    return indices


def get_cell_addresses(df, condition):
    """
    DataFrame과 조건을 받아 조건을 만족하는 셀의 주소(A1, B2 등)를 반환
    - df: pandas DataFrame
    - condition: 불리언 Series 또는 DataFrame
    """
    cells = []

    # ✅ 1. Series인 경우 (특정 컬럼에만 조건 적용)
    if isinstance(condition, pd.Series):
        col_name = condition.name  # 조건이 적용된 컬럼 이름
        col_idx = df.columns.get_loc(col_name)  # 해당 컬럼의 인덱스 (0부터 시작)
        col_letter = chr(65 + col_idx)  # A, B, C ... (엑셀 열 이름)

        # 조건이 True인 경우 해당 셀 주소 저장
        for row_idx, match in condition.items():
            if match:
                excel_row = row_idx + 2  # 헤더가 1행, 데이터는 2행부터 시작
                cell_ref = f"{col_letter}{excel_row}"  # 셀 주소 (예: B2)
                cells.append(cell_ref)

    # ✅ 2. DataFrame인 경우 (여러 컬럼에 조건 적용)
    elif isinstance(condition, pd.DataFrame):
        for row_idx, row in condition.iterrows():
            for col_idx, match in enumerate(row):
                if match:
                    col_letter = chr(65 + col_idx)
                    excel_row = row_idx + 2
                    cell_ref = f"{col_letter}{excel_row}"
                    cells.append(cell_ref)

    else:
        raise ValueError("Condition must be a Series or DataFrame")

    return cells


def find_columns_with_nulls(worksheet):
    """
    결측치가 있는 컬럼 반환
    """
    headers = [cell.value for cell in next(worksheet.iter_rows(min_row=1, max_row=1))]
    null_columns = []

    for idx, col in enumerate(worksheet.iter_cols(min_row=2), start=1):
        col_data = [cell.value for cell in col]
        if any(pd.isnull(value) for value in col_data):
            null_columns.append(headers[idx - 1])

    return null_columns


def find_columns_by_type(worksheet, data_type):
    """
    특정 데이터 타입(int, str 등)을 가진 열 찾기
    """
    headers = [cell.value for cell in next(worksheet.iter_rows(min_row=1, max_row=1))]
    type_columns = []

    for idx, col in enumerate(worksheet.iter_cols(min_row=2), start=1):
        col_data = [cell.value for cell in col]
        if all(isinstance(value, data_type) or pd.isnull(value) for value in col_data):
            type_columns.append(headers[idx - 1])

    return type_columns
