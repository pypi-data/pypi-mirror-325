from openpyxl.styles import Alignment, PatternFill, Font, Border, Side
from openpyxl.utils import get_column_letter

from .utils import map_column_names_to_letters

COLOR_MAP = {
    "black": "000000",
    "white": "FFFFFF",
    "red": "FF0000",
    "green": "00FF00",
    "blue": "0000FF",
    "yellow": "FFFF00",
    "cyan": "00FFFF",
    "magenta": "FF00FF",
    "gray": "808080",
    "orange": "FFA500",
    "purple": "800080",
    "pink": "FFC0CB",
    "brown": "A52A2A",
    "gold": "FFD700",
    "silver": "C0C0C0",
}


class Style:

    @staticmethod
    def apply_auto_wrap(worksheet):
        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
            for cell in row:
                if cell.value:
                    cell.alignment = Alignment(wrap_text=True, vertical="center")

    @staticmethod
    def freeze_first_row(worksheet):
        worksheet.freeze_panes = worksheet["A2"]

    @staticmethod
    def set_column_width(worksheet, width_map):
        """컬럼 이름 또는 엑셀 열 문자로 열 너비 설정"""
        col_letter_map = map_column_names_to_letters(worksheet, width_map)
        for col, width in col_letter_map.items():
            worksheet.column_dimensions[col].width = width

    @staticmethod
    def auto_adjust_column_widths(worksheet):
        """데이터에 맞게 자동으로 열 너비 조정"""
        for col in worksheet.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            adjusted_width = max_length + 2
            worksheet.column_dimensions[col_letter].width = adjusted_width

    @staticmethod
    def set_font(worksheet, font_name="Arial", font_size=12, bold=False, italic=False):
        font = Font(name=font_name, size=font_size, bold=bold, italic=italic)
        for row in worksheet.iter_rows(min_row=1, max_row=1):
            for cell in row:
                cell.font = font

    @staticmethod
    def apply_border(worksheet, border_style="thin"):
        border = Border(
            left=Side(style=border_style),
            right=Side(style=border_style),
            top=Side(style=border_style),
            bottom=Side(style=border_style),
        )
        for row in worksheet.iter_rows():
            for cell in row:
                cell.border = border

    @staticmethod
    def apply_color(worksheet, color):
        """
        셀 배경색 적용
        - 16진수 색상 코드("FFFF00") 또는 색상 이름("red") 지원
        """
        # ✅ 색상 이름을 16진수로 변환
        if color.lower() in COLOR_MAP:
            color = COLOR_MAP[color.lower()]

        # ✅ 16진수 형식 보정
        if not color.startswith("#") and len(color) == 6:
            color = f"FF{color}"  # openpyxl은 ARGB 포맷을 사용

        fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        for row in worksheet.iter_rows():
            for cell in row:
                if cell.value:
                    cell.fill = fill
