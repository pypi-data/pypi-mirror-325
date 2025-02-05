import copy
from functools import wraps

import pandas as pd
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook

from .sheet import Sheet


def transactional(method):
    """트랜잭션 데코레이터: 에러 발생 시 자동 롤백"""

    @wraps(method)
    def wrapper(self, *args, **kwargs):
        if not hasattr(self, "_backup") or self._backup is None:
            self._backup = copy.deepcopy(self.workbook)  # ✅ 백업 생성
            print(f"Transaction started for {method.__name__}")
        try:
            result = method(self, *args, **kwargs)  # 메서드 실행
            return result
        except Exception as e:
            self.rollback()  # ❌ 에러 발생 시 롤백
            print(f"Error in {method.__name__}: {e}")
            raise e

    return wrapper


class ExcelCore:
    def __init__(self, filename=None, remove_default_sheet=True):
        self.workbook = Workbook()
        self.filename = filename or "untitled.xlsx"
        self.active_sheet = None
        self._backup = None  # ✅ 백업 저장소

        if remove_default_sheet:
            default_sheet = self.workbook.active
            self.workbook.remove(default_sheet)

    @property
    def sheet_names(self):
        return self.workbook.sheetnames

    @property
    def df(self):
        return self.get_dataframe(self.workbook.active.title)

    @property
    def activate_sheet_name(self):
        return self.active_sheet.worksheet.title if self.active_sheet else None

    def get_dataframe(self, sheet_name=None):
        """워크시트의 데이터를 DataFrame으로 변환"""
        sheet = self.workbook[sheet_name] if sheet_name else self.active_sheet
        if not sheet:
            raise ValueError("No active sheet selected.")

        data = list(sheet.values)
        if not data:
            return pd.DataFrame()

        headers = data[0]
        rows = data[1:]
        return pd.DataFrame(rows, columns=headers)

    def rollback(self):
        """롤백: 백업으로 복원"""
        if self._backup:
            self.workbook = self._backup
            self.active_sheet = None
            self._backup = None
            print("Transaction rolled back.")
        return self

    def commit(self):
        """커밋: 백업 삭제"""
        self._backup = None
        print("Transaction committed.")
        return self

    @transactional
    def write(self, sheet_name, data):
        if sheet_name not in self.workbook.sheetnames:
            self.workbook.create_sheet(sheet_name)
        sheet = Sheet(self.workbook[sheet_name])
        sheet.write(data)
        self.active_sheet = sheet
        return self

    @transactional
    def select(self, sheet_name):
        if sheet_name not in self.workbook.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' does not exist.")
        self.active_sheet = Sheet(self.workbook[sheet_name])
        return self

    @transactional
    def remove(self, sheet_name):
        """시트 삭제 메서드"""
        if sheet_name not in self.workbook.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' does not exist.")

        # 시트 삭제
        sheet_to_remove = self.workbook[sheet_name]
        self.workbook.remove(sheet_to_remove)
        print(f"Removed sheet: {sheet_name}")

        # 삭제된 시트가 활성 시트였을 경우, 다른 시트를 활성화
        if self.active_sheet and self.active_sheet.worksheet.title == sheet_name:
            if self.workbook.sheetnames:
                # 남아있는 시트 중 첫 번째 시트 활성화
                self.active_sheet = Sheet(self.workbook[self.workbook.sheetnames[0]])
            else:
                self.active_sheet = None  # 더 이상 시트가 없으면 None 처리

        return self

    @transactional
    def filter(self, include=None, exclude=None):
        if self.active_sheet:
            self.active_sheet.filter(include=include, exclude=exclude)
        return self

    @transactional
    def style(self, **kwargs):
        if self.active_sheet:
            self.active_sheet.style(**kwargs)
        return self

    @transactional
    def add_hyperlinks_to_column(self, column_name, urls, display_texts=None):
        if self.active_sheet:
            self.active_sheet.add_hyperlinks_to_column(column_name, urls, display_texts)
        return self

    @transactional
    def save(self, filename=None):
        self.workbook.save(filename or self.filename)
        print(f"Saved to {filename or self.filename}")
        self.commit()  # ✅ 저장 성공 시 커밋
        self.close()
        return self

    def end(self):
        # self.commit()
        print("End of chain.")
        return None

    def close(self):
        self.workbook = None
        self.active_sheet = None
        print("Workbook closed and resources released.")
