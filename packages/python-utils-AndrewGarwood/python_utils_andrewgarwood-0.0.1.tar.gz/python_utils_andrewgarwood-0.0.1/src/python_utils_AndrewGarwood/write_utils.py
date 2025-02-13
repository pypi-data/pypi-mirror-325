from typing import List, Dict, Any, Literal, Tuple
import os
from dotenv import load_dotenv
from datetime import datetime
import pandas as pd
from pandas import DataFrame
from typing import Dict, List, Literal
from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils.dataframe import dataframe_to_rows
from .file_utils import validate_file_extension

load_dotenv()
DEFAULT_LOG: str = os.getenv('DEFAULT_LOG_FILE')
FIELD_UPDATE_LOG: str = os.getenv('FIELD_UPDATE_LOG')

__all__ = [
    'print_group', 'concatenate_dataframes_to_excel', 'concatenate_dataframes_to_excel_sheet', 'write_dataframes_to_excel'
]

# like a worse version of json pretty printing...
def print_group(
    label: str = f'Print Group {datetime.now()}', 
    data: List[str] = [], 
    indent: int = 0,
    print_to_console: bool = True, 
    log_path: str = None,
    mode: Literal['w', 'a'] = 'a'
) -> None:
    timestamp: str = f"{datetime.now()}".split('.')[0]
    if not log_path and print_to_console:
        print("\t"*indent + f"({timestamp}) {label}")
        for item in data:
            print("\t"*(indent+1) + f"{item}")
    elif log_path and print_to_console:
        output_path = log_path if log_path and os.path.isfile(log_path) else DEFAULT_LOG
        with open(file=output_path, mode=mode) as f:
            f.write("\t"*indent + f"({timestamp}) {label}\n")
            for item in data:
                f.write("\t"*(indent+1) + f"{item}\n")
        print("\t"*indent + f"({timestamp}) {label}")
        for item in data:
            print("\t"*(indent+1) + f"{item}")
    else:
        output_path = log_path if log_path and os.path.isfile(log_path) else DEFAULT_LOG
        with open(file=output_path, mode=mode) as f:
            f.write("\t"*indent + f"({timestamp}) {label}\n")
            for item in data:
                f.write("\t"*(indent+1) + f"{item}\n")
                


def concatenate_dataframes_to_excel(
    concat_dict: Dict[Tuple[str, Literal['horizontal', 'vertical', 'h', 'v', 'horiz', 'vert', 'row', 'col']], List[DataFrame]],
    output_path: str,
    empty_spaces: int = 1
) -> None:
    """
    Concatenate dataframes to excel sheets.
    
    Args:
        concat_dict (Dict[Tuple[str, Literal['horizontal', 'vertical', 'h', 'v', 'horiz', 'vert', 'row', 'col']], List[DataFrame]]):
            Dictionary of (sheet name, orientation) tuples mapped to dataframes.
            {(sheet1, orientation1) : [dataframes to concatenate to sheet1], (sheet2, orientation2) : [dataframes to concatenate to sheet2]}
        output_path (str): verified by file_utils.validate_file_extension
        empty_spaces (int, optional): number of empty rows/cols between concatenated dataframes. Defaults to 1.
    """
    for (sheet_name, orientation), dataframes in concat_dict.items():
        concatenate_dataframes_to_excel_sheet(
            dataframes, 
            orientation, 
            output_path, 
            sheet_name, 
            empty_spaces
        )

def concatenate_dataframes_to_excel_sheet(
    dataframes: List[DataFrame], 
    orientation: Literal['horizontal', 'vertical', 'h', 'v', 'horiz', 'vert', 'row', 'col'],
    output_path: str, 
    sheet_name: str = 'Sheet1',
    empty_spaces: int = 1
) -> None:
    """_summary_

    Args:
        dataframes (List[DataFrame]): _description_
        orientation (Literal[&#39;horizontal&#39;, &#39;vertical&#39;, &#39;h&#39;, &#39;v&#39;, &#39;horiz&#39;, &#39;vert&#39;, &#39;row&#39;, &#39;col&#39;]): _description_
        output_path (str): _description_
        sheet_name (str, optional): _description_. Defaults to 'Sheet1'.
        empty_spaces (int, optional): _description_. Defaults to 1.

    Raises:
        ValueError: If given invalid orientation.
        FileNotFoundError: if given invalid output_path (see file_utils.validate_file_extension).
    """
    output_path = validate_file_extension(output_path, '.xlsx')
    try:
        wb: Workbook = load_workbook(output_path)
    except FileNotFoundError:
        wb = Workbook()
        wb.remove(wb.active)  # Remove the default sheet created with a new workbook

    ws: Worksheet = wb.create_sheet(title=sheet_name)
    ws.freeze_panes = 'A2'  # Freeze the first row
    if orientation.lower() in ['horizontal', 'h', 'horiz', 'row']:
        current_col = 1
        for df in dataframes:
            for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
                for c_idx, value in enumerate(row, current_col):
                    ws.cell(row=r_idx, column=c_idx, value=value)
            current_col += df.shape[1] + empty_spaces
    elif orientation.lower() in ['vertical', 'v', 'vert', 'col']:
        current_row = 1
        for df in dataframes:
            for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), current_row):
                for c_idx, value in enumerate(row, 1):
                    ws.cell(row=r_idx, column=c_idx, value=value)
            current_row += df.shape[0] + empty_spaces + 1

    else:
        raise ValueError("Orientation must be either 'horizontal' or 'vertical'")
    wb.save(output_path)


def write_dataframes_to_excel(
    output_path: str, 
    df_dict: Dict[str, DataFrame]
) -> None:
    """_summary_
    Write dataframes to excel using Pandas ExcelWriter. 
    If the file does not exist, it will be created. 
    If the file does exist, the dataframes will be appended to the existing file.
    Freezes the first row of each sheet.
    Args:
        output_path (str): _description_
        df_dict (Dict[str, DataFrame]): _description_
    """
    output_path = validate_file_extension(output_path, '.xlsx')
    try: # Load the existing workbook  
        workbook: Workbook = load_workbook(output_path)
        sheet_names = workbook.sheetnames
    except FileNotFoundError: # If the file does not exist, create a new workbook
        workbook = None
        sheet_names = []
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            df = pd.DataFrame({'': ['']})
            df.to_excel(writer, sheet_name='Sheet0', index=False, freeze_panes=(1, 0))
        workbook = load_workbook(output_path)
        sheet_names = workbook.sheetnames

    with pd.ExcelWriter(output_path, engine='openpyxl', mode='a') as writer:
        # writer.book = workbook
        for df_name, df in df_dict.items():
            # Ensure unique sheet names
            original_sheet_name = df_name
            index = 1
            while df_name in sheet_names:
                df_name = f"{original_sheet_name}{index}"
                index += 1
            df.to_excel(
                writer, sheet_name=df_name, freeze_panes=(1, 0),
                merge_cells=False, index=False
            )
            sheet_names.append(df_name)