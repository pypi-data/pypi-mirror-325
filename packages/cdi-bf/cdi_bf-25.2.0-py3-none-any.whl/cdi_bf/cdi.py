import pandas as pd
import openpyxl as px
import os
from datetime import datetime


class ExcelDataHandler(object):
    """
    Cette classe est utilisée pour lire et concaténer des données à partir de feuilles Excel.
    """

    def __init__(self, file_path):
        self.wb = None
        self.file_path = file_path

    def read_file(self):
        """Charge le classeur Excel en utilisant openpyxl."""
        try:
            self.wb = px.load_workbook(self.file_path)
        except FileNotFoundError as e:
            raise FileNotFoundError(f"Fichier non trouvé : {self.file_path}") from e
        return self.wb

    def sheet_names(self):
        """Retourne les noms des feuilles dans le classeur."""
        if self.wb is None:
            try:
                self.read_file()
            except FileNotFoundError as e:
                print(e)
                return []
        return self.wb.sheetnames

    def infos(self):
        """Retourne les informations des feuilles dans le classeur.
        ------------------------------------------------------
        """
        infos = {}
        for sheet in self.sheet_names():
            infos[sheet] = {
                "max_row": self.wb[sheet].max_row,
                "max_column": self.wb[sheet].max_column,
            }
        return infos

    def get_column_names(self, sheet, row=8):
        """Extrait les noms des colonnes à partir de la ligne spécifiée de la feuille.
        ------------------------------------------------------------
        Paramètres :
            sheet: L'objet feuille à partir duquel extraire les noms des colonnes.
            row: Le numéro de la ligne contenant les noms des colonnes.

        """
        column_labels = [
            sheet.cell(row=row, column=i).value for i in range(1, sheet.max_column + 1)
        ]
        column_labels_wth_none = [col for col in column_labels if col is not None]

        return column_labels_wth_none  # Retourne une liste plate

    def sheets_data(
        self,
        header: int = 8,
        drop_last: bool = True,
        last_rows: int = 2,
        drop_first: bool = True,
        first_rows: int = 2,
        column_indexes: list = None,
    ):
        """Lit et concatène les données de toutes les feuilles.
        ------------------------------------------------------------------------------
        Paramètres :
            row: Le numéro de la ligne à utiliser comme en-tête (par défaut 8).
            header: Le numéro de la ligne à utiliser comme en-tête (par défaut 8).
            drop_last: Indique s'il faut supprimer la dernière ligne de chaque feuille (par défaut True).
            last_rows: Le nombre de lignes à supprimer à la fin de chaque feuille (par défaut 1).
            first_rows: Le nombre de lignes à supprimer au début de chaque feuille (par défaut 1).
            drop_first: Indique s'il faut supprimer la première ligne de chaque feuille (par défaut True).
            index_column_specified: Indique si les index des colonnes à conserver sont spécifiés (par défaut False).
            column_indexes: Les index des colonnes à conserver (par défaut None).
        """

        # Vérifie si le fichier existe en le lisant
        try:
            self.read_file()
        except FileNotFoundError as e:
            print(e)
            return pd.DataFrame()

        all_data = []

        # Récupère les noms des feuilles
        sheet_names = self.sheet_names()
        if not sheet_names:
            raise ValueError("Aucune feuille trouvée.")

        # En supposant que toutes les feuilles ont la même structure, on récupère les colonnes de la première feuille
        column_labels_sheet_0 = self.get_column_names(
            self.wb[sheet_names[0]], row=header
        )
        column_labels_sheet_1 = self.get_column_names(
            self.wb[sheet_names[1]], row=header
        )
        #! verification de l'existance des noms de colonnes
        if not column_labels_sheet_0:
            raise ValueError("Aucun nom de colonne trouvé.")

        #! verification de la correspondance des noms de colonnes
        if column_labels_sheet_0 != column_labels_sheet_1:
            raise ValueError("Les noms de colonnes ne correspondent pas.")

        for sheet in sheet_names:
            # Lit les données, en utilisant row-1 car pandas est indexé à partir de 0 et openpyxl à partir de 1
            df = pd.read_excel(
                self.file_path,
                sheet_name=sheet,
                header=header - 1,  # Ajuste pour l'indexation basée sur 0 dans pandas
                engine="openpyxl",
            )

            if column_indexes is not None:
                column_index = column_indexes
                column_index = [
                    i - 1 for i in column_index
                ]  # Ajuste pour l'indexation basée sur 0 dans pandas
                df = df.iloc[:, column_index]  # Sélectionne les colonnes spécifiées
            else:
                df = df[column_labels_sheet_0]  # Sélectionne toutes les colonnes

            # Assure que les colonnes sont correctement nommées (en cas de divergences)
            if drop_last:
                df = df.iloc[
                    :-last_rows
                ]  # Supprime la dernière ligne (généralement les totaux)
            if drop_first:
                df = df.iloc[
                    first_rows:
                ]  # Supprime la première ligne (généralement les en-têtes)

            df.columns = column_labels_sheet_0

            # Ajoute les colonnes de date, code, label, recorded_date et recorded_time
            sixth_col = self.get_column_names(self.wb[sheet], row=6)
            forth_col = self.get_column_names(self.wb[sheet], row=4)
            third_col = self.get_column_names(self.wb[sheet], row=3)

            date = third_col[0]
            code = forth_col[1]
            label = forth_col[2]
            recorded_date = sixth_col[2]
            recorded_time = sixth_col[3]

            row = len(df)

            df["Periode"] = [date for i in range(row)]
            df["code"] = [code for i in range(row)]
            df["Libele"] = [label for i in range(row)]
            df["Date_enregistrement"] = [recorded_date for i in range(row)]
            df["Heure_enregistrement"] = [recorded_time for i in range(row)]

            all_data.append(df)

        data = pd.concat(all_data, ignore_index=True)

        # Supprime les colonnes avec toutes les valeurs None
        # none_columns = [col for col in data.columns if col is None]
        # data.drop(none_columns, axis=1, inplace=True)
        return data

    # methode pour sauvegarder les données dans un
    def save(self, file_name, file_path=".", file_format="xlsx", **kwargs):
        """Sauvegarde les données dans un nouveau fichier specifié."""

        # Vérifie si le fichier existe en le lisant
        if not os.path.exists(file_path):
            os.makedirs(file_path)
        file_name = f"{file_path}/{file_name}.{file_format}"

        data = self.sheets_data(**kwargs)
        # sauvegarde des données sous le format spécifié
        if file_format == "csv":
            data.to_csv(file_name, index=False)
        elif file_format == "xlsx" or file_format == "xls":
            data.to_excel(file_name, index=False)
        elif file_format == "json":
            data.to_json(file_name, orient="records")
        elif file_format == "parquet":
            data.to_parquet(file_name, index=False)
        else:
            raise ValueError("Format de fichier non pris en charge.")
        print(f"Les données ont été sauvegardées dans {file_name}")
        return file_name
