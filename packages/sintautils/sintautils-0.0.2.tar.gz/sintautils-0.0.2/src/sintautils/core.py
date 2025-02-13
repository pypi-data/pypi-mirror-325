"""
sintautils
Created by: Samarthya Lykamanuella
Establishment year: 2025

LICENSE NOTICE:
===============

This file is part of sintautils.

sintautils is free software: you can redistribute it and/or modify it
under the terms of the GNU General Public License as published by the
Free Software Foundation, either version 3 of the License, or (at your
option) any later version.

sintautils is distributed in the hope that it will be useful, but WITHOUT
ANY WARRANTY; without even the implied warranty of MERCHANTABILITY or
FITNESS FOR A PARTICULAR PURPOSE. See the GNU General Public License
for more details.

You should have received a copy of the GNU General Public License along
with sintautils. If not, see <https://www.gnu.org/licenses/>.
"""
import copy
from datetime import datetime as dt
import json
from openpyxl import Workbook
import os
import requests as rq
import time

from .exceptions import EmptyFieldException
from .exceptions import InvalidLoginCredentialException
from .exceptions import InvalidParameterException
from .exceptions import NoLoginCredentialsException
from .exceptions import NonStringParameterException
from .backend import UtilBackEnd


# noinspection SpellCheckingInspection
class SintaScraper(object):
    """ The super-class for all SINTA scrapers in sintautils. Do not invoke directly! """

    # Determine the verbosity of loggings and debug messages.
    # Must be of: 0 (no verbose message), 1 (moderate logging level), and 2 (log all events)
    # If less than 0, it will be interpreted as 0.
    # If greater than 2, it will be interpreted as 2.
    verbosity = 1

    # Determine if the timestamp should be given in the logging message.
    log_timestamp = False

    def __init__(self, username: str = '', password: str = ''):
        self.username = username
        self.password = password

        # Initiating the session.
        self.s = rq.Session()

    def print(self, msg, verbose_level: int):
        """ Debug logging with verbosity control.
        :param msg: the message to be logged to the output terminal.
        :param verbose_level: between 0 and 2, determining on which verbosity level this message will be shown.
        """
        if self.verbosity < 0:
            self.verbosity = 0
        elif self.verbosity > 2:
            self.verbosity = 2

        # Do the message logging.
        if verbose_level <= self.verbosity:
            if self.log_timestamp:
                print('::: [' + str(dt.now()) + '] + ' + str(msg))
            else:
                print(str(msg))


class AV(SintaScraper):
    """ The scraper for the SINTA author verification site (https://sinta.kemdikbud.go.id/authorverification). """

    LOGIN_URL = 'https://sinta.kemdikbud.go.id/authorverification/login/do_login'

    def __init__(self, username: str = '', password: str = '', autologin: bool = False):
        if type(username) is not str or type(password) is not str:
            raise NonStringParameterException()

        elif username.strip() == '' or password.strip() == '':
            raise NoLoginCredentialsException()

        else:
            # We got the credential! Now let's proceed.
            super().__init__(username, password)

            # Initializing the back-end.
            self.backend = UtilBackEnd(self.s, self.print)

            if autologin:
                self.login()

    # noinspection PyDefaultArgument,PyTypeChecker
    def _get_dump(self, author_id: str, fields: list = ['*']):
        """ This function actually gets the dump information requested by the function `dump_author()`.

        :param author_id: the author ID of one, and only one author.
        :param fields: the kind of dump information to be returned.
        :return: a dict that corresponds to the data fields requested.

        The return dict, if specified in the `fields` parameter, will have the following keys:
        - "book"
        - "garuda"
        - "gscholar"
        - "ipr"
        - "research"
        - "scopus"
        - "service"
        - "wos"
        """
        ret_dict = {}

        if type(author_id) is not str:
            raise NonStringParameterException()

        if '*' in fields or 'book' in fields:
            self.print('Obtaining book data...', 0)
            ret_dict['book'] = self.get_book(author_id, out_format='json')

        if '*' in fields or 'garuda' in fields:
            self.print('Obtaining garuda data...', 0)
            ret_dict['garuda'] = self.get_garuda(author_id, out_format='json')

        if '*' in fields or 'gscholar' in fields:
            self.print('Obtaining gscholar data...', 0)
            ret_dict['gscholar'] = self.get_gscholar(author_id, out_format='json')

        if '*' in fields or 'ipr' in fields:
            self.print('Obtaining ipr data...', 0)
            ret_dict['ipr'] = self.get_ipr(author_id, out_format='json')

        if '*' in fields or 'research' in fields:
            self.print('Obtaining research data...', 0)
            ret_dict['research'] = self.get_research(author_id, out_format='json')

        if '*' in fields or 'scopus' in fields:
            self.print('Obtaining scopus data...', 0)
            ret_dict['scopus'] = self.get_scopus(author_id, out_format='json')

        if '*' in fields or 'service' in fields:
            self.print('Obtaining service data...', 0)
            ret_dict['service'] = self.get_service(author_id, out_format='json')

        if '*' in fields or 'wos' in fields:
            self.print('Obtaining wos data...', 0)
            ret_dict['wos'] = self.get_wos(author_id, out_format='json')

        return ret_dict

    # noinspection PyDefaultArgument
    def dump_author(
            self,
            author_id: list = [],
            out_folder: str = os.getcwd(),
            out_prefix: str = 'sintautils_dump_author-',
            out_format: str = 'xlsx',
            fields: list = ['*'],
            use_fullname_prefix: bool = True,
    ):
        """ Performs the scraping of an author's all-information data, and save it as a file.

        :param author_id: the list of author IDs to be scraped.
        :param out_folder: the output folder to which all scraping result files will be saved. Defaults to the current working directory.
        :param out_prefix: the prefix file name into which the scraping result(s) will be saved. The default output filename prefix is "sintautils_dump_author-".
        :param use_fullname_prefix: whether to use the corresponding author's full name as file prefix. If set to "True" (the default), this parameter overrides the values passed to `out_prefix`.

        :param out_format: the format of the output result document.

        Currently, the only supported formats are as follows:
        - "csv"
        - "json"
        - "json-pretty"
        - "xlsx"

        :param fields: the types of scraping data to be scraped.

        Currently, the only supported fields are as follows:
        - "*"
        - "book"
        - "garuda"
        - "gscholar"
        - "ipr"
        - "research"
        - "scopus"
        - "service"
        - "wos"

        You can input more than one field. For instance:
        - ["gscholar", "scopus"]
        - ["wos", "research", "scopus"]

        Use asterisk in order to return all fields:
        - ["*"]
        """

        # Validating the fields.
        if fields.__len__() < 1:
            raise EmptyFieldException('book, garuda, gscholar, ipr, research, scopus, service, wos')

        # Validating the output format.
        if type(out_format) is not str or out_format not in ['csv', 'json', 'xlsx']:
            raise InvalidParameterException('"out_format" must be one of "csv", "json", and "xlsx"')

        # This local function carries out the actual dumping of author data.
        def dump(dump_id):
            di = str(dump_id).strip()
            a: dict = self._get_dump(di, fields=fields)
            r: str = self.backend.get_author_full_name(di)

            # Whether to use author's full name or just a computer-readable file name.
            if use_fullname_prefix:
                op = str(out_folder) + os.sep + r + '_' + str(di)
            else:
                op = str(out_folder) + os.sep + str(out_prefix) + str(di)

            if out_format == 'csv':
                for m in a.keys():
                    with open(op + '-' + m + '.csv', 'w') as fo:
                        b: list = a[m]

                        # Length validation.
                        if b.__len__() < 1:
                            continue

                        # Write the CSV header.
                        headers: str = str()
                        for n in b[0].keys():
                            headers += f'"{n}",'
                        fo.write(headers[:-1] + '\n')

                        # Write the content.
                        for n in b:
                            n: dict = n
                            rows: str = str()
                            for c in n.keys():
                                rows += f'"{n[c]}",'
                            fo.write(rows[:-1] + '\n')

            elif out_format == 'json' or out_format == 'json-pretty':
                b: dict = {
                    'data': {
                        'author_id': di,
                        'scraping_date': int(time.time()),
                        'scraping_result': copy.deepcopy(a)
                    }
                }

                # Saving the JSON file.
                with open(op + '.json', 'w') as fo:
                    if out_format == 'json-pretty':
                        json.dump(b, fo, indent=4)
                    else:
                        json.dump(b, fo)

            elif out_format == 'xlsx':
                wb = Workbook()
                for m in sorted(a.keys()):
                    ws = wb.create_sheet(m, -1)

                    # Obtaining the data list and validate the data length.
                    b: list = a[m]
                    if b.__len__() < 1:
                        continue

                    # Write the spreadsheet header.
                    headers: list = list(b[0].keys())
                    for i in range(len(headers)):
                        n = headers[i]
                        ws.cell(row=1, column=(i + 1), value=n)

                        # Write the column's content.
                        for j in range(len(b)):
                            c: dict = b[j]
                            # Offset the row number by two, because the first row is header.
                            ws.cell(row=(j + 2), column=(i + 1), value=c[n])

                # Remove sheets that do not represent data type.
                if wb.sheetnames.__len__() > 0:
                    for d in wb.sheetnames:
                        if d not in a.keys():
                            wb.remove(wb[d])

                # Saving the spreadsheet.
                save_file = op + '.xlsx'
                wb.save(save_file)

        if type(author_id) in (str, int):
            dump(dump_id=str(author_id))

        elif type(author_id) is list:
            # Remove duplicates.
            author_id = list(dict.fromkeys(author_id))

            # Validating individual item type.
            for l in author_id:
                if type(l) not in (str, int):
                    raise InvalidParameterException('You can only pass list, string, or integer into this function')

            for l in author_id:
                self.print(f'Dumping author data for author ID: {l}...', 0)
                dump(dump_id=str(l))

        else:
            raise InvalidParameterException('You can only pass list, string, or integer into this function')

    # noinspection PyDefaultArgument
    def get_book(self, author_id: list = [], out_format: str = 'csv', fields: list = ['*']):
        """ Performs the scraping of individual author's book data.

        :param author_id: the list of author IDs to be scraped.
        :param out_format: the format of the output result document.

        Currently, the only supported formats are as follows:
        - "csv"
        - "json"

        You can only specify one output format at a time.

        :param fields: the types of field to be scraped.

        Currently, the only supported fields are as follows:
        - "*"
        - "title"
        - "isbn"
        - "author"
        - "publisher"
        - "page"
        - "year"
        - "location"
        - "thumbnail"
        - "url"

        You can input more than one field. For instance:
        - ["title", "thumbnail"]
        - ["isbn", "year", "thumbnail"]

        Use asterisk in order to return all fields:
        - ["*"]
        """

        if type(author_id) in (str, int):
            a = self.backend.scrape_book(author_id=str(author_id), out_format=out_format, fields=fields)

        elif type(author_id) is list:
            # Remove duplicates.
            author_id = list(dict.fromkeys(author_id))

            # Validating individual item type.
            for l in author_id:
                if type(l) not in (str, int):
                    raise InvalidParameterException('You can only pass list, string, or integer into this function')

            a = {}
            for l in author_id:
                self.print(f'Scraping for author ID: {l}...', 1)
                a[l] = self.backend.scrape_book(author_id=str(l), out_format=out_format, fields=fields)

        else:
            raise InvalidParameterException('You can only pass list, string, or integer into this function')

        return a

    # noinspection PyDefaultArgument
    def get_garuda(self, author_id: list = [], out_format: str = 'csv', fields: list = ['*']):
        """ Performs the scraping of individual author's Garuda data.

        :param author_id: the list of author IDs to be scraped.
        :param out_format: the format of the output result document.

        Currently, the only supported formats are as follows:
        - "csv"
        - "json"

        You can only specify one output format at a time.

        :param fields: the types of field to be scraped.

        Currently, the only supported fields are as follows:
        - "*"
        - "title"
        - "publisher"
        - "author"
        - "journal"
        - "doi"
        - "year"
        - "quartile"
        - "url"

        You can input more than one field. For instance:
        - ["journal", "url"]
        - ["quartile", "doi", "year"]

        Use asterisk in order to return all fields:
        - ["*"]
        """

        if type(author_id) in (str, int):
            a = self.backend.scrape_garuda(author_id=str(author_id), out_format=out_format, fields=fields)

        elif type(author_id) is list:
            # Remove duplicates.
            author_id = list(dict.fromkeys(author_id))

            # Validating individual item type.
            for l in author_id:
                if type(l) not in (str, int):
                    raise InvalidParameterException('You can only pass list, string, or integer into this function')

            a = {}
            for l in author_id:
                self.print(f'Scraping for author ID: {l}...', 1)
                a[l] = self.backend.scrape_garuda(author_id=str(l), out_format=out_format, fields=fields)

        else:
            raise InvalidParameterException('You can only pass list, string, or integer into this function')

        return a

    # noinspection PyDefaultArgument
    def get_gscholar(self, author_id: list = [], out_format: str = 'csv', fields: list = ['*']):
        """ Performs the scraping of individual author's Google Scholar data.

        :param author_id: the list of author IDs to be scraped.
        :param out_format: the format of the output result document.

        Currently, the only supported formats are as follows:
        - "csv"
        - "json"

        You can only specify one output format at a time.

        :param fields: the types of field to be scraped.

        Currently, the only supported fields are as follows:
        - "*"
        - "title"
        - "author"
        - "journal"
        - "year"
        - "citations"
        - "url"

        You can input more than one field. For instance:
        - ["journal", "url"]
        - ["title", "author", "year"]

        Use asterisk in order to return all fields:
        - ["*"]
        """

        if type(author_id) in (str, int):
            a = self.backend.scrape_gscholar(author_id=str(author_id), out_format=out_format, fields=fields)

        elif type(author_id) is list:
            # Remove duplicates.
            author_id = list(dict.fromkeys(author_id))

            # Validating individual item type.
            for l in author_id:
                if type(l) not in (str, int):
                    raise InvalidParameterException('You can only pass list, string, or integer into this function')

            a = {}
            for l in author_id:
                self.print(f'Scraping for author ID: {l}...', 1)
                a[l] = self.backend.scrape_gscholar(author_id=str(l), out_format=out_format, fields=fields)

        else:
            raise InvalidParameterException('You can only pass list, string, or integer into this function')

        return a

    # noinspection PyDefaultArgument
    def get_ipr(self, author_id: list = [], out_format: str = 'csv', fields: list = ['*']):
        """ Performs the scraping of individual author's IPR data.

        :param author_id: the list of author IDs to be scraped.
        :param out_format: the format of the output result document.

        Currently, the only supported formats are as follows:
        - "csv"
        - "json"

        You can only specify one output format at a time.

        :param fields: the types of field to be scraped.

        Currently, the only supported fields are as follows:
        - "*"
        - "title"
        - "application_no"
        - "inventor"
        - "patent_holder"
        - "category"
        - "year"
        - "status"

        You can input more than one field. For instance:
        - ["inventor", "status"]
        - ["title", "patent_holder", "status"]

        Use asterisk in order to return all fields:
        - ["*"]
        """

        if type(author_id) in (str, int):
            a = self.backend.scrape_ipr(author_id=str(author_id), out_format=out_format, fields=fields)

        elif type(author_id) is list:
            # Remove duplicates.
            author_id = list(dict.fromkeys(author_id))

            # Validating individual item type.
            for l in author_id:
                if type(l) not in (str, int):
                    raise InvalidParameterException('You can only pass list, string, or integer into this function')

            a = {}
            for l in author_id:
                self.print(f'Scraping for author ID: {l}...', 1)
                a[l] = self.backend.scrape_ipr(author_id=str(l), out_format=out_format, fields=fields)

        else:
            raise InvalidParameterException('You can only pass list, string, or integer into this function')

        return a

    # noinspection PyDefaultArgument
    def get_research(self, author_id: list = [], out_format: str = 'csv', fields: list = ['*']):
        """ Performs the scraping of individual author's research data.

        :param author_id: the list of author IDs to be scraped.
        :param out_format: the format of the output result document.

        Currently, the only supported formats are as follows:
        - "csv"
        - "json"

        You can only specify one output format at a time.

        :param fields: the types of field to be scraped.

        Currently, the only supported fields are as follows:
        - "*"
        - "title"
        - "funds"
        - "program"
        - "schema"
        - "year"
        - "membership"
        - "url"

        You can input more than one field. For instance:
        - ["program", "membership"]
        - ["title", "funds", "membership"]

        Use asterisk in order to return all fields:
        - ["*"]
        """

        if type(author_id) in (str, int):
            a = self.backend.scrape_research(author_id=str(author_id), out_format=out_format, fields=fields)

        elif type(author_id) is list:
            # Remove duplicates.
            author_id = list(dict.fromkeys(author_id))

            # Validating individual item type.
            for l in author_id:
                if type(l) not in (str, int):
                    raise InvalidParameterException('You can only pass list, string, or integer into this function')

            a = {}
            for l in author_id:
                self.print(f'Scraping for author ID: {l}...', 1)
                a[l] = self.backend.scrape_research(author_id=str(l), out_format=out_format, fields=fields)

        else:
            raise InvalidParameterException('You can only pass list, string, or integer into this function')

        return a

    # noinspection PyDefaultArgument
    def get_scopus(self, author_id: list = [], out_format: str = 'csv', fields: list = ['*']):
        """ Performs the scraping of individual author's scopus data.

        :param author_id: the list of author IDs to be scraped.
        :param out_format: the format of the output result document.

        Currently, the only supported formats are as follows:
        - "csv"
        - "json"

        You can only specify one output format at a time.

        :param fields: the types of field to be scraped.

        Currently, the only supported fields are as follows:
        - "*"
        - "title"
        - "author"
        - "journal"
        - "type"
        - "year"
        - "citations"
        - "quartile"
        - "url"

        You can input more than one field. For instance:
        - ["journal", "url"]
        - ["quartile", "citations", "year"]

        Use asterisk in order to return all fields:
        - ["*"]
        """

        if type(author_id) in (str, int):
            a = self.backend.scrape_scopus(author_id=str(author_id), out_format=out_format, fields=fields)

        elif type(author_id) is list:
            # Remove duplicates.
            author_id = list(dict.fromkeys(author_id))

            # Validating individual item type.
            for l in author_id:
                if type(l) not in (str, int):
                    raise InvalidParameterException('You can only pass list, string, or integer into this function')

            a = {}
            for l in author_id:
                self.print(f'Scraping for author ID: {l}...', 1)
                a[l] = self.backend.scrape_scopus(author_id=str(l), out_format=out_format, fields=fields)

        else:
            raise InvalidParameterException('You can only pass list, string, or integer into this function')

        return a

    # noinspection PyDefaultArgument
    def get_service(self, author_id: list = [], out_format: str = 'csv', fields: list = ['*']):
        """ Performs the scraping of individual author's community service data.

        :param author_id: the list of author IDs to be scraped.
        :param out_format: the format of the output result document.

        Currently, the only supported formats are as follows:
        - "csv"
        - "json"

        You can only specify one output format at a time.

        :param fields: the types of field to be scraped.

        Currently, the only supported fields are as follows:
        - "*"
        - "title"
        - "funds"
        - "program"
        - "schema"
        - "year"
        - "membership"
        - "url"

        You can input more than one field. For instance:
        - ["program", "membership"]
        - ["title", "funds", "membership"]

        Use asterisk in order to return all fields:
        - ["*"]
        """

        if type(author_id) in (str, int):
            a = self.backend.scrape_service(author_id=str(author_id), out_format=out_format, fields=fields)

        elif type(author_id) is list:
            # Remove duplicates.
            author_id = list(dict.fromkeys(author_id))

            # Validating individual item type.
            for l in author_id:
                if type(l) not in (str, int):
                    raise InvalidParameterException('You can only pass list, string, or integer into this function')

            a = {}
            for l in author_id:
                self.print(f'Scraping for author ID: {l}...', 1)
                a[l] = self.backend.scrape_service(author_id=str(l), out_format=out_format, fields=fields)

        else:
            raise InvalidParameterException('You can only pass list, string, or integer into this function')

        return a

    # noinspection PyDefaultArgument
    def get_wos(self, author_id: list = [], out_format: str = 'csv', fields: list = ['*']):
        """ Performs the scraping of individual author's Web of Science (WOS) data.

        :param author_id: the list of author IDs to be scraped.
        :param out_format: the format of the output result document.

        Currently, the only supported formats are as follows:
        - "csv"
        - "json"

        You can only specify one output format at a time.

        :param fields: the types of field to be scraped.

        Currently, the only supported fields are as follows:
        - "*"
        - "title"
        - "author"
        - "journal"
        - "year"
        - "citations"
        - "quartile"
        - "url"

        You can input more than one field. For instance:
        - ["journal", "url"]
        - ["quartile", "citations", "year"]

        Use asterisk in order to return all fields:
        - ["*"]
        """

        if type(author_id) in (str, int):
            a = self.backend.scrape_wos(author_id=str(author_id), out_format=out_format, fields=fields)

        elif type(author_id) is list:
            # Remove duplicates.
            author_id = list(dict.fromkeys(author_id))

            # Validating individual item type.
            for l in author_id:
                if type(l) not in (str, int):
                    raise InvalidParameterException('You can only pass list, string, or integer into this function')

            a = {}
            for l in author_id:
                self.print(f'Scraping for author ID: {l}...', 1)
                a[l] = self.backend.scrape_wos(author_id=str(l), out_format=out_format, fields=fields)

        else:
            raise InvalidParameterException('You can only pass list, string, or integer into this function')

        return a

    def login(self):
        """ Performs the credential login and obtains the session cookie for this account. """
        r = self.s.post(self.LOGIN_URL, data={'username': self.username, 'password': self.password})
        self.print(r.status_code, 2)
        self.print(r.url, 2)

        if 'dashboard' in r.url:
            self.print('Login successful!', 1)
        else:
            raise InvalidLoginCredentialException()

    # noinspection PyDefaultArgument
    def sync_dikti(self, author_id: list = []):
        """ Performs the syncing of an author's PD-DIKTI profile data.
        :param author_id: the list of author IDs to be synced.
        :return: nothing.
        """

        if type(author_id) in (str, int):
            a = self.backend.sync_dikti(author_id=str(author_id))

        elif type(author_id) is list:
            # Remove duplicates.
            author_id = list(dict.fromkeys(author_id))

            # Validating individual item type.
            for l in author_id:
                if type(l) not in (str, int):
                    raise InvalidParameterException('You can only pass list, string, or integer into this function')

            a = {}
            for l in author_id:
                self.print(f'Syncing author ID: {l}...', 1)
                a[l] = self.backend.sync_dikti(author_id=str(l))

        else:
            raise InvalidParameterException('You can only pass list, string, or integer into this function')

    # noinspection PyDefaultArgument
    def sync_garuda(self, author_id: list = []):
        """ Performs the syncing of an author's Garuda data.
        :param author_id: the list of author IDs to be synced.
        :return: nothing.
        """

        if type(author_id) in (str, int):
            a = self.backend.sync_garuda(author_id=str(author_id))

        elif type(author_id) is list:
            # Remove duplicates.
            author_id = list(dict.fromkeys(author_id))

            # Validating individual item type.
            for l in author_id:
                if type(l) not in (str, int):
                    raise InvalidParameterException('You can only pass list, string, or integer into this function')

            a = {}
            for l in author_id:
                self.print(f'Syncing author ID: {l}...', 1)
                a[l] = self.backend.sync_garuda(author_id=str(l))

        else:
            raise InvalidParameterException('You can only pass list, string, or integer into this function')

    # noinspection PyDefaultArgument
    def sync_gscholar(self, author_id: list = []):
        """ Performs the syncing of an author's Google Scholar data.
        :param author_id: the list of author IDs to be synced.
        :return: nothing.
        """

        if type(author_id) in (str, int):
            a = self.backend.sync_gscholar(author_id=str(author_id))

        elif type(author_id) is list:
            # Remove duplicates.
            author_id = list(dict.fromkeys(author_id))

            # Validating individual item type.
            for l in author_id:
                if type(l) not in (str, int):
                    raise InvalidParameterException('You can only pass list, string, or integer into this function')

            a = {}
            for l in author_id:
                self.print(f'Syncing author ID: {l}...', 1)
                a[l] = self.backend.sync_gscholar(author_id=str(l))

        else:
            raise InvalidParameterException('You can only pass list, string, or integer into this function')

    # noinspection PyDefaultArgument
    def sync_research(self, author_id: list = []):
        """ Performs the syncing of an author's research data.
        :param author_id: the list of author IDs to be synced.
        :return: nothing.
        """

        if type(author_id) in (str, int):
            a = self.backend.sync_research(author_id=str(author_id))

        elif type(author_id) is list:
            # Remove duplicates.
            author_id = list(dict.fromkeys(author_id))

            # Validating individual item type.
            for l in author_id:
                if type(l) not in (str, int):
                    raise InvalidParameterException('You can only pass list, string, or integer into this function')

            a = {}
            for l in author_id:
                self.print(f'Syncing author ID: {l}...', 1)
                a[l] = self.backend.sync_research(author_id=str(l))

        else:
            raise InvalidParameterException('You can only pass list, string, or integer into this function')

    # noinspection PyDefaultArgument
    def sync_scopus(self, author_id: list = []):
        """ Performs the syncing of an author's Scopus data.
        :param author_id: the list of author IDs to be synced.
        :return: nothing.
        """

        if type(author_id) in (str, int):
            a = self.backend.sync_scopus(author_id=str(author_id))

        elif type(author_id) is list:
            # Remove duplicates.
            author_id = list(dict.fromkeys(author_id))

            # Validating individual item type.
            for l in author_id:
                if type(l) not in (str, int):
                    raise InvalidParameterException('You can only pass list, string, or integer into this function')

            a = {}
            for l in author_id:
                self.print(f'Syncing author ID: {l}...', 1)
                a[l] = self.backend.sync_scopus(author_id=str(l))

        else:
            raise InvalidParameterException('You can only pass list, string, or integer into this function')

    # noinspection PyDefaultArgument
    def sync_service(self, author_id: list = []):
        """ Performs the syncing of an author's community services data.
        :param author_id: the list of author IDs to be synced.
        :return: nothing.
        """

        if type(author_id) in (str, int):
            a = self.backend.sync_service(author_id=str(author_id))

        elif type(author_id) is list:
            # Remove duplicates.
            author_id = list(dict.fromkeys(author_id))

            # Validating individual item type.
            for l in author_id:
                if type(l) not in (str, int):
                    raise InvalidParameterException('You can only pass list, string, or integer into this function')

            a = {}
            for l in author_id:
                self.print(f'Syncing author ID: {l}...', 1)
                a[l] = self.backend.sync_service(author_id=str(l))

        else:
            raise InvalidParameterException('You can only pass list, string, or integer into this function')

    # noinspection PyDefaultArgument
    def sync_wos(self, author_id: list = []):
        """ Performs the syncing of an author's Web of Science (WOS) publication data.
        :param author_id: the list of author IDs to be synced.
        :return: nothing.
        """

        if type(author_id) in (str, int):
            a = self.backend.sync_wos(author_id=str(author_id))

        elif type(author_id) is list:
            # Remove duplicates.
            author_id = list(dict.fromkeys(author_id))

            # Validating individual item type.
            for l in author_id:
                if type(l) not in (str, int):
                    raise InvalidParameterException('You can only pass list, string, or integer into this function')

            a = {}
            for l in author_id:
                self.print(f'Syncing author ID: {l}...', 1)
                a[l] = self.backend.sync_wos(author_id=str(l))

        else:
            raise InvalidParameterException('You can only pass list, string, or integer into this function')
