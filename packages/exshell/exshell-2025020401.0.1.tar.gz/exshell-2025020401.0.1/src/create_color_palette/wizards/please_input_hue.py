import openpyxl as xl
import subprocess
import time

from openpyxl.styles import PatternFill

from src.create_color_palette import Color, ToneSystem


class PleaseInputHue():


    def play(exshell):

        # ワークブックを新規生成
        wb = xl.Workbook()

        # ワークシート
        ws = wb['Sheet']

        cell = ws[f'B1']
        cell.value = "色"

        cell = ws[f'C1']
        cell.value = "この番号を入力してください"


        number_of_colors = 12

        for index in range(0, number_of_colors):
            # 小数点以下第２位で丸め
            hue = round(index / number_of_colors, 2)

            tone_system = ToneSystem(
                    low=0,
                    high=255,
                    hue=hue)
            color_obj = Color(tone_system.get_red(), tone_system.get_green(), tone_system.get_blue())

            web_safe_color = color_obj.to_web_safe_color()
            xl_color = web_safe_color[1:]
            try:
                pattern_fill = PatternFill(
                        patternType='solid',
                        fgColor=xl_color)
            except:
                print(f'{xl_color=}')
                raise


            row_th = index + 2

            # 色
            cell = ws[f'B{row_th}']
            cell.fill = pattern_fill

            # コメント
            cell = ws[f'C{row_th}']
            cell.value = hue

        try:
            # ワークブック保存
            exshell.save_workbook(wb=wb)

        except PermissionError:
            # TODO ファイルが既に開かれているかも
            message = f"""\
🙋　Error
----------
📄［ {exshell.abs_path_to_workbook} ］ファイルが既に開かれているかもしれません？
原因を取り除いたあと、Enter キーを空打ちしてください。

    Example of input
    ----------------
    

Input
-----
"""
            input(message)
            return True, None


        # エクセル開く
        exshell.open_virtual_display()


        message = """\
🙋　Please input
-----------------
開かれたワークシートから、好きな色を１つ選んで番号を入力してください。
番号はワークシートに書いていない番号でも、 0 以上 1 以下の実数で入力できます。
分からなかったら 0 を入力してください。

    Example of input
    ----------------
    0.8123

Input
-----
"""
        line = input(message)
        number_of_hue = float(line)
        print() # 空行


        # エクセルを閉じる
        exshell.close_virtual_display()


        return False, number_of_hue
