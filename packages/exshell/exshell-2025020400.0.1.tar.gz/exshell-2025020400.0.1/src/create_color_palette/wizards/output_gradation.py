import openpyxl as xl
#import random

from openpyxl.styles import Font, PatternFill
from openpyxl.styles.alignment import Alignment

from src.create_color_palette import Color, ToneSystem


class OutputGradation():


    def play(number_of_color_samples, start_hue, saturation, brightness, exshell):
        # ワークブックを新規生成
        wb = xl.Workbook()

        # ワークシート
        ws = wb['Sheet']

        low, high = OutputGradation.create_tone(
                saturation=saturation,
                brightness=brightness)
        
        # 色相 [0.0, 1.0]
        #cur_hue = random.uniform(0, 1)
        cur_hue = start_hue
        step_hue = 1 / number_of_color_samples
#     print(f"""\
# {step_hue=}""")

        ws.column_dimensions['A'].width = 2.7 * 2    # 2.7 characters = about 30 pixels
        ws.column_dimensions['B'].width = 2.7 * 5    # 2.7 characters = about 30 pixels
        ws.column_dimensions['C'].width = 2.7 * 7    # 2.7 characters = about 30 pixels

        title_font = Font(color='F8F8F8')
        title_pattern_fill = PatternFill(
                patternType='solid',
                fgColor='333333')

        cell = ws[f'A1']
        cell.value = "No"
        cell.font = title_font
        cell.fill = title_pattern_fill
        cell.alignment = Alignment(horizontal='right', vertical='center')

        cell = ws[f'B1']
        cell.value = "色"
        cell.font = title_font
        cell.fill = title_pattern_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')

        cell = ws[f'C1']
        cell.value = "ウェブ・セーフ・カラー"
        cell.font = title_font
        cell.fill = title_pattern_fill
        cell.alignment = Alignment(horizontal='left', vertical='center')

        # デバッグ用情報
        # cell = ws[f'A1']
        # cell.value = "色相"

        # cell = ws[f'B1']
        # cell.value = "色相種類"

        # cell = ws[f'C1']
        # cell.value = "色相内段階"


        for index, row_th in enumerate(range(2, 2 + number_of_color_samples)):

            tone_system = ToneSystem(
                    low=low,
                    high=high,
                    hue=cur_hue)

            color_obj = Color(tone_system.get_red(), tone_system.get_green(), tone_system.get_blue())
    
            web_safe_color = color_obj.to_web_safe_color()
            xl_color = web_safe_color[1:]
            try:
                pattern_fill = PatternFill(
                        patternType='solid',
                        fgColor=xl_color)
            except:
                print(f'{xl_color=}')
                raise

            # 連番
            cell = ws[f'A{row_th}']
            cell.value = index

            # 色
            cell = ws[f'B{row_th}']
            cell.fill = pattern_fill

            # ウェブ・セーフ・カラー
            cell = ws[f'C{row_th}']
            cell.value = web_safe_color.upper()

            # デバッグ情報
            # cell = ws[f'A{row_th}']
            # cell.value = cur_hue

            # cell = ws[f'B{row_th}']
            # cell.value = tone_system.get_phase_name()

            # cell = ws[f'C{row_th}']
            # cell.value = tone_system.get_value_of_hue_in_phase()

            cur_hue += step_hue
            if 1 < cur_hue:
                cur_hue -= 1


        # ワークブック保存
        exshell.save_workbook(wb=wb)


        is_successful = False

        # エクセル開く
        exshell.open_virtual_display()


        message = f"""\
🙋　Please input
-----------------
グラデーションを作成しました。

アプリケーションを終了するなら `exit` を、
やり直す場合は　それ以外を入力してください。

    Example of input
    ----------------
    exit

Input
-----
"""
        line = input(message)
        print() # 空行


        # エクセル閉じる
        exshell.close_virtual_display()


        return line == 'exit'


    @staticmethod
    def create_tone(saturation, brightness):
        """色調を１つに決めます。

        Parameters
        ----------
        saturation : int
            彩度。[0, 255] の整数
            NOTE モノクロに近づくと、標本数が多くなると、色の違いを出しにくいです。
        brightness : int
            明度
        """

        # NOTE ウェブ・セーフ・カラーは、暗い色の幅が多めに取られています。 0～255 のうち、 180 ぐらいまで暗い色です。
        # NOTE 色の標本数が多くなると、 low, high は極端にできません。変化の幅が狭まってしまいます。

        # 上限
        high = brightness
        # 下限
        low = brightness - saturation

        if 255 < high:
            raise ValueError(f'{high=} Others: {brightness=} {saturation=}')

        if low < 0:
            raise ValueError(f'{low=} Others: {brightness=} {saturation=}')


        return low, high
