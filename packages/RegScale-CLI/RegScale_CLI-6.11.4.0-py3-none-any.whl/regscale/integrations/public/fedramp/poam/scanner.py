#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""FedRAMP Scanner Integration"""
import logging
import re
from typing import Iterator, Optional

from openpyxl import load_workbook  # type: ignore
from openpyxl.utils import column_index_from_string  # type: ignore
from openpyxl.utils.exceptions import InvalidFileException  # type: ignore
from openpyxl.workbook import Workbook

from regscale.core.app.utils.app_utils import get_current_datetime
from regscale.core.utils.date import date_str
from regscale.integrations.scanner_integration import (
    IntegrationAsset,
    IntegrationFinding,
    ScannerIntegration,
    issue_due_date,
)
from regscale.integrations.variables import ScannerVariables
from regscale.models import IssueSeverity, regscale_models

logger = logging.getLogger("rich")


class FedrampPoamIntegration(ScannerIntegration):
    """Integration class for FedRAMP POAM scanning."""

    title = "FedRAMP"
    asset_identifier_field = "otherTrackingNumber"
    finding_severity_map = {
        "Low": regscale_models.IssueSeverity.Low,
        "Moderate": regscale_models.IssueSeverity.Moderate,
        "High": regscale_models.IssueSeverity.High,
    }
    poam_id_header = "POAM ID"
    blank_records: int = 0
    blank_threshold: int = 3
    error_records: int = 0
    skipped_records: int = 0
    processed_assets: set[str] = set()  # Track processed assets across all methods

    def __init__(self, plan_id: int):
        """Initialize the FedRAMP integration.

        :param int plan_id: The ID of the security plan
        """
        super().__init__(plan_id=plan_id)
        self.processed_assets = set()  # Reset processed assets on init
        self.workbook: Optional[Workbook] = None  # Type hint for workbook

    def __enter__(self):
        """Context manager entry."""
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        """Context manager exit - cleanup resources."""
        if self.workbook:
            self.workbook.close()

    def fetch_findings(self, *args, **kwargs) -> Iterator[IntegrationFinding]:
        """
        Fetches findings from FedRAMP POAM files.

        :param args: Variable length argument list
        :param kwargs: Arbitrary keyword arguments
        :yield: Iterator[IntegrationFinding]
        """
        file_path = kwargs.get("path")
        if not file_path:
            raise ValueError("File path is required")

        try:
            # Use read_only mode for memory efficiency
            self.workbook = load_workbook(filename=file_path, data_only=True, read_only=True)
        except (FileNotFoundError, InvalidFileException) as e:
            logger.error(f"Failed to load workbook: {e}")
            return

        try:
            poam_sheets = [sheet for sheet in self.workbook.sheetnames if re.search("POA&M Items", sheet)]

            for sheet in poam_sheets:
                yield from self._process_sheet(sheet)
        finally:
            # Ensure workbook is closed
            if self.workbook:
                self.workbook.close()
                self.workbook = None

    def _process_sheet(self, sheet: str) -> Iterator[IntegrationFinding]:
        """Process a single sheet from the POAM workbook."""
        if not self.workbook:
            return
        ws = self.workbook[sheet]
        category = ws["C3"].value or "Low"
        if not ws["C3"].value:
            logger.warning(f"Category is required in cell C3. Defaulting to Low for sheet {sheet}.")

        status = self.determine_status(sheet)
        if status is None:
            logger.warning(f"Unable to determine POA&M status for sheet {sheet}. Skipping import.")
            return

        start_row = self.find_start_row(ws)
        if start_row is None:
            logger.warning(f"No POAM entries found in sheet {sheet}. Skipping.")
            return

        for index, row in enumerate(ws.iter_rows(min_row=start_row, values_only=True), start_row):
            if parsed_category := self.get_row_val(row, "S"):
                category = parsed_category
            try:
                for finding in self.parse_finding(row, status, category, index, sheet):
                    if finding:
                        yield finding

                if self.blank_records >= self.blank_threshold:
                    # logger.warning(f"Too many empty records in sheet {sheet}. Stopping import.")
                    # break
                    pass

            except Exception as e:
                logger.error(f"Error processing row {index} in sheet {sheet}: {str(e)}")
                self.error_records += 1
                continue

    @staticmethod
    def is_poam(finding: IntegrationFinding) -> bool:
        """
        Determine if this finding is a POAM.

        :param IntegrationFinding finding: The finding to check
        :return: True if this is a POAM finding
        :rtype: bool
        """
        return True  # All FedRAMP findings are POAMs

    @staticmethod
    def get_issue_title(finding: IntegrationFinding) -> str:
        """
        Get the title for an issue.

        :param IntegrationFinding finding: The finding
        :return: The issue title
        :rtype: str
        """
        return finding.title[:255]  # Enforce title length limit

    def parse_finding(
        self, row: tuple, status: str, category: str, index: int, sheet: str
    ) -> Iterator[IntegrationFinding]:
        """
        Parse a single row from the POAM spreadsheet into IntegrationFinding objects.
        Creates a separate finding for each asset and CVE combination.

        :param tuple row: The row data
        :param str status: The finding status
        :param str category: The finding category
        :param int index: The row index
        :param str sheet: The sheet name
        :yield: Iterator[IntegrationFinding]
        """
        try:
            poam_id = self.get_row_val(row, "A")
            weakness_name = str(self.get_row_val(row, "C"))

            if not poam_id and weakness_name in [None, "None", ""]:
                self.blank_records += 1
                return

            if not poam_id or not poam_id.upper():
                print(weakness_name, poam_id)
                logger.warning(f"Invalid POAM ID on row {index}, sheet {sheet}. Skipping.")
                return

            if not weakness_name:
                logger.warning(f"Title is required on row {index}, sheet {sheet}. Unable to import")
                return

            # Get and validate plugin ID
            raw_plugin_id = self.get_row_val(row, "F")
            try:
                plugin_id_int = (
                    int(raw_plugin_id)
                    if raw_plugin_id and str(raw_plugin_id).isdigit()
                    else abs(hash(str(raw_plugin_id or ""))) % (10**9)
                )
            except (ValueError, TypeError):
                plugin_id_int = abs(hash(poam_id)) % (10**9)

            # Get asset identifiers
            asset_ids = self.get_row_val(row, "G")
            if not asset_ids:
                logger.warning(f"No asset identifier found on row {index}, sheet {sheet}. Skipping.")
                return

            # Clean asset identifiers
            asset_id_list = [aid.strip() for aid in asset_ids.split(",") if aid.strip()]
            if not asset_id_list:
                logger.warning(f"No valid asset identifiers found on row {index}, sheet {sheet}. Skipping.")
                return

            # Get and validate CVEs
            cves = self.process_cve(self.get_row_val(row, "AD"), index, sheet)
            cve_list = cves.split("\n") if cves else [""]  # Use empty string if no CVEs

            # Create a finding for each asset and CVE combination
            for asset_id in asset_id_list:
                for cve in cve_list:
                    # Create unique plugin ID for each CVE
                    if cve:
                        unique_plugin_id = abs(hash(f"{plugin_id_int}:{cve}")) % (10**9)
                    else:
                        unique_plugin_id = plugin_id_int

                    date_created = date_str(self.get_row_val(row, "K")) or get_current_datetime()
                    due_date = date_str(self.get_row_val(row, "L") if self.get_row_val(row, "L") != "#REF!" else "")
                    severity: IssueSeverity = getattr(IssueSeverity, category.title(), IssueSeverity.NotAssigned)
                    if date_created and not due_date:
                        due_date = issue_due_date(severity, date_created)

                    yield IntegrationFinding(
                        control_labels=[],
                        title=f"{weakness_name[:240]} - {cve}" if cve else weakness_name[:255],
                        category=f"FedRAMP POAM: {category}",
                        description=self.get_row_val(row, "D") or "",
                        severity=severity,
                        status=(
                            regscale_models.IssueStatus.Closed
                            if status.lower() == "closed"
                            else regscale_models.IssueStatus.Open
                        ),
                        asset_identifier=asset_id,
                        external_id=f"{poam_id}:{cve}" if cve else poam_id,
                        date_created=date_created,
                        date_last_updated=date_str(self.get_row_val(row, "O")),
                        due_date=due_date,
                        cve=cve,  # Single CVE per finding
                        plugin_name=self.get_row_val(row, "E") or "",
                        plugin_id=str(unique_plugin_id),
                        observations=self.get_row_val(row, "N") or "",
                        poam_comments=self.empty(self.get_row_val(row, "Z")),
                        remediation=self.empty(self.get_row_val(row, "J")),
                        basis_for_adjustment=self.get_basis_for_adjustment(row),
                        vulnerability_type="FedRAMP",
                        source_report=self.get_row_val(row, "E"),
                        point_of_contact=self.get_row_val(row, "H"),
                        milestone_changes=self.get_row_val(row, "N"),
                        adjusted_risk_rating=self.get_row_val(row, "T"),
                        risk_adjustment=self.get_row_val(row, "U"),
                        operational_requirements=str(self.get_row_val(row, "W")),
                        deviation_rationale=self.get_row_val(row, "X"),
                        poam_id=poam_id,
                    )

        except Exception as e:
            logger.error(f"Error processing row {index} in sheet {sheet}: {str(e)}")
            self.error_records += 1
            return

    # flake8: noqa: C901
    def parse_asset(self, row: tuple) -> Iterator[IntegrationAsset]:
        """
        Parse a single row from the POAM spreadsheet into IntegrationAsset objects.
        Handles multiple comma-separated asset identifiers.

        :param tuple row: The row data from the spreadsheet
        :yield: The parsed IntegrationAsset objects
        :rtype: Iterator[IntegrationAsset]
        """
        try:
            asset_ids = self.get_row_val(row, "G")
            if not asset_ids:
                return

            asset_id_list = [aid.strip() for aid in asset_ids.split(",") if aid.strip()]
            if not asset_id_list:
                return

            def clean_str(val: Optional[str], default: str = "") -> str:
                """Clean and validate string values."""
                if not val:
                    return default
                if not isinstance(val, str):
                    return default

                # Remove problematic patterns
                val = str(val).strip()
                if any(
                    pattern in val.lower()
                    for pattern in [
                        "n/a",
                        "none",
                        "null",
                        "undefined",
                        "planned",
                        "pending",
                        "tbd",
                        "remediation",
                        "deviation",
                        "request",
                        "vulnerability",
                    ]
                ):
                    return default

                # Remove date-like strings
                if re.search(r"\d{4}[-/]\d{1,2}[-/]\d{1,2}", val):
                    return default

                # Remove long descriptions
                if len(val) > 100 or "\n" in val:
                    return default

                return val

            def determine_asset_type(asset_id: str, raw_type: str) -> str:
                """Determine asset type based on asset ID and raw type."""
                if not raw_type or raw_type == "Other":
                    # Check for common patterns in asset ID
                    if any(pattern in asset_id.lower() for pattern in ["docker", "container", "image", "registry"]):
                        return "Container"
                    elif any(pattern in asset_id.lower() for pattern in ["lambda", "function", "azure-function"]):
                        return "Function"
                    elif any(pattern in asset_id.lower() for pattern in ["s3", "bucket", "blob", "storage"]):
                        return "Storage"
                    elif any(pattern in asset_id.lower() for pattern in ["db", "database", "rds", "sql"]):
                        return "Database"
                    elif any(pattern in asset_id.lower() for pattern in ["ec2", "vm", "instance"]):
                        return "Virtual Machine"
                    else:
                        return "Other"
                return raw_type

            for asset_id in asset_id_list:
                # Get raw values and clean them
                raw_values = {
                    "ip": clean_str(self.get_row_val(row, "H")),
                    "type": clean_str(self.get_row_val(row, "I")),
                    "owner": clean_str(self.get_row_val(row, "J"), ScannerVariables.userId),
                    "loc": clean_str(self.get_row_val(row, "K")),
                    "fqdn": clean_str(self.get_row_val(row, "L")),
                    "mac": clean_str(self.get_row_val(row, "M")),
                    "os": clean_str(self.get_row_val(row, "N")),
                    "ver": clean_str(self.get_row_val(row, "O")),
                    "name": clean_str(self.get_row_val(row, "P")),
                    "vendor": clean_str(self.get_row_val(row, "Q")),
                }

                # Determine proper asset type
                asset_type = determine_asset_type(asset_id, raw_values["type"])

                yield IntegrationAsset(
                    name=asset_id,
                    identifier=asset_id,
                    asset_type=asset_type,  # Use determined asset type
                    asset_category=regscale_models.AssetCategory.Hardware,
                    parent_id=self.plan_id,
                    parent_module=regscale_models.SecurityPlan.get_module_string(),
                    status="Active (On Network)",
                    ip_address=raw_values["ip"],
                    fqdn=raw_values["fqdn"],
                    mac_address=raw_values["mac"],
                    location=raw_values["loc"],
                    operating_system=raw_values["os"],
                    software_version=raw_values["ver"],
                    software_name=raw_values["name"],
                    software_vendor=raw_values["vendor"],
                    date_last_updated=get_current_datetime(),
                )

        except Exception as e:
            logger.error(f"Error parsing asset from row: {str(e)}")
            return

    # Helper methods from the original POAM class
    @staticmethod
    def get_row_val(row: tuple, column_name: str) -> Optional[str]:
        """
        Get the value from a specific column in the row.

        :param tuple row: The row data
        :param str column_name: The column name
        :return: The value from the specified column
        :rtype: Optional[str]
        """
        try:
            index = column_index_from_string(column_name) - 1
            return row[index] if index < len(row) else None
        except Exception as e:
            logger.error(f"Error getting value for column {column_name}: {str(e)}")
            return None

    @staticmethod
    def empty(string: Optional[str]) -> Optional[str]:
        """
        Convert empty strings and "None" to None.

        :param Optional[str] string: The input string
        :return: The processed string or None
        :rtype: Optional[str]
        """
        if not isinstance(string, str):
            return None
        if string.lower() in ["none", "n/a"]:
            return None
        return string

    @staticmethod
    def determine_status(sheet: str) -> Optional[str]:
        """
        Determine the status based on sheet name.

        :param str sheet: The sheet name
        :return: The status (Open/Closed) or None
        :rtype: Optional[str]
        """
        if "closed" in sheet.lower():
            return "Closed"
        elif "open" in sheet.lower():
            return "Open"
        return None

    def find_start_row(self, ws) -> Optional[int]:
        """
        Find the first row containing POAM data.

        :param ws: The worksheet
        :return: The row number where POAM entries start
        :rtype: Optional[int]
        """
        for row_index, row in enumerate(ws.iter_rows(min_row=1, max_col=1, values_only=True), 1):
            if row[0] and self.poam_id_header in str(row[0]):
                return row_index + 1
        return None

    def get_basis_for_adjustment(self, row: tuple) -> Optional[str]:
        """
        Get the basis for risk adjustment.

        :param tuple row: The row data
        :return: The basis for adjustment
        :rtype: Optional[str]
        """
        basis_for_adjustment = self.empty(row[23])
        risk_rating = self.get_row_val(row, "S")
        adjusted_risk_rating = self.get_row_val(row, "T")

        if (adjusted_risk_rating != risk_rating) and not basis_for_adjustment:
            return "POAM Import"
        if adjusted_risk_rating == risk_rating:
            return None
        return basis_for_adjustment

    def process_cve(self, cve: Optional[str], index: int, sheet: str) -> Optional[str]:
        """
        Process and validate CVE string. Handles multiple comma-separated CVEs.

        :param Optional[str] cve: The CVE string
        :param int index: The row index
        :param str sheet: The sheet name
        :return: The processed CVE string, multiple CVEs joined by newlines
        :rtype: Optional[str]
        """
        cve = self.empty(cve)
        if not cve:
            return None

        # Split by comma and clean
        cve_list = [c.strip() for c in cve.split(",") if c.strip()]
        if not cve_list:
            return None

        valid_cves = []
        cve_pattern = r"(?:CVE-\d{4}-\d{4,7}|RHSA-\d{4}:\d+|GHSA-[a-z0-9]{4}-[a-z0-9]{4}-[a-z0-9]{4})"

        for single_cve in cve_list:
            # Search for CVE pattern in the string
            cve_match = re.search(cve_pattern, single_cve, re.IGNORECASE)
            if cve_match:
                valid_cves.append(cve_match.group(0).upper())
            else:
                logger.warning(f"Invalid CVE format: {single_cve} on row {index}, sheet {sheet}. Skipping this CVE.")

        # Return newline-separated CVEs or None if no valid CVEs found
        return "\n".join(valid_cves) if valid_cves else None

    def fetch_assets(self, *args, **kwargs) -> Iterator[IntegrationAsset]:
        """
        Fetches assets from FedRAMP POAM files.

        :yield: Iterator[IntegrationAsset]
        """
        file_path = kwargs.get("path")
        if not file_path:
            raise ValueError("File path is required")

        try:
            # Use read_only mode for memory efficiency
            self.workbook = load_workbook(filename=file_path, data_only=True, read_only=True)
        except (FileNotFoundError, InvalidFileException) as e:
            logger.error(f"Failed to load workbook: {e}")
            return

        try:
            poam_sheets = [sheet for sheet in self.workbook.sheetnames if re.search("POA&M Items", sheet)]

            with self._get_lock("processed_assets"):
                for sheet in poam_sheets:
                    ws = self.workbook[sheet]
                    start_row = self.find_start_row(ws)
                    if start_row is None:
                        logger.warning(f"No POAM entries found in sheet {sheet}. Skipping.")
                        continue

                    for row in ws.iter_rows(min_row=start_row, values_only=True):
                        try:
                            for asset in self.parse_asset(row):
                                if asset and asset.identifier not in self.processed_assets:
                                    self.processed_assets.add(asset.identifier)
                                    yield asset
                        except Exception as e:
                            logger.error(f"Error parsing asset from row in sheet {sheet}: {str(e)}")
                            continue
        finally:
            # Ensure workbook is closed
            if self.workbook:
                self.workbook.close()
                self.workbook = None
