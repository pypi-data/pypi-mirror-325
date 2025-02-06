from typing import List
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font, Alignment, PatternFill, Border
from table import Cell

class ExcelStyles:

    @staticmethod
    def set_background_color(worksheet: Worksheet, cells: List[Cell], fill: PatternFill) -> None:
        """Set the background color of the cells.
        
        Parameters
        ----------
        worksheet : Worksheet
            The worksheet of the excel.
        cells : List[Cell]
            The cells to set the background color.
        fill : PatternFill
            The fill of the cells.
        """
        for cell in cells:
            worksheet[cell.get_cell()].fill = fill

    @staticmethod
    def get_background_color(worksheet: Worksheet, cell: Cell) -> str:
        """Returns the background color of the cell.

        Parameters
        ----------
        worksheet : Worksheet
            The worksheet of the excel.
        cell : Cell
            The cell to get the background color.
        """
        return worksheet[cell.get_cell()].fill.start_color.index

    @staticmethod
    def set_font(worksheet: Worksheet, 
                cells: List[Cell], 
                font: Font) -> None:
        """Set the font of the cells.

        Parameters
        ----------
        worksheet : Worksheet
            The worksheet of the excel.
        cells : List[Cell]
            The cells to set the font.
        font : Font
            The font of the cells.
        """
        for cell in cells:
            worksheet[cell.get_cell()].font = font

    @staticmethod
    def get_font(worksheet: Worksheet, cell: Cell) -> Font:
        """Returns the font of the cell.

        Parameters
        ----------
        worksheet : Worksheet
            The worksheet of the excel.
        cell : Cell
            The cell to get the font.
        """
        return worksheet[cell.get_cell()].font
    
    @staticmethod
    def set_alignment(worksheet: Worksheet, cells: List[Cell], alignment: Alignment) -> None:
        """Set the alignment of the cells.

        Parameters
        ----------
        worksheet : Worksheet
            The worksheet of the excel.
        cells : List[Cell]
            The cell to set the alignment.
        aligment : Alignment
            The alignment of the cells.
        """ 
        for cell in cells:
            worksheet[cell.get_cell()].alignment = alignment

    @staticmethod
    def get_alignment(worksheet: Worksheet, cell: Cell) -> Alignment:
        """Returns the alignment of the cell.

        Parameters
        ----------
        worksheet : Worksheet
            The worksheet of the excel.
        cell : Cell
            The cell to get the alignment.
        """
        alignment = worksheet[cell.get_cell()].alignment
        return alignment

    @staticmethod
    def set_border(worksheet: Worksheet, cells: list[Cell], border: Border) -> None:
        """Set the border of the cells.

        Parameters
        ----------
        worksheet : Worksheet
            The worksheet of the excel.
        cells : list[Cell]
            The cells to set the border.
        border : Border
            The border to set.
        """
        for cell in cells:
            worksheet[cell.get_cell()].border = border

    @staticmethod
    def get_border(worksheet: Worksheet, cell: Cell) -> Border:
        """Returns the border of the cell.

        Parameters
        ----------
        worksheet : Worksheet
            The worksheet of the excel.
        cell : Cell
            The cell to get the border.
        """
        return worksheet[cell.get_cell()].border

    @staticmethod
    def set_height(worksheet: Worksheet, cells: List[Cell], height: int) -> None:
        """Set the height of the cells.

        Parameters
        ----------
        worksheet : Worksheet
            The worksheet of the excel.
        cells : List[Cell]
            The cell to set the height.
        height : int
            The height of the cell.
        """ 
        for cell in cells:
            worksheet.row_dimensions[cell.get_row()].height = height

    @staticmethod
    def get_height(worksheet: Worksheet, cell: Cell) -> int:
        """Returns the height of the cell.

        Parameters
        ----------
        worksheet : Worksheet
            The worksheet of the excel.
        cell : Cell
            The cell to get the height.
        """
        return worksheet.row_dimensions[cell.get_row()].height
    
    @staticmethod
    def set_width(worksheet: Worksheet, cells: List[Cell], width: int) -> None:
        """Set the width of the cells.
        
        Parameters
        ----------
        worksheet : Worksheet
            The worksheet of the excel.
        cells : List[Cell]
            The cell to set the width.
        width : int
            The width of the cell.        
        """
        for cell in cells:
            worksheet.column_dimensions[cell.get_column()].width = width

    @staticmethod
    def get_width(worksheet: Worksheet, cell: Cell) -> int:
        """Returns the width of the cell.

        Parameters
        ----------
        worksheet : Worksheet
            The worksheet of the excel.
        cell : Cell
            The cell to get the width.
        """
        return worksheet.column_dimensions[cell.get_column()].width