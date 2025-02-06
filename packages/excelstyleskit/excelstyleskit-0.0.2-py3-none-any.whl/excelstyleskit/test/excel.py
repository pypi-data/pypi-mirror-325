import unittest
import os
from openpyxl.workbook import Workbook
from openpyxl.styles import Font, Alignment
from excel import Excel
from table import Table

LIST_CELLS = ['A1', 'B1', 'C1', 'D1', 'A2', 'B2', 'C2', 'D2', 'A3', 'B3', 'C3', 'D3', 'A4', 'B4', 'C4', 'D4']
DEFAULT_COLOR_BG = '00000000'
COLOR_BG = 'FF00A933'
RANDOM_COLOR_ONE = '00666699'
RANDOM_COLOR_TWO = '00993366'
RANDOM_COLOR_THREE = '00FF99CC'
DEFAULT_FONT = Font(
    name='Calibri',
    size=11, 
    bold=False, 
    italic=False, 
    vertAlign=None, 
    underline=None, 
    strike=None
)
FONT_ONE = Font(
    name='Arial',
    size=16, 
    bold=True, 
    italic=True, 
    vertAlign='superscript', 
    underline='double', 
    strike=True, 
    color=RANDOM_COLOR_ONE
)
FONT_TWO = Font(
    name='Time New Roman',
    size=12
)
DEFAULT_ALIGNMENT = Alignment(
    horizontal=None, 
    vertical=None, 
    text_rotation=0, 
    wrap_text=None, 
    shrink_to_fit=None, 
    indent=0
)
ALIGNMENT_ONE = Alignment(
    horizontal='center', 
    vertical='top', 
    text_rotation=5,
    wrap_text=True, 
    shrink_to_fit=True, 
    indent=1.3
)
ALIGNMENT_TWO = Alignment(
    horizontal='left', 
    vertical='bottom', 
)
DEFAULT_HEIGHT_ROW = None
HEIGHT_ROW_ONE = 28
HEIGHT_ROW_TWO = 16
DEFAULT_WIDTH = 13
WIDHT = 30

class TestExcel(unittest.TestCase):
    FIRST_COLUMN = 'A'
    FIRST_ROW = 1
    LAST_COLUMN = 'D'
    LAST_ROW = 4
    DEFAULT_COLOR_BG = '00000000'
    COLOR_BG = 'FF00A933'
    RANDOM_COLOR_ONE = '00666699'
    RANDOM_COLOR_TWO = '00993366'
    RANDOM_COLOR_THREE = '00FF99CC'

    def get_excel(self) -> Excel:
        filepath = os.getcwd() + '/test/test.xlsx'
        if os.path.exists(filepath): os.remove(filepath)
        woorkbook = Workbook()
        woorksheet = woorkbook.active
        woorksheet.cell(row=1, column=1).value = 'Header A1'
        woorksheet.cell(row=1, column=2).value = 'Header B2'
        woorksheet.cell(row=1, column=3).value = 'Header C3'
        woorksheet.cell(row=1, column=4).value = 'Header D4'
        woorksheet.cell(row=2, column=1).value = 'Content A2'
        woorksheet.cell(row=2, column=2).value = 'Content B2'
        woorksheet.cell(row=2, column=3).value = 'Content C2'
        woorksheet.cell(row=2, column=4).value = 'Content D2'
        woorksheet.cell(row=3, column=1).value = 'Content A3'
        woorksheet.cell(row=3, column=2).value = 'Content B3'
        woorksheet.cell(row=3, column=3).value = 'Content C3'
        woorksheet.cell(row=3, column=4).value = 'Content D3'
        woorksheet.cell(row=4, column=1).value = 'Content A4'
        woorksheet.cell(row=4, column=2).value = 'Content B4'
        woorksheet.cell(row=4, column=3).value = 'Content C4'
        woorksheet.cell(row=4, column=4).value = 'Content D4'
        woorkbook.save(filepath)
        return Excel(filepath)
    
    def delete_file(self) -> None:
        filepath = os.getcwd() + '/test/test.xlsx'
        if os.path.exists(filepath): os.remove(filepath)
    
    def create_table(self, excel: Excel) -> Excel:
        excel.set_table(
            self.FIRST_COLUMN, 
            self.FIRST_ROW, 
            self.LAST_COLUMN, 
            self.LAST_ROW
        )
        excel.add_row_header(1)
        return excel

    def test_get_workbook(self):
        excel = self.get_excel()
        workbook = excel.get_workbook()
        self.assertEqual(type(workbook), Workbook)
        self.delete_file()
    
    def test_get_table(self):
        excel = self.get_excel()
        excel = self.create_table(excel)
        table = excel.get_table()
        self.assertEqual(
            Table.get_cells_str(table.get_cells()), 
            LIST_CELLS
        )
        self.delete_file()

    def test_get_background_color(self):
        excel = self.get_excel()
        excel = self.create_table(excel)
        color_content = excel.get_background_color_cell(
            excel.get_table().get_cells_content()[0]
        )
        color_header = excel.get_background_color_cell(
            excel.get_table().get_cells_header()[0]
        )
        self.assertEqual(color_content, DEFAULT_COLOR_BG)
        self.assertEqual(color_header, DEFAULT_COLOR_BG)
        self.delete_file()

    def test_set_background_color(self):
        excel = self.get_excel()
        excel = self.create_table(excel)
        excel.set_background_color_header(
            start_color=COLOR_BG,
            end_color=COLOR_BG
        )
        color_header = excel.get_background_color_cell(
            excel.get_table().get_cells_header()[0]
        )
        color_content = excel.get_background_color_cell(
            excel.get_table().get_cells_content()[0]
        )
        self.assertEqual(color_header, COLOR_BG)
        self.assertEqual(color_content, DEFAULT_COLOR_BG)
        self.delete_file()

    def test_get_font(self):
        excel = self.get_excel()
        excel = self.create_table(excel)
        header_font = excel.get_font_cell(
            excel.get_table().get_cells_header()[0]
        )
        content_font = excel.get_font_cell(
            excel.get_table().get_cells_content()[0]
        )
        self.assertEqual(header_font.name, DEFAULT_FONT.name)
        self.assertEqual(header_font.size, DEFAULT_FONT.size)
        self.assertEqual(header_font.bold, DEFAULT_FONT.bold)
        self.assertEqual(header_font.italic, DEFAULT_FONT.italic)
        self.assertEqual(header_font.vertAlign, DEFAULT_FONT.vertAlign)
        self.assertEqual(header_font.underline, DEFAULT_FONT.underline)
        self.assertEqual(header_font.strike, DEFAULT_FONT.strike)
        self.assertEqual(content_font.name, DEFAULT_FONT.name)
        self.assertEqual(content_font.size, DEFAULT_FONT.size)
        self.assertEqual(content_font.bold, DEFAULT_FONT.bold)
        self.assertEqual(content_font.italic, DEFAULT_FONT.italic)
        self.assertEqual(content_font.vertAlign, DEFAULT_FONT.vertAlign)
        self.assertEqual(content_font.underline, DEFAULT_FONT.underline)
        self.assertEqual(content_font.strike, DEFAULT_FONT.strike)
        self.delete_file()

    def test_set_font(self):
        excel = self.get_excel()
        excel = self.create_table(excel)
        excel.set_font_header(
            name=FONT_ONE.name,
            size=FONT_ONE.size,
            bold=FONT_ONE.bold,
            italic=FONT_ONE.italic,
            vertAlign=FONT_ONE.vertAlign,
            underline=FONT_ONE.underline,
            strike=FONT_ONE.strike,
            color=FONT_ONE.color
        )
        excel.set_font_content(
            name=FONT_TWO.name,
            size=FONT_TWO.size
        )
        header_font = excel.get_font_cell(
            excel.get_table().get_cells_header()[0]
        )
        content_font = excel.get_font_cell(
            excel.get_table().get_cells_content()[0]
        )
        self.assertEqual(header_font.name, FONT_ONE.name)
        self.assertEqual(header_font.size, FONT_ONE.size)
        self.assertEqual(header_font.bold, FONT_ONE.bold)
        self.assertEqual(header_font.italic, FONT_ONE.italic)
        self.assertEqual(header_font.vertAlign, FONT_ONE.vertAlign)
        self.assertEqual(header_font.underline, FONT_ONE.underline)
        self.assertEqual(header_font.strike, FONT_ONE.strike)
        self.assertEqual(header_font.color.value, FONT_ONE.color.value)
        self.assertEqual(content_font.name, FONT_TWO.name)
        self.assertEqual(content_font.size, FONT_TWO.size)
        self.delete_file()

    def test_get_alignment(self):
        excel = self.get_excel()
        excel = self.create_table(excel)
        header_alignment = excel.get_alignment_cell(
            excel.get_table().get_cells_header()[0]
        )
        content_alignment = excel.get_alignment_cell(
            excel.get_table().get_cells_content()[0]
        )
        self.assertEqual(header_alignment.horizontal, DEFAULT_ALIGNMENT.horizontal)
        self.assertEqual(header_alignment.vertical, DEFAULT_ALIGNMENT.vertical)
        self.assertEqual(header_alignment.text_rotation, DEFAULT_ALIGNMENT.text_rotation)
        self.assertEqual(header_alignment.wrap_text, DEFAULT_ALIGNMENT.wrap_text)
        self.assertEqual(header_alignment.shrink_to_fit, DEFAULT_ALIGNMENT.shrink_to_fit)
        self.assertEqual(header_alignment.indent, DEFAULT_ALIGNMENT.indent)
        self.assertEqual(content_alignment.horizontal, DEFAULT_ALIGNMENT.horizontal)
        self.assertEqual(content_alignment.vertical, DEFAULT_ALIGNMENT.vertical)
        self.assertEqual(content_alignment.text_rotation, DEFAULT_ALIGNMENT.text_rotation)
        self.assertEqual(content_alignment.wrap_text, DEFAULT_ALIGNMENT.wrap_text)
        self.assertEqual(content_alignment.shrink_to_fit, DEFAULT_ALIGNMENT.shrink_to_fit)
        self.assertEqual(content_alignment.indent, DEFAULT_ALIGNMENT.indent)
        self.delete_file()

    def test_set_alignment(self):
        excel = self.get_excel()
        excel = self.create_table(excel)
        excel.set_alignment_header(
            horizontal=ALIGNMENT_ONE.horizontal,
            vertical=ALIGNMENT_ONE.vertical,
            text_rotation=ALIGNMENT_ONE.text_rotation,
            wrap_text=ALIGNMENT_ONE.wrap_text,
            shrink_to_fit=ALIGNMENT_ONE.shrink_to_fit,
            indent=ALIGNMENT_ONE.indent
        )
        excel.set_alignment_content(
            horizontal=ALIGNMENT_TWO.horizontal,
            vertical=ALIGNMENT_TWO.vertical
        )
        header_alignment = excel.get_alignment_cell(
            excel.get_table().get_cells_header()[0]
        )
        content_alignment = excel.get_alignment_cell(
            excel.get_table().get_cells_content()[0]
        )
        self.assertEqual(header_alignment.horizontal, ALIGNMENT_ONE.horizontal)
        self.assertEqual(header_alignment.vertical, ALIGNMENT_ONE.vertical)
        self.assertEqual(header_alignment.text_rotation, ALIGNMENT_ONE.text_rotation)
        self.assertEqual(header_alignment.wrap_text, ALIGNMENT_ONE.wrap_text)
        self.assertEqual(header_alignment.shrink_to_fit, ALIGNMENT_ONE.shrink_to_fit)
        self.assertEqual(header_alignment.indent, ALIGNMENT_ONE.indent)
        self.assertEqual(content_alignment.horizontal, ALIGNMENT_TWO.horizontal)
        self.assertEqual(content_alignment.vertical, ALIGNMENT_TWO.vertical)
        self.delete_file()

    def test_get_border(self):
        excel = self.get_excel()
        excel = self.create_table(excel)
        header_border = excel.get_border_cell(
            excel.get_table().get_cells_header()[0]
        )
        content_border = excel.get_border_cell(
            excel.get_table().get_cells_content()[0]
        )
        self.assertEqual(header_border.left.color, None)
        self.assertEqual(header_border.left.style, None)
        self.assertEqual(header_border.top.color, None)
        self.assertEqual(header_border.top.style, None)
        self.assertEqual(header_border.right.color, None)
        self.assertEqual(header_border.right.style, None)
        self.assertEqual(header_border.bottom.color, None)
        self.assertEqual(header_border.bottom.style, None)
        self.assertEqual(content_border.left.color, None)
        self.assertEqual(content_border.left.style, None)
        self.assertEqual(content_border.top.color, None)
        self.assertEqual(content_border.top.style, None)
        self.assertEqual(content_border.right.color, None)
        self.assertEqual(content_border.right.style, None)
        self.assertEqual(content_border.bottom.color, None)
        self.assertEqual(content_border.bottom.style, None)
        self.delete_file()

    def test_set_border(self):
        excel = self.get_excel()
        excel = self.create_table(excel)
        excel.set_border_header(
            color=RANDOM_COLOR_TWO
        )
        excel.set_border_content(
            color=RANDOM_COLOR_THREE
        )
        header_border = excel.get_border_cell(
            excel.get_table().get_cells_header()[0]
        )
        content_border = excel.get_border_cell(
            excel.get_table().get_cells_content()[0]
        )
        self.assertEqual(header_border.left.color.value, RANDOM_COLOR_TWO)
        self.assertEqual(header_border.top.color.value, RANDOM_COLOR_TWO)
        self.assertEqual(header_border.right.color.value, RANDOM_COLOR_TWO)
        self.assertEqual(header_border.bottom.color.value, RANDOM_COLOR_TWO)
        self.assertEqual(content_border.left.color.value, RANDOM_COLOR_THREE)
        self.assertEqual(content_border.top.color.value, RANDOM_COLOR_THREE)
        self.assertEqual(content_border.right.color.value, RANDOM_COLOR_THREE)
        self.assertEqual(content_border.bottom.color.value, RANDOM_COLOR_THREE)
        self.delete_file()

    def test_get_height_row(self):
        excel = self.get_excel()
        excel = self.create_table(excel)
        height_row = excel.get_height_row_cell(
            excel.get_table().get_cells_header()[0]
        )
        content_row = excel.get_height_row_cell(
            excel.get_table().get_cells_content()[0]
        )
        self.assertEqual(height_row, DEFAULT_HEIGHT_ROW)
        self.assertEqual(content_row, DEFAULT_HEIGHT_ROW)
        self.delete_file()

    def test_set_height_row(self):
        excel = self.get_excel()
        excel = self.create_table(excel)
        excel.set_height_row_header(HEIGHT_ROW_ONE)
        excel.set_height_row_content(HEIGHT_ROW_TWO)
        height_row = excel.get_height_row_cell(
            excel.get_table().get_cells_header()[0]
        )
        content_row = excel.get_height_row_cell(
            excel.get_table().get_cells_content()[0]
        )
        self.assertEqual(height_row, HEIGHT_ROW_ONE)
        self.assertEqual(content_row, HEIGHT_ROW_TWO)
        self.delete_file()

    def test_get_width(self):
        excel = self.get_excel()
        excel = self.create_table(excel)
        width = excel.get_width_column_cell(
            excel.get_table().get_cells_header()[0]
        )
        self.assertEqual(width, DEFAULT_WIDTH)
        self.delete_file()

    def test_set_width(self):
        excel = self.get_excel()
        excel = self.create_table(excel)
        excel.set_width_column_table(WIDHT)
        width = excel.get_width_column_cell(
            excel.get_table().get_cells_header()[0]
        )
        self.assertEqual(width, WIDHT)
        self.delete_file()

if __name__ == '__main__':
    unittest.main()