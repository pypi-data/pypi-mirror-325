from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles.colors import Color
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from table import Table, Cell
from excel import ExcelStyles
from utils import is_valid_filepath

class Excel:
    _filepath: str
    _workbook: Workbook
    _worksheet: Worksheet
    _table: Table

    def __init__(self, filepath: str, sheetname: (str | None)=None) -> None:
        if not is_valid_filepath(filepath): 
            raise FileNotFoundError(f'The file {filepath} does not exist.')
        self._filepath = filepath
        self._workbook = load_workbook(self._filepath)
        if sheetname: self._worksheet = self._workbook[sheetname]
        else: self._worksheet = self._workbook.active

    def set_table(self, first_column: str, first_row: int, last_column: str, last_row: int) -> None:
        """Set the table of the excel.

        Parameters
        ----------
        first_column : str
            The first column of the table.
        first_row : int
            The first row of the table.
        last_column : str
            The last column of the table.
        last_row : int
            The last row of the table.
        """
        self._table = Table(
            first_column, 
            first_row, 
            last_column, 
            last_row
        )

    def add_row_header(self, row: int) -> None:
        """Add the all cells of the row as the header of the table. 
        
        Parameters
        ----------
        row : int
            The row to add as header.
        """
        self._table.add_row_header(row)

    def get_workbook(self) -> Workbook:
        """Returns the workbook of the excel. """
        return self._workbook
    
    def get_worksheet(self) -> Worksheet:
        """Returns the worksheet of the excel. """
        return self._worksheet
    
    def get_table(self) -> Table:
        """Returns the object table of the excel. """
        return self._table
    
    def save_work(self) -> None:
        """Save the changes in the excel. """
        self._workbook.save(self._filepath)
    
    def set_background_color_header(self, start_color: str='FFFFFF', end_color: str='FFFFFF') -> None:
        """Set the background color of the header of the table.
        
        Parameters
        ----------
        start_color : str, optional
            The start color in hexadecimal format.
        end_color : str, optional
            The end color in hexadecimal format.
        """
        start_color = Color(rgb=start_color)
        end_color = Color(rgb=end_color)
        fill = PatternFill(
            patternType='solid', 
            start_color=start_color, 
            end_color=end_color
        )
        ExcelStyles.set_background_color(
            self._worksheet, 
            self._table.get_cells_header(), 
            fill
        )

    def set_background_color_content(self, start_color: str='FFFFFF', end_color: str='FFFFFF') -> None:
        """Set the background color of the content of the table.
        
        Parameters
        ----------
        start_color : str, optional
            The start color in hexadecimal format.
        end_color : str, optional
            The end color in hexadecimal format.
        """
        start_color = Color(rgb=start_color)
        end_color = Color(rgb=end_color)
        fill = PatternFill(
            patternType='solid', 
            start_color=start_color, 
            end_color=end_color
        )
        ExcelStyles.set_background_color(
            self._worksheet,
            self._table.get_cells_content(), 
            fill
        )

    def get_background_color_cell(self, cell: Cell) -> str:
        """Returns the background color of the cell. 
        
        Parameters
        ----------
        cell : Cell
            The cell to get the background color.
        """
        return ExcelStyles.get_background_color(self._worksheet, cell)
    
    def set_font_header(self,
                name: str='Arial', 
                size: int=10, 
                bold: bool=False, 
                italic: bool=False, 
                vertAlign: (str | None)=None, 
                underline: (str | None)=None, 
                strike: bool=False, 
                color: str='FFFF0000') -> None:
        """Set the font of the header of the table.

        Parameters
        ----------
        worksheet : Workbook
            The worksheet of the excel.
        cells : List[Cell]
            The cells to set the font.
        name : str
            The name of the font.
        size : int
            The size of the font.
        bold : bool
            The bold of the font. 
        italic : bool
            The italic of the font.
        vertAlign : str
            Value must be one of {‘superscript’, ‘baseline’, ‘subscript’} to the font.
        underline : str
            Value must be one of {‘single’, ‘double’, ‘doubleAccounting’, ‘singleAccounting’} to the font.
        strike : bool
            The strike of the font.
        color: str
            The color of the font.
        """
        font_settigns = Font(
                name=name,
                size=size,
                bold=bold,
                italic=italic,
                vertAlign=vertAlign,
                underline=underline,
                strike=strike,
                color=color
        )
        ExcelStyles.set_font(
            self._worksheet,
            self._table.get_cells_header(),
            font_settigns
        )

    def set_font_content(self,
                name: str='Arial', 
                size: int=10, 
                bold: bool=False, 
                italic: bool=False, 
                vertAlign: (str | None)=None, 
                underline: (str | None)=None, 
                strike: bool=False, 
                color: str='FFFF0000') -> None:
        """Set the font of the content of the table.

        Parameters
        ----------
        worksheet : Workbook
            The worksheet of the excel.
        cells : List[Cell]
            The cells to set the font.
        name : str
            The name of the font.
        size : int
            The size of the font.
        bold : bool
            The bold of the font. 
        italic : bool
            The italic of the font.
        vertAlign : str
            Value must be one of {‘superscript’, ‘baseline’, ‘subscript’} to the font.
        underline : str
            Value must be one of {‘single’, ‘double’, ‘doubleAccounting’, ‘singleAccounting’} to the font.
        strike : bool
            The strike of the font.
        color: str
            The color of the font.
        """
        font_settigns = Font(
                name=name,
                size=size,
                bold=bold,
                italic=italic,
                vertAlign=vertAlign,
                underline=underline,
                strike=strike,
                color=color
        )
        ExcelStyles.set_font(
            self._worksheet,
            self._table.get_cells_content(),
            font_settigns
        )

    def get_font_cell(self, cell: Cell) -> Font:
        """Returns the font settings of the cell.
        
        Parameters
        ----------
        cell : Cell
            The cell to get the font.
        """
        return ExcelStyles.get_font(self._worksheet, cell)
    
    def set_alignment_header(self, 
                horizontal: str='left', 
                vertical: str='center', 
                text_rotation: int=0, 
                wrap_text: bool=False, 
                shrink_to_fit: bool=False, 
                indent: int=0) -> None:
        """Set the alignment of the header of the table.
        
        Parameters
        ----------
        horizontal : str
            Value must be one of {‘left’, ‘center’, ‘right’, ‘fill’, ‘justify’, ‘centerContinuous’, ‘distributed’} to the horizontal.
        vertical : str
            Value must be one of {‘top’, ‘center’, ‘bottom’, ‘justify’, ‘distributed’} to the vertical.
        text_rotation : int
            The rotation of the text.
        wrap_text : bool
            The wrap text.
        shrink_to_fit : bool
            The shrink.
        indent : int
            The indent.
        """
        aligment = Alignment(
            horizontal=horizontal,
            vertical=vertical, 
            text_rotation=text_rotation, 
            wrap_text=wrap_text, 
            shrink_to_fit=shrink_to_fit, 
            indent=indent
        )
        ExcelStyles.set_alignment(
            self._worksheet, 
            self._table.get_cells_header(), 
            aligment
        )

    def set_alignment_content(self, 
                horizontal: str='left', 
                vertical: str='center', 
                text_rotation: int=0, 
                wrap_text: bool=False, 
                shrink: bool=False, 
                indent: int=0) -> None:
        """Set the alignment of the content of the table.
        
        Parameters
        ----------
        horizontal : str
            Value must be one of {‘left’, ‘center’, ‘right’, ‘fill’, ‘justify’, ‘centerContinuous’, ‘distributed’} to the horizontal.
        vertical : str
            Value must be one of {‘top’, ‘center’, ‘bottom’, ‘justify’, ‘distributed’} to the vertical.
        text_rotation : int
            The rotation of the text.
        wrap_text : bool
            The wrap text.
        shrink : bool
            The shrink.
        indent : int
            The indent.
        """
        aligment = Alignment(
            horizontal=horizontal,
            vertical=vertical, 
            text_rotation=text_rotation, 
            wrap_text=wrap_text, 
            shrink_to_fit=shrink, 
            indent=indent
        )
        ExcelStyles.set_alignment(
            self._worksheet,
            self._table.get_cells_content(), 
            aligment
        )

    def get_alignment_cell(self, cell: Cell) -> Alignment:
        """Returns the alignment of the cell.
        
        Parameters
        ----------
        cell : Cell
            The cell to get the alignment.
        """
        return ExcelStyles.get_alignment(self._worksheet, cell)
    
    def set_border_header(self, color: str="000000", style: str="thin") -> None:
        """Set the border of the header of the table.

        Parameters
        ----------
        color : str, optional
            The color of the border.
        style : str, optional
            Value must be one of {'dashDot','dashDotDot', 'dashed','dotted',
                                'double','hair', 'medium', 'mediumDashDot', 'mediumDashDotDot',
                                'mediumDashed', 'slantDashDot', 'thick', 'thin'} to the style.
        """
        border = Border(
            left=Side(style=style, color=color),
            right=Side(style=style, color=color),
            top=Side(style=style, color=color),
            bottom=Side(style=style, color=color),
        )
        ExcelStyles.set_border(
            self._worksheet,
            self._table.get_cells_header(),
            border
        )

    def set_border_content(self, color: str="000000", style: str="thin") -> None:
        """Set the border of the content of the table.

        Parameters
        ----------
        color : str, optional
            The color of the border.
        style : str, optional
            Value must be one of {'dashDot','dashDotDot', 'dashed','dotted',
                                'double','hair', 'medium', 'mediumDashDot', 'mediumDashDotDot',
                                'mediumDashed', 'slantDashDot', 'thick', 'thin'} to the style.
        """
        border = Border(
            left=Side(style=style, color=color),
            right=Side(style=style, color=color),
            top=Side(style=style, color=color),
            bottom=Side(style=style, color=color),
        )
        ExcelStyles.set_border(
            self._worksheet,
            self._table.get_cells_content(),
            border
        )

    def get_border_cell(self, cell: Cell) -> Border:
        """Returns the border of the cell.
        
        Parameters
        ----------
        cell : Cell
            The cell to get the border.
        """
        return ExcelStyles.get_border(self._worksheet, cell)
    
    def set_height_row_header(self, height: int) -> None:
        """Set the height of the row header of the table.
        
        Parameters
        ----------
        height : int
            The height of the row header.
        """
        ExcelStyles.set_height(
            self._worksheet, 
            self._table.get_cells_header(), 
            height
        )

    def set_height_row_content(self, height: int) -> None:
        """Set the height of the row content of the table.
        
        Parameters
        ----------
        height : int
            The height of the row content.
        """
        ExcelStyles.set_height(
            self._worksheet, 
            self._table.get_cells_content(), 
            height
        )

    def get_height_row_cell(self, cell: Cell) -> int:
        """Returns the height of the row of a cell.
        
        Parameters
        ----------
        cell : Cell
            The cell to get the height.
        """
        return ExcelStyles.get_height(self._worksheet, cell)
    
    def set_width_column_table(self, width: int) -> None:
        """Set the width of the column of the table.
        
        Parameters
        ----------
        width : int
            The width of the column.
        """
        ExcelStyles.set_width(
            self._worksheet, 
            self._table.get_cells(), 
            width
        )

    def get_width_column_cell(self, cell: Cell) -> int:
        """Returns the width of the column of a cell.
        
        Parameters
        ----------
        cell : Cell
            The cell to get the width.
        """
        return ExcelStyles.get_width(self._worksheet, cell)