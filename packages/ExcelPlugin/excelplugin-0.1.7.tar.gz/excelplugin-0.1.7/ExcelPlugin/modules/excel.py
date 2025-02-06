# ExcelPlugin/modules/excel.py

import pandas as pd
import numpy as np
from openpyxl import Workbook
from typing import Union, List, Tuple
from ..utils.header import Header
from ..utils.funstions import get_column_letter


from ..utils.formatter import format_money, format_percent, format_date

# Пример значений по умолчанию для заголовка
DEFAULT_HEADER = {
    'start_color': 'A0A0A0',  # Цвет фона в формате HEX
    'end_color': 'A0A0A0',  # Цвет фона в формате HEX
    'text_color': 'FFFFFF',  # Цвет текста
    'font_size': 12,    # Размер шрифта
    'font_name': 'Roboto Condensed',  # Название шрифта
    'bold': True,       # Жирный шрифт
    'text_align': 'center',  # Горизонтальное выравнивание
    'valign_align': 'center',  # Вертикальное выравнивание
    'row_height': 45,    # Высота строки
    'border_style': 'thin',
    'border_color': 'FFFFFF'
}


class Excel:
    def __init__(
        self,
        data: pd.DataFrame,
        workbook: Workbook,
        worksheet: str
    ):
        self._wb = workbook
        self._ws = self._wb.create_sheet(title=worksheet)
        self._data = data
        self._header = None
        self._styles = None


    def add_header(self, rows: Union[List[List[Tuple[Tuple[int, int], str]]], None]=None, format=DEFAULT_HEADER):
        """_summary_

        Args:
            rows (Union[List[List[Tuple[Tuple[int, int], str]]], None], optional): _description_. Defaults to None.
            format (_type_, optional): _description_. Defaults to DEFAULT_HEADER.

        Returns:
            _type_: _description_
        """
        headers_row = [((_[0], _[0],), _[1]) for _ in  enumerate(self._data.columns, 1)]
        if rows:
            pass
        else:
            rows=[]
        rows.append(headers_row)
        self.header = Header(self._ws, rows, format)
        return self

    def _write_data(self):
        start_row = len(self.header.rows) + 1 if self.header else 1
        for row_idx, row in enumerate(self._data.values, start_row):
            for col_idx, value in enumerate(row, 1):
                # Универсальная проверка на все виды NA (None, np.nan, pd.NA)
                if pd.isna(value):
                    value = None

                # Обработка pandas IntegerArray (Int64)
                elif isinstance(value, pd._libs.missing.NAType):
                    value = None

                # Преобразование numpy-типов
                elif isinstance(value, np.generic):
                    value = value.item()

                # Преобразование pandas Timestamp
                elif isinstance(value, pd.Timestamp):
                    value = value.to_pydatetime()

                self._ws.cell(row=row_idx, column=col_idx, value=value)
        self._set_columns_width(24)
        return self

    def _set_columns_width(self, width: float) -> None:
        for col_idx in range(1, len(self._data.columns) + 1):
            col_letter = get_column_letter(col_idx)
            self._ws.column_dimensions[col_letter].width = width

    def set_money_column(self, money_columns: Union[List[str], None]=None):
        # Приведение к денежному формату
        [[format_money(self._data, get_column_letter(index+1), self._ws), col] for index, col in enumerate(self._data.columns) if col in money_columns]
        return self

    def set_percent_column(self, percent_columns: Union[List[str], None]=None):
        # Приведение к процентному формату
        [[format_percent(self._data, get_column_letter(index+1), self._ws), col] for index, col in enumerate(self._data.columns) if col in percent_columns]
        return self

    def set_date_column(self, date_columns: Union[List[str], None]=None):
        # Приведение к процентному формату
        [[format_date(self._data, get_column_letter(index+1), self._ws), col] for index, col in enumerate(self._data.columns) if col in date_columns]
        return self
