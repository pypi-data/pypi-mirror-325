# ExcelPlugin/utils/header.py

from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from typing import List, Tuple, Union, Dict, Optional

class Header:
    def __init__(
        self,
        worksheet: Worksheet,
        rows: List[List[Tuple[Tuple[int, int], str]]],
        format: dict
    ):
        self.ws = worksheet
        self.rows = rows
        self.format = format
        self._apply_merges()
        self._apply_format()

    def _apply_merges(self):
        for row_idx, row_data in enumerate(self.rows, 1):
            for cell_info in row_data:
                (start_col, end_col), value = cell_info
                self.ws.merge_cells(
                    start_row=row_idx,
                    start_column=start_col,
                    end_row=row_idx,
                    end_column=end_col
                )
                self.ws.cell(row=row_idx, column=start_col, value=value)

    def _apply_format(self):
        """Применяет стили к заголовкам."""
        # Создаем стили один раз для всех ячеек
        side = Side(
            border_style=self.format['border_style'],
            color=self.format['border_color']
        )
        border = Border(left=side, right=side, top=side, bottom=side)
        fill = PatternFill(
            start_color=self.format['start_color'],
            end_color=self.format['end_color'],
            fill_type='solid'
        )
        font = Font(
            name=self.format['font_name'],
            size=self.format['font_size'],
            bold=self.format['bold'],
            color = self.format['text_color']
        )
        alignment = Alignment(
            horizontal=self.format['text_align'],
            vertical=self.format['valign_align']
        )

        for row_idx, row_data in enumerate(self.rows, 1):
            self.ws.row_dimensions[row_idx].height = self.format['row_height']

            for cell_info in row_data:
                (start_col, end_col), _ = cell_info
                # Применяем стиль только к первой ячейке объединенного диапазона
                for col in range(start_col, end_col + 1):
                    cell = self.ws.cell(row=row_idx, column=col)

                    cell.fill = fill
                    cell.font = font
                    cell.alignment = alignment
                    cell.border = border
