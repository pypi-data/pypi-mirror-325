# ExcelPlugin/modules/excel.py

import pandas as pd
from openpyxl import Workbook
from typing import Union, List, Tuple
from ..utils.header import Header
from ..utils.funstions import get_column_letter


from ..utils.formatter import format_money, format_percent, format_date

# Пример значений по умолчанию для заголовка
DEFAULT_HEADER = {
    'start_color': 'A0A0A0',  # Цвет фона в формате HEX
    'end_color': 'A0A0A0',  # Цвет фона в формате HEX
    'text_color': 'FFFFFF',  # Цвет текста
    'font_size': 12,    # Размер шрифта
    'font_name': 'Roboto Condensed',  # Название шрифта
    'bold': True,       # Жирный шрифт
    'text_align': 'center',  # Горизонтальное выравнивание
    'valign_align': 'center',  # Вертикальное выравнивание
    'row_height': 45,    # Высота строки
    'border_style': 'thin',
    'border_color': 'FFFFFF'
}


class Excel:
    def __init__(
        self,
        data: pd.DataFrame,
        workbook: Workbook,
        worksheet: str
    ):
        self.wb = workbook
        self.ws = self.wb.create_sheet(title=worksheet)
        self.data = data
        self.header = None
        self.styles = None

    def add_header(self, rows: Union[List[List[Tuple[Tuple[int, int], str]]], None]=None, format=DEFAULT_HEADER):
        """Добавляет заголовок с объединенными ячейками."""
        headers_row = [((_[0], _[0],), _[1]) for _ in  enumerate(self.data.columns, 1)]
        if rows:
            pass
        else:
            rows=[]
        rows.append(headers_row)
        self.header = Header(self.ws, rows, format)
        return self

    def _write_data(self):
        start_row = len(self.header.rows) + 1 if self.header else 1
        for row_idx, row in enumerate(self.data.values, start_row):
            for col_idx, value in enumerate(row, 1):
                self.ws.cell(row=row_idx, column=col_idx, value=value)

        self._set_columns_width(24)
        return self

    def _set_columns_width(self, width: float) -> None:
        for col_idx in range(1, len(self.data.columns) + 1):
            col_letter = get_column_letter(col_idx)
            self.ws.column_dimensions[col_letter].width = width

    def set_money_column(self, money_columns: Union[List[str], None]=None):
        # Приведение к денежному формату
        [[format_money(self.data, get_column_letter(index+1), self.ws), col] for index, col in enumerate(self.data.columns) if col in money_columns]
        return self

    def set_percent_column(self, percent_columns: Union[List[str], None]=None):
        # Приведение к процентному формату
        [[format_percent(self.data, get_column_letter(index+1), self.ws), col] for index, col in enumerate(self.data.columns) if col in percent_columns]
        return self

    def set_date_column(self, date_columns: Union[List[str], None]=None):
        # Приведение к процентному формату
        [[format_date(self.data, get_column_letter(index+1), self.ws), col] for index, col in enumerate(self.data.columns) if col in date_columns]
        return self
